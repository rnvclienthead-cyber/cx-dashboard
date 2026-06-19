"""
Импортирует xlsx-отчёт 'Рейтинг товаров' из личного кабинета ЯМ в базу данных.

Использование:
  python3 import_ym_ratings_xlsx.py <путь_к_файлу.xlsx> [YYYY-MM-DD]

  Если дата не указана — берётся из имени файла (ищет YYYY-MM-DD) или сегодняшняя.

Пример:
  python3 import_ym_ratings_xlsx.py business_rating_report_203987989_11-06-2026.xlsx 2026-06-11
"""

import io
import os
import re
import sys
from datetime import date, datetime

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

DATABASE_URL = os.getenv(
    "DATABASE_URL_LOCAL",
    "postgresql://db_user:RDB_r6o_BA0qSlVVGjb_2026@127.0.0.1:5432/cx_dashboard",
)
engine = create_engine(DATABASE_URL, pool_pre_ping=True)


def parse_date_from_filename(path):
    """Пытается извлечь дату из имени файла."""
    name = os.path.basename(path)
    m = re.search(r'(\d{2}-\d{2}-\d{4})', name)
    if m:
        try:
            return datetime.strptime(m.group(1), "%d-%m-%Y").date()
        except ValueError:
            pass
    m = re.search(r'(\d{4}-\d{2}-\d{2})', name)
    if m:
        try:
            return datetime.strptime(m.group(1), "%Y-%m-%d").date()
        except ValueError:
            pass
    return None


def parse_ratings_xlsx(path, report_date):
    wb = openpyxl.load_workbook(path, data_only=True)

    if "Рейтинг товаров" not in wb.sheetnames:
        print("ОШИБКА: лист 'Рейтинг товаров' не найден")
        return []

    ws = wb["Рейтинг товаров"]
    rows = []
    current_name   = None
    current_rating = None
    current_delta  = None

    for row_vals in ws.iter_rows(min_row=3, values_only=True):
        if len(row_vals) < 11:
            continue
        name, _link, sku, rating, delta, total, s5, s4, s3, s2, s1 = row_vals[:11]

        if name:
            current_name   = str(name).strip()
            current_rating = float(rating) if rating not in (None, '') else None
            current_delta  = float(delta)  if delta  not in (None, '') else 0.0

        if sku and str(sku).strip():
            rows.append({
                "report_date":      report_date,
                "supplier_article": str(sku).strip(),
                "product_name":     current_name,
                "group_rating":     current_rating,
                "weekly_delta":     current_delta,
                "review_count":     int(total) if total else 0,
                "stars_5":          int(s5)    if s5    else 0,
                "stars_4":          int(s4)    if s4    else 0,
                "stars_3":          int(s3)    if s3    else 0,
                "stars_2":          int(s2)    if s2    else 0,
                "stars_1":          int(s1)    if s1    else 0,
            })

    return rows


def ingest(rows):
    if not rows:
        return
    with engine.begin() as conn:
        for r in rows:
            conn.execute(text("""
                INSERT INTO ym_ratings_report
                    (report_date, supplier_article, product_name, group_rating,
                     weekly_delta, review_count, stars_5, stars_4, stars_3, stars_2, stars_1)
                VALUES
                    (:report_date, :supplier_article, :product_name, :group_rating,
                     :weekly_delta, :review_count, :stars_5, :stars_4, :stars_3, :stars_2, :stars_1)
                ON CONFLICT (report_date, supplier_article) DO UPDATE SET
                    product_name = EXCLUDED.product_name,
                    group_rating = EXCLUDED.group_rating,
                    weekly_delta = EXCLUDED.weekly_delta,
                    review_count = EXCLUDED.review_count,
                    stars_5 = EXCLUDED.stars_5, stars_4 = EXCLUDED.stars_4,
                    stars_3 = EXCLUDED.stars_3, stars_2 = EXCLUDED.stars_2,
                    stars_1 = EXCLUDED.stars_1,
                    synced_at = NOW()
            """), r)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"ОШИБКА: файл не найден: {path}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        try:
            report_date = datetime.strptime(sys.argv[2], "%Y-%m-%d").date()
        except ValueError:
            print("ОШИБКА: неверный формат даты, используй YYYY-MM-DD")
            sys.exit(1)
    else:
        report_date = parse_date_from_filename(path) or date.today()

    print(f"Импорт: {path}")
    print(f"Дата отчёта: {report_date}")

    rows = parse_ratings_xlsx(path, report_date)
    if not rows:
        print("Нет данных для импорта")
        sys.exit(1)

    print(f"Найдено SKU: {len(rows)}")
    ingest(rows)
    print(f"Готово! Сохранено {len(rows)} строк на дату {report_date}")

    # Показываем первые 5 строк для проверки
    print("\nПервые 5 записей:")
    for r in rows[:5]:
        print(f"  {r['supplier_article']:20s} | рейтинг={r['group_rating']} | "
              f"всего={r['review_count']} | ★5={r['stars_5']} ★4={r['stars_4']} ★3={r['stars_3']}")
