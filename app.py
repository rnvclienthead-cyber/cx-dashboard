import streamlit as st
import asyncio
import aiohttp
import gspread
import json
import re
import pandas as pd
import plotly.express as px
from datetime import datetime, timedelta
from google.oauth2.service_account import Credentials
import io
import zipfile
import urllib.request
import base64
import streamlit.components.v1 as componentsаа
import time
import os
import xlsxwriter
import requests
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from sqlalchemy import create_engine, text

st.set_page_config(page_title="CX AI Enterprise", layout="wide")

st.markdown("""
    <style>
    /* НАВИГАЦИЯ: Стилизуем кнопки под плоские пункты меню */
    div[data-testid="stSidebar"] button {
        justify-content: flex-start;
        text-align: left;
        padding: 8px 12px;
        border-radius: 6px;
        border: none;
        background-color: transparent;
        color: #475569;
        font-weight: 500;
        transition: all 0.2s ease;
        margin-bottom: 2px;
    }
    div[data-testid="stSidebar"] button:hover { background-color: #f1f5f9; color: #0f172a; }
    div[data-testid="stSidebar"] button[kind="primary"] {
        background-color: #eff6ff !important;
        color: #2563eb !important;
        font-weight: 600 !important;
    }

    [data-testid="stDataFrame"] { font-size: 11px !important; }
    .detail-card { border: 1px solid #ddd; padding: 15px; border-radius: 8px; margin-bottom: 15px; background-color: #fcfcfc; }
    .media-row { display: flex; flex-wrap: wrap; gap: 15px; margin-bottom: 10px; }
    .media-row a { background: transparent !important; padding: 0 !important; margin: 0 !important; border: none !important; display: inline-flex; }
    .photo-zoom { width: 140px !important; height: 140px !important; object-fit: cover !important; border-radius: 8px !important; transition: transform 0.3s ease; cursor: pointer; }
    .photo-zoom:hover { transform: scale(4); z-index: 9999; position: relative; border-radius: 0px !important; box-shadow: 0 20px 50px rgba(0,0,0,0.8) !important; }
    .video-link-btn { display: inline-block; padding: 8px 14px; background-color: #2563eb; color: white !important; border-radius: 6px; text-decoration: none; font-weight: bold; font-size: 13px; }
    .video-link-btn:hover { background-color: #1d4ed8; }
    .ai-tags-box { background-color: #f0fdf4; padding: 10px 14px; border-radius: 6px; font-size: 14px; color: #166534; margin-bottom: 15px; font-weight: 500; border-left: 4px solid #22c55e; }
    </style>
""", unsafe_allow_html=True)

# --- ИСПРАВЛЕННАЯ СТАБИЛЬНАЯ НАВИГАЦИЯ С MATERIAL ICONS ---
tech_menu = {
    "Синхронизатор": ("Робот-Синхронизатор", ":material/cloud_sync:"),
    "ИИ Тегирование": ("ИИ Тегирование", ":material/biotech:"),
    "Модерация": ("Модерация", ":material/fact_check:"),
    "Обучение ИИ": ("Обучение ИИ", ":material/psychology:"),
    "Системный Журнал": ("Системный Журнал", ":material/receipt_long:")
}

ops_menu = {
    "Отчет производства": ("Отчет производства", ":material/insights:"),
    "Уровень PPM": ("Уровень PPM", ":material/warning:"),
    "Рейтинг товаров": ("Рейтинг товаров", ":material/star:")
}

if 'active_tab' not in st.session_state:
    st.session_state.active_tab = "Робот-Синхронизатор"

st.sidebar.markdown("### :material/settings: Технический блок")
for key, (page_name, icon) in tech_menu.items():
    is_active = st.session_state.active_tab == page_name
    # ИСПОЛЬЗУЕМ параметр icon= вместо добавления в текст
    if st.sidebar.button(key, icon=icon, key=f"nav_{key}", type="primary" if is_active else "secondary", use_container_width=True):
        st.session_state.active_tab = page_name
        st.rerun()

st.sidebar.markdown("### :material/trending_up: Операционный блок")
for key, (page_name, icon) in ops_menu.items():
    is_active = st.session_state.active_tab == page_name
    # ИСПОЛЬЗУЕМ параметр icon= вместо добавления в текст
    if st.sidebar.button(key, icon=icon, key=f"nav_{key}", type="primary" if is_active else "secondary", use_container_width=True):
        st.session_state.active_tab = page_name
        st.rerun()

page = st.session_state.active_tab

if st.session_state.get('current_tab') != page:
    st.session_state.current_tab = page
    st.session_state.matrix_key = int(time.time())
    st.session_state.show_detail_trigger = None
    st.session_state.last_click_id = None

try:
    YANDEX_API_KEY = st.secrets["YANDEX_API_KEY"]
    FOLDER_ID = st.secrets["FOLDER_ID"]
    XAI_API_KEY = st.secrets["XAI_API_KEY"]
    SPREADSHEET_ID_MAIN = st.secrets["SPREADSHEET_ID_MAIN"]
    SPREADSHEET_ID_INVOICES = st.secrets["SPREADSHEET_ID_INVOICES"]
    GOOGLE_CREDS = dict(st.secrets["gcp_service_account"])
    DB_URL = st.secrets.get("DB_URL") 
    engine = create_engine(DB_URL) if DB_URL else None
except Exception as e:
    st.error(f":material/warning: Ошибка в Secrets: {e}")
    st.stop()

CATEGORIES = {
    1: "Некомплект: Фурнитура", 2: "Некомплект: Несущие детали", 3: "Состояние упаковки",
    4: "Производственный дефект", 5: "Механические повреждения", 6: "Инструкция и сборка",
    7: "Хлипкость / Устойчивость", 8: "Пересорт / Ошибка склада", 9: "Качество материалов",
    10: "Габариты и Размер", 11: "Несоответствие описанию", 12: "Субъективное 'Не подошло'",
    13: "Следы использования / Б/У"
}

# Словари для автоматического заполнения Рекламационных Актов
CLAIM_MAPPING = {
    "defect_descr": {
        "Некомплект": ("Не хватает комплектующих изделий", "缺少配件"),
        "Дефект": ("Повреждения деталей", "零部件损坏 / 部件有损坏"),
        "Сварка": ("Некачественная сварка", "焊接质量不良"),
    },
    "causes": {
        "Производство": ("Нарушение при производственном процессе", "生产过程异常"),
        "Сборка": ("Отклонение в процессе сборки (комплектации)", "装配/配套过程偏差"),
    }
}

# Маппинг категорий вашего приложения (ID) в текст Рекламации
CAT_TO_CLAIM_TEXT = {
    1: ("Не хватает комплектующих (Фурнитура)", "缺少配件（五金）"),
    2: ("Не хватает деталей", "缺少零件"),
    4: ("Производственный дефект / Сварка", "生产过程异常 / 焊接质量不良"),
    5: ("Механические повреждения", "零部件损坏"),
}

# Правила объединения категорий и перевода для рекламации
CLAIM_CATEGORIES_LOGIC = {
    "Shortage": {
        "ids": [1, 2],
        "ru": "Не хватает комплектующих изделий",
        "cn": "缺少配件",
        "cause_ru": "Отклонение в процессе сборки (комплектации)",
        "cause_cn": "装配/配套过程偏差"
    },
    "Damage": {
        "ids": [4, 5],
        "ru": "Повреждения деталей",
        "cn": "零部件损坏 / 部件有损坏",
        "cause_ru": "Нарушение при производственном процессе",
        "cause_cn": "生产过程异常"
    },
    "Flimsy": {
        "ids": [12], # Предположим, 12 это Хлипкость/устойчивость
        "ru": "Хлипкость",
        "cn": "不牢固",
        "cause_ru": "Конструктивный недостаток / Нарушение процесса",
        "cause_cn": "设计缺陷 / 生产过程异常"
    }
}

# ==========================================
# БАЗОВЫЕ ФУНКЦИИ 
# ==========================================
def safe_read(file_obj):
    bytes_data = file_obj.getvalue()
    name = file_obj.name.lower()
    if name.endswith('.xlsx') or name.endswith('.xls'):
        try:
            eng = 'calamine' if name.endswith('.xlsx') else 'xlrd'
            return pd.read_excel(io.BytesIO(bytes_data), engine=eng)
        except Exception:
            try: 
                return pd.read_excel(io.BytesIO(bytes_data), engine='openpyxl')
            except Exception:
                try: 
                    return pd.read_html(io.BytesIO(bytes_data))[0]
                except Exception: 
                    pass
                    
    for enc in ['utf-8-sig', 'utf-8', 'windows-1251', 'utf-16']:
        for sep in [';', '\t', ',']:
            try:
                text_data = bytes_data.decode(enc)
                df = pd.read_csv(io.StringIO(text_data), sep=sep, engine='python', on_bad_lines='skip')
                if len(df.columns) > 1: return df
            except Exception: 
                continue
    st.error(f"⚠️ Не удалось прочитать файл {file_obj.name}.")
    return pd.DataFrame()

@st.cache_resource
def get_gspread_client():
    scopes = ['https://www.googleapis.com/auth/spreadsheets']
    return gspread.authorize(Credentials.from_service_account_info(GOOGLE_CREDS, scopes=scopes))

def get_memory_records():
    try: 
        return get_gspread_client().open_by_key(SPREADSHEET_ID_MAIN).worksheet("Память_ИИ").get_all_records()
    except: 
        return []

def add_system_log(action, status, details=""):
    if not engine:
        return
    
    sql = text("""
        INSERT INTO system_logs (action, status, details) 
        VALUES (:action, :status, :details)
    """)
    
    try:
        with engine.begin() as conn:
            conn.execute(sql, {
                "action": action, 
                "status": status, 
                "details": details
            })
    except Exception as e:
        # Если база упала, выводим ошибку хотя бы в консоль сервера
        print(f"🚨 КРИТИЧЕСКАЯ ОШИБКА SQL-ЛОГИРОВАНИЯ: {e}")

def update_db_row(srid, updates_dict):
    if not engine or not updates_dict: return False
    
    # Сохраняем в основную базу
    set_clauses = [f"{k} = :{k}" for k in updates_dict.keys()]
    sql = text(f"UPDATE wb_claims SET {', '.join(set_clauses)} WHERE srid = :srid")
    
    try:
        with engine.begin() as conn:
            conn.execute(sql, {**updates_dict, "srid": srid})
            
            # САМООБУЧЕНИЕ: Если была ручная корректировка, сохраняем её как опыт
            # ВАЖНО: Исключаем системные статусы, чтобы ИИ не выучил их как теги
            if "correction" in updates_dict and updates_dict["correction"] and updates_dict["correction"] not in ["Подтверждено", "Нет тегов"]:
                # Достаем текст отзыва для этой заявки
                claim_text = conn.execute(text("SELECT user_comment FROM wb_claims WHERE srid = :srid"), {"srid": srid}).scalar()
                if claim_text:
                    conn.execute(text("""
                        INSERT INTO ai_knowledge_base (content, tags, source) 
                        VALUES (:txt, :tgs, 'manual')
                        ON CONFLICT (content) DO UPDATE SET tags = EXCLUDED.tags
                    """), {"txt": claim_text, "tgs": updates_dict["correction"]})
        return True
    except Exception as e:
        return False

def create_images_zip(urls):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for i, url in enumerate(urls):
            try:
                req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=3) as response:
                    img_data = response.read()
                    ext = ".jpg"
                    if ".png" in url.lower(): ext = ".png"
                    elif ".jpeg" in url.lower(): ext = ".jpeg"
                    zip_file.writestr(f"photo_{i+1}{ext}", img_data)
            except Exception: pass
    return zip_buffer.getvalue()

def get_media_for_srid(srid):
    """Ленивая загрузка фото и видео только для одной конкретной заявки"""
    if not engine or not srid: return "", ""
    query = text("SELECT photos, video_paths FROM wb_claims WHERE srid = :srid")
    try:
        with engine.connect() as conn:
            result = conn.execute(query, {"srid": str(srid)}).fetchone()
            if result:
                return str(result[0] or ''), str(result[1] or '')
    except: pass
    return "", ""

# ==========================================
# НОВАЯ ФУНКЦИЯ ГЕНЕРАЦИИ (ПО ШАБЛОНУ)
# ==========================================
# Умная функция для записи текста (обходит блокировки объединенных ячеек)
def safe_write(sheet, coord, value):
    """Умная запись с поддержкой объединенных ячеек"""
    try:
        sheet[coord].value = value
    except AttributeError:
        for merged_range in sheet.merged_cells.ranges:
            if coord in merged_range:
                top_left = str(merged_range).split(':')[0]
                sheet[top_left].value = value
                break

def generate_claim_from_template(data, chart_fig=None, template_path="template_ra.xlsx"):
    try:
        wb = openpyxl.load_workbook(template_path)
    except FileNotFoundError:
        return b""
        
    sheet = wb.active
    
    # 1. ШАПКА (Смещено на 1 строку вверх: 2 -> 1)
    safe_write(sheet, 'G1', f"Рекламационный акт № {data['number']}")
    safe_write(sheet, 'O1', f"质量投诉报告 № {data['number']}")
    
    # 2. РУССКИЙ БЛОК (Смещено на 1 строку вверх: 4->3, 5->4 и т.д.)
    safe_write(sheet, 'C3', data['date'])          
    safe_write(sheet, 'G3', data['supplier'])      
    safe_write(sheet, 'C4', data['period'])        
    safe_write(sheet, 'G4', data['invoice'])       
    safe_write(sheet, 'C5', "Возвраты с маркетплейсов")
    safe_write(sheet, 'C6', data['sku'])           
    safe_write(sheet, 'C7', data['name'])          
    safe_write(sheet, 'G7', f"{data['defects']} ({data['ppm_pct']} %)") 
    safe_write(sheet, 'C8', data['desc_ru'])       
    safe_write(sheet, 'G8', data['cause_ru'])      
    
    # 3. КИТАЙСКИЙ БЛОК (Смещено на 1 строку вверх)
    safe_write(sheet, 'L3', data['date'])          
    safe_write(sheet, 'P3', data['supplier'])      
    safe_write(sheet, 'L4', data['period'])        
    safe_write(sheet, 'P4', data['invoice'])       
    safe_write(sheet, 'L5', "电商平台退货")
    safe_write(sheet, 'L6', data['sku'])           
    safe_write(sheet, 'L7', data.get('name_cn', '产品'))          
    safe_write(sheet, 'P7', f"{data['defects']} ({data['ppm_pct']} %)") 
    safe_write(sheet, 'L8', data.get('desc_cn', '')) 
    safe_write(sheet, 'P8', data.get('cause_cn', '')) 
    
    # 4. ГРАФИКИ (Теперь в 10-й строке)
    if chart_fig:
        try:
            # Генерация картинки (уменьшена на 10% для точности)
            img_bytes = chart_fig.to_image(format="png", width=700, height=320)
            
            # RU
            img_ru = OpenpyxlImage(io.BytesIO(img_bytes))
            img_ru.width, img_ru.height = 350, 160
            sheet.add_image(img_ru, 'B10') 
            
            # CN
            img_cn = OpenpyxlImage(io.BytesIO(img_bytes))
            img_cn.width, img_cn.height = 350, 160
            sheet.add_image(img_cn, 'K10') 
        except Exception:
            pass

    # 5. ФОТОГРАФИИ (Начинаем с 25 строки, чтобы не было дыр)
    photo_row = 25 
    for cat_name, photos in data.get('photo_groups', {}).items():
        if photos:
            # Заголовок категории
            safe_write(sheet, f'B{photo_row}', f"Категория дефекта / 缺陷类别: {cat_name}")
            sheet.cell(row=photo_row, column=2).font = openpyxl.styles.Font(bold=True)
            photo_row += 1
            
            col_off = 2 # Колонка B
            for url in photos[:3]:
                try:
                    # Усиленная загрузка фото
                    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
                    resp = requests.get(url, headers=headers, timeout=10)
                    if resp.status_code == 200:
                        img = OpenpyxlImage(io.BytesIO(resp.content))
                        img.width, img.height = 130, 130
                        col_letter = openpyxl.utils.get_column_letter(col_off)
                        sheet.add_image(img, f'{col_letter}{photo_row}')
                        col_off += 3
                except Exception:
                    continue
            photo_row += 8 # Шаг для следующей группы

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ==========================================
# ИИ ДВИЖОК
# ==========================================
def parse_ai_response(text):
    try:
        clean_text = re.sub(r'```json|```', '', str(text)).strip()
        # Защита от мусора до/после JSON
        start = min([i for i in [clean_text.find('['), clean_text.find('{')] if i >= 0] or [0])
        end = max(clean_text.rfind(']'), clean_text.rfind('}')) + 1
        if start >= 0 and end > 0:
            clean_text = clean_text[start:end]
            
        parsed = json.loads(clean_text)
        if isinstance(parsed, dict): return parsed.get('results', [])
        elif isinstance(parsed, list): return parsed
        else: return [{"error": f"Неожиданный формат: {type(parsed)}"}]
    except Exception as e:
        return [{"error": f"Сбой формата JSON: {str(e)} | Ответ ИИ: {text}"}]
        
def find_similar_examples_sql(target_text, engine, top_n=15):
    """
    Ищет похожие примеры в SQL базе знаний, используя умный алгоритм Триграмм (pg_trgm).
    Понимает опечатки и находит смысл.
    """
    if not target_text or not engine:
        return "Опыта пока нет."

    # SQL запрос: ищем тексты, похожесть которых больше 10% (0.1), сортируем от самых похожих к менее
    sql = text("""
        SELECT content, tags, similarity(content, :target_text) as sml
        FROM ai_knowledge_base
        WHERE similarity(content, :target_text) > 0.05
        ORDER BY sml DESC
        LIMIT :top_n
    """)

    try:
        with engine.connect() as conn:
            results = conn.execute(sql, {"target_text": target_text, "top_n": top_n}).fetchall()

        if not results:
            return "Прямых совпадений в опыте не найдено."

        # Собираем результаты в текст для ИИ
        best_matches = [f"Текст: {row[0]} -> Тег: {row[1]}" for row in results]
        return "\n".join(best_matches)

    except Exception as e:
        print(f"Ошибка поиска в базе знаний: {e}")
        return "Ошибка доступа к опыту."

async def fetch_ai_tags(session, batch, memory_string, model="yandex"):
    content_lines = []
    for i in batch:
        content_lines.append(f"ID {i['id']}: {i['text']}")
    content = "\n".join(content_lines)

    system_prompt = f"""Ты эксперт контроля качества. 
    Категории (ID: Название): {json.dumps(CATEGORIES, ensure_ascii=False)}
    ПРАВИЛО 12: Если клиент хвалит, но есть мелкий дефект (рейтинг 4-5) - СТРОГО Категория 12.
    ВОТ ПРИМЕРЫ ПОХОЖИХ СИТУАЦИЙ ИЗ БАЗЫ:
    {memory_string}
    
    ИНСТРУКЦИЯ: Верни ТОЛЬКО массив category_ids (цифры подходящих категорий). Никакого текста!
    ОТВЕТЬ СТРОГО JSON: {{"results": [{{"id": "...", "category_ids": [1, 5]}}]}}"""

    if "yandex" in model:
        url = 'https://llm.api.cloud.yandex.net/foundationModels/v1/completion'
        headers = {"Authorization": f"Api-Key {YANDEX_API_KEY}", "x-folder-id": FOLDER_ID}
        yandex_model_name = "yandexgpt-lite" if model == "yandex-lite" else "yandexgpt"
        payload = {
            "modelUri": f"gpt://{FOLDER_ID}/{yandex_model_name}/latest",
            "completionOptions": {"temperature": 0.1, "maxTokens": 2000},
            "messages": [{"role": "system", "text": system_prompt}, {"role": "user", "text": content}]
        }
        try:
            async with session.post(url, headers=headers, json=payload, timeout=45) as resp:
                if resp.status == 200:
                    res = await resp.json()
                    return parse_ai_response(res['result']['alternatives'][0]['message']['text'])
                else: 
                    return [{"error": f"Ошибка Яндекса ({resp.status}): {await resp.text()}"}]
        except Exception as e: 
            return [{"error": f"Системная ошибка Яндекса: {str(e)}"}]

    elif model == "grok":
        url = "https://api.x.ai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {XAI_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "grok-beta",
            "temperature": 0.1,
            "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": content}]
        }
        try:
            async with session.post(url, headers=headers, json=payload, timeout=45) as resp:
                if resp.status == 200:
                    res = await resp.json()
                    return parse_ai_response(res['choices'][0]['message']['content'])
                else: 
                    return [{"error": f"Ошибка Grok ({resp.status}): {await resp.text()}"}]
        except Exception as e: 
            return [{"error": f"Системная ошибка Grok: {str(e)}"}]
    return []

async def fetch_ai_crosscheck(session, batch, engine, model="grok"):
    content = "\n".join([f"ID {i['id']}: {i['text']}" for i in batch])
    combined_target_text = " ".join([i['text'] for i in batch])
    relevant_memory = find_similar_examples_sql(combined_target_text, engine)

    system_prompt = f"""Ты — строгий аудитор. Твоя задача — проверить правильность тегов, которые поставила другая нейросеть.
ДОСТУПНЫЕ КАТЕГОРИИ: {json.dumps(CATEGORIES, ensure_ascii=False)}
--- ОПЫТ ---
{relevant_memory}
ПРАВИЛО: Если текущие теги противоречат опыту, исправь их, записав тег в `correction`.
ОТВЕТЬ СТРОГО JSON: {{"results": [{{"id": "...", "audit_status": "Согласен", "audit_comment": "", "correction": ""}}]}}"""

    if "yandex" in model:
        url = 'https://llm.api.cloud.yandex.net/foundationModels/v1/completion'
        headers = {"Authorization": f"Api-Key {YANDEX_API_KEY}", "x-folder-id": FOLDER_ID}
        yandex_model_name = "yandexgpt-lite" if model == "yandex-lite" else "yandexgpt"
        payload = {
            "modelUri": f"gpt://{FOLDER_ID}/{yandex_model_name}/latest",
            "completionOptions": {"temperature": 0.1, "maxTokens": 2000},
            "messages": [{"role": "system", "text": system_prompt}, {"role": "user", "text": content}]
        }
        try:
            async with session.post(url, headers=headers, json=payload, timeout=45) as resp:
                if resp.status == 200:
                    res = await resp.json()
                    return parse_ai_response(res['result']['alternatives'][0]['message']['text'])
        except Exception as e: print(f"Ошибка Yandex Аудит: {e}")
        
    elif model == "grok":
        url = "https://api.x.ai/v1/chat/completions"
        headers = {"Authorization": f"Bearer {XAI_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "grok-beta", "temperature": 0.1,
            "messages": [{"role": "system", "content": system_prompt}, {"role": "user", "content": content}]
        }
        try:
            async with session.post(url, headers=headers, json=payload, timeout=45) as resp:
                if resp.status == 200:
                    res = await resp.json()
                    return parse_ai_response(res['choices'][0]['message']['content'])
        except Exception as e: print(f"Ошибка Grok Аудит: {e}")

    return []

async def run_ai_batch_processing(df_to_tag, model_choice, memory_string="", mode="tagging"):
    results = []
    async with aiohttp.ClientSession() as session:
        batch = []
        for idx, row in df_to_tag.iterrows():
            # Формируем данные для батча в зависимости от режима
            if mode == "tagging":
                batch.append({
                    "id": str(row['srid']), 
                    "text": f"Артикул: {row.get('supplier_article','')}. Текст: {row.get('user_comment','')}"
                })
            else:
                current_tags = [str(c) for c in range(1, 14) if row.get(f'cat_{c}') == True]
                tags_str = ", ".join(current_tags) if current_tags else "Нет тегов"
                batch.append({
                    "id": str(row['srid']), 
                    "text": f"Текст: {row.get('user_comment','')}. Текущие теги (ID): {tags_str}"
                })
            
            # Обработка накопленной пачки
            if len(batch) >= 10:
                if mode == "tagging": 
                    res = await fetch_ai_tags(session, batch, memory_string, model_choice)
                else: 
                    # Для аудита оставляем передачу engine, так как там логика не менялась
                    res = await fetch_ai_crosscheck(session, batch, engine, model_choice) 
                if res: 
                    results.extend(res)
                batch = []
                
        # Обработка остатков в батче
        if batch:
            if mode == "tagging": 
                res = await fetch_ai_tags(session, batch, memory_string, model_choice)
            else: 
                res = await fetch_ai_crosscheck(session, batch, engine)
            if res: 
                results.extend(res)
            
    return results

# ==========================================
# ИНТЕРФЕЙС И НАВИГАЦИЯ
# ==========================================

@st.cache_data(ttl=120) 
def load_cached_hybrid_data():
    # ОСТАВЛЯЕМ ТОЛЬКО ЭТУ ВЕРСИЮ (с JOIN инвойсов)
    query = """
        SELECT 
            v."SRID", v."Дата и время оформления заявки на возврат", v."Дата заказа", v."Дата и время получения заказа покупателем",
            v."Артикул продавца", v."Комментарий покупателя", v."Решение по возврату покупателю", v."Статус товара",
            v."1", v."2", v."3", v."4", v."5", v."6", v."7", v."8", v."9", v."10", v."11", v."12", v."13",
            v."Корректировка", v."Номер поставки",
            COALESCE(inv.invoice_num, 'Не указан') AS "Инвойс"
        FROM view_cx_dashboard v
        LEFT JOIN wb_invoices inv 
            ON TRIM(v."Номер поставки") = inv.supply_id 
            AND TRIM(v."Артикул продавца") = inv.supplier_article
    """
    df_temp = pd.DataFrame()
    try:
        df_temp = pd.read_sql(query, engine)
    except Exception as e:
        print(f"Ошибка SQL: {e}")
        return df_temp
        
    if df_temp.empty: 
        return df_temp

    # Обработка дат
    date_col = next((c for c in df_temp.columns if 'оформления заявки' in str(c).lower()), None)
    if date_col:
        df_temp['Дата_ДТ'] = pd.to_datetime(df_temp[date_col], errors='coerce')
        df_temp[date_col] = df_temp['Дата_ДТ'].dt.strftime('%d.%m.%Y %H:%M').fillna('Не указана')
    else:
        df_temp['Дата_ДТ'] = pd.NaT
        
    # Фильтрация только одобренных
    valid_statuses = ['одобрено', '2', '2.0', 'да', 'true']
    df_temp = df_temp[
        df_temp['Решение по возврату покупателю'].astype(str).str.strip().str.lower().isin(valid_statuses) |
        df_temp['Статус товара'].astype(str).str.strip().str.lower().isin(valid_statuses)
    ]
    
    df_temp['Артикул продавца'] = df_temp['Артикул продавца'].astype(str).str.strip()
    df_temp = df_temp[~df_temp['Артикул продавца'].str.lower().isin(['nan', 'none', '', 'null'])]
    df_temp['Номер поставки_ОРИГИНАЛ'] = df_temp['Номер поставки'].astype(str).replace(['nan', 'None', ''], 'Не указан').str.strip()
        
    return df_temp

    # --- Обработка дат ---
    date_col = next((c for c in df_temp.columns if 'оформления заявки' in str(c).lower()), None)
    if date_col:
        df_temp['Дата_ДТ'] = pd.to_datetime(df_temp[date_col], errors='coerce')
        df_temp[date_col] = df_temp['Дата_ДТ'].dt.strftime('%d.%m.%Y %H:%M').fillna('Не указана')
    else:
        df_temp['Дата_ДТ'] = pd.NaT
        
    # --- НАДЕЖНАЯ ПИТОНОВСКАЯ ФИЛЬТРАЦИЯ (Срезает любой скрытый мусор) ---
    valid_statuses = ['одобрено', '2', '2.0', 'да', 'true']
    df_temp = df_temp[
        df_temp['Решение по возврату покупателю'].astype(str).str.strip().str.lower().isin(valid_statuses) |
        df_temp['Статус товара'].astype(str).str.strip().str.lower().isin(valid_statuses)
    ]
    
    # --- Очистка артикулов ---
    df_temp['Артикул продавца'] = df_temp['Артикул продавца'].astype(str).str.strip()
    df_temp = df_temp[~df_temp['Артикул продавца'].str.lower().isin(['nan', 'none', '', 'null'])]

    if 'Номер поставки' not in df_temp.columns:
        df_temp['Номер поставки'] = 'Не указан'
    
    df_temp['Номер поставки_ОРИГИНАЛ'] = df_temp['Номер поставки'].astype(str).replace(['nan', 'None', ''], 'Не указан').str.strip()
        
    # БЛОК СКАЧИВАНИЯ ИЗ GOOGLE SHEETS УДАЛЕН ЗА НЕНАДОБНОСТЬЮ
    # Теперь колонка "Инвойс" приходит сразу готовой из базы!
        
    return df_temp

@st.cache_data(ttl=120)
def load_cached_orders():
    # ОПТИМИЗАЦИЯ: Фильтрация отмен и группировка по месяцам перенесена в SQL!
    query = """
        SELECT 
            TRIM(supplier_article) AS "Артикул продавца", 
            DATE_TRUNC('month', dt) AS "Месяц_ДТ",
            COUNT(*) AS "Чистые_заказы"
        FROM wb_orders
        WHERE cancel_dt IS NULL
          AND supplier_article IS NOT NULL 
          AND TRIM(supplier_article) != ''
        GROUP BY 1, 2
    """
    try: 
        final_orders = pd.read_sql(query, engine)
        
        if final_orders.empty:
            return pd.DataFrame()

        # SQL отдает дату, просто убедимся, что Pandas понимает её как datetime
        final_orders['Месяц_ДТ'] = pd.to_datetime(final_orders['Месяц_ДТ'])
        
        return final_orders
    except Exception as e: 
        print(f"Ошибка загрузки заказов: {e}")
        return pd.DataFrame()

@st.cache_data(ttl=3600) # Кешируем на час (справочник меняется редко)
def load_cached_abc():
    try:
        return pd.read_sql("SELECT article as \"Артикул\", class_abc as \"ABC_Группа\", class_xyz as \"Класс XYZ\" FROM product_classification", engine)
    except: return pd.DataFrame()

@st.cache_data(ttl=300) # Кешируем на 5 минут
def load_cached_history():
    try:
        df = pd.read_sql("SELECT article as \"Артикул\", month_date as \"Месяц_ДТ\", defects as \"Брак\", orders as \"Заказы\", source as \"Source\" FROM historical_ppm", engine)
        if not df.empty: df['Месяц_ДТ'] = pd.to_datetime(df['Месяц_ДТ'])
        return df
    except: return pd.DataFrame()

@st.cache_data(ttl=120)
def build_ppm_base_dataset():
    df_sys = load_cached_hybrid_data()
    df_orders_sys = load_cached_orders()
    df_hist = load_cached_history()
    df_abc = load_cached_abc()
    
    if not df_orders_sys.empty and not df_sys.empty:
        # Векторизованная разметка тегов (в 100 раз быстрее df.apply)
        valid_tag_vals = ['1', '1.0', '+', 'true', 'да']
        tag_cols = [str(i) for i in range(1, 14) if str(i) in df_sys.columns]
        if tag_cols:
            df_tags = df_sys[tag_cols].fillna('').astype(str).apply(lambda x: x.str.strip().str.lower())
            df_sys['Размечено'] = df_tags.isin(valid_tag_vals).any(axis=1)
        else:
            df_sys['Размечено'] = False
            
        df_app_sys = df_sys[df_sys['Размечено'] == True].copy()
        if not df_app_sys.empty:
            df_app_sys['Месяц_ДТ'] = df_app_sys['Дата_ДТ'].dt.to_period('M').dt.to_timestamp()
        
        sys_metrics = df_app_sys.groupby(['Артикул продавца', 'Месяц_ДТ']).size().reset_index(name='Брак') if not df_app_sys.empty else pd.DataFrame(columns=['Артикул продавца', 'Месяц_ДТ', 'Брак'])
        sys_metrics = pd.merge(df_orders_sys, sys_metrics, left_on=['Артикул продавца', 'Месяц_ДТ'], right_on=['Артикул продавца', 'Месяц_ДТ'], how='left').fillna(0)
        sys_metrics.rename(columns={'Артикул продавца':'Артикул', 'Чистые_заказы':'Заказы'}, inplace=True)
        sys_metrics['Source'] = 'System'

        df_total = pd.concat([df_hist, sys_metrics], ignore_index=True)
        if not df_abc.empty:
            df_total = pd.merge(df_total, df_abc, on='Артикул', how='left')
            df_total['ABC_Группа'] = df_total['ABC_Группа'].fillna('C')
            df_total['Класс XYZ'] = df_total['Класс XYZ'].fillna('-')
        else:
            df_total['ABC_Группа'] = 'C'; df_total['Класс XYZ'] = '-'

        months_ru = {1:'Янв', 2:'Фев', 3:'Мар', 4:'Апр', 5:'Май', 6:'Июн', 7:'Июл', 8:'Авг', 9:'Сен', 10:'Окт', 11:'Ноя', 12:'Дек'}
        df_total['Месяц_Стр'] = df_total['Месяц_ДТ'].dt.month.map(months_ru) + " " + df_total['Месяц_ДТ'].dt.year.astype(str)
        return df_total
    return pd.DataFrame()
    
from datetime import timedelta

if page == "Робот-Синхронизатор":
    st.title(":material/database: Статус Базы Данных (Supabase)")
    st.info("Сбор логистики и обращений работает автоматически через GitHub Actions.")
    
    if engine:
        try:
            with engine.connect() as conn:
                # 1. Считаем общее количество
                claims_count = conn.execute(text("SELECT COUNT(*) FROM wb_claims")).scalar() or 0
                orders_count = conn.execute(text("SELECT COUNT(*) FROM wb_orders")).scalar() or 0
                sales_count = conn.execute(text("SELECT COUNT(*) FROM wb_logistics WHERE doc_type='SALE'")).scalar() or 0
                
                # 2. Находим время последнего успешного завершения воркера
                last_sync_raw = conn.execute(text("SELECT MAX(last_sync) FROM wb_claims")).scalar()
                
                # Инициализируем дельты нулями
                claims_delta, orders_delta, sales_delta = 0, 0, 0
                sync_time_str = "Нет данных"

                if last_sync_raw:
                    # Корректировка времени на Московское (UTC+3)
                    # Если last_sync_raw без часового пояса, просто прибавляем 3 часа
                    last_sync_msk = last_sync_raw + timedelta(hours=3)
                    sync_time_str = last_sync_msk.strftime('%d.%m.%Y в %H:%M')
                    
                    # 3. Считаем, сколько записей было затронуто именно в этот запуск (по точному времени)
                    claims_delta = conn.execute(
                        text("SELECT COUNT(*) FROM wb_claims WHERE last_sync = :ts"), 
                        {"ts": last_sync_raw}
                    ).scalar() or 0
                    
                    # Для заказов и продаж (если в них тоже есть колонка last_sync, иначе используем дату)
                    # Если в wb_orders нет last_sync, оставим фильтр по дате или уберем дельту
                    try:
                        orders_delta = conn.execute(
                            text("SELECT COUNT(*) FROM wb_orders WHERE last_sync = :ts"), 
                            {"ts": last_sync_raw}
                        ).scalar() or 0
                    except:
                        orders_delta = 0 # Если колонки нет в этой таблице

            # Визуализация метрик
            c1, c2, c3 = st.columns(3)
            
            c1.metric(
                "Всего Обращений в БД", 
                f"{claims_count:,}".replace(',', ' '), 
                delta=f"+{claims_delta} в последнем пакете" if claims_delta > 0 else None
            )
            c2.metric(
                "Строк Заказов (ORDER)", 
                f"{orders_count:,}".replace(',', ' '), 
                delta=f"+{orders_delta} новых" if orders_delta > 0 else None
            )
            c3.metric(
                "Строк Продаж (SALE)", 
                f"{sales_count:,}".replace(',', ' ')
            )
            
            st.success(f"✅ Последняя синхронизация (МСК): **{sync_time_str}**")
            
        except Exception as e: 
            st.error(f"⚠️ Ошибка получения метрик: {e}")
    else: 
        st.warning("⚠️ База данных не подключена. Проверьте DB_URL.")

elif page == "Обучение ИИ":
    st.title(":material/psychology: База знаний ИИ (Умный импорт)")
    st.markdown("Загрузите исторический файл с проверенными отзывами. Робот всё поймет, расшифрует теги и загрузит в свою память")
    f_import = st.file_uploader("📂 Загрузить базу знаний (Excel/CSV)", type=['xlsx', 'csv', 'xls'])

    if st.button("📥 Загрузить и обновить память", type="primary"):
        if f_import:
            with st.spinner("Анализируем структуру файла и разрешаем конфликты..."):
                # УМНОЕ ЧТЕНИЕ ФАЙЛА
                bytes_data = f_import.getvalue()
                name = f_import.name.lower()
                df_import = pd.DataFrame()
                
                try:
                    if name.endswith(('.xlsx', '.xls')):
                        eng = 'calamine' if name.endswith('.xlsx') else 'xlrd'
                        try:
                            df_import = pd.read_excel(io.BytesIO(bytes_data), engine=eng)
                        except:
                            df_import = pd.read_excel(io.BytesIO(bytes_data), engine='openpyxl')
                    else:
                        for enc in ['utf-8-sig', 'utf-8', 'windows-1251']:
                            for sep in [';', '\t', ',']:
                                try:
                                    df_import = pd.read_csv(io.BytesIO(bytes_data), sep=sep, engine='python', encoding=enc)
                                    if len(df_import.columns) > 1: break
                                except: pass
                                
                    if not df_import.empty:
                        # Проверка на сбитые заголовки (как в "Датасете для отзывов")
                        if df_import.columns.astype(str).str.contains('Unnamed').sum() > 2:
                            row0 = df_import.iloc[0].astype(str).str.lower()
                            if any('текст' in s for s in row0.values):
                                new_header = df_import.iloc[0]
                                df_import = df_import[1:]
                                df_import.columns = new_header
                except Exception as e:
                    st.error(f"⚠️ Ошибка чтения файла: {e}")

                if not df_import.empty:
                    # Ищем все возможные текстовые колонки (чтобы склеить Достоинства и Недостатки)
                    text_cols = [str(c) for c in df_import.columns if any(kw in str(c).lower() for kw in ['текст', 'достоинства', 'недостатки', 'comment', 'комментарий покупателя'])]
                    
                    # Ищем колонку с готовыми текстовыми решениями (как в "Лилиях")
                    corr_col = next((str(c) for c in df_import.columns if any(kw in str(c).lower() for kw in ['корректировка', 'исправление', 'комментарий']) and str(c).lower() not in [str(x).lower() for x in text_cols]), None)
                    
                    if not text_cols: 
                        st.error("❌ Ошибка: В файле не найдены колонки с текстом.")
                    else:
                        new_memory_dict = {}
                        for idx, row in df_import.iterrows():
                            # Собираем текст, игнорируя ошибки типов
                            parts = []
                            for tc in text_cols:
                                if tc in row and pd.notna(row[tc]):
                                    cell_val = str(row[tc]).strip()
                                    if cell_val.lower() not in ['nan', 'none', '']:
                                        parts.append(cell_val)
                            
                            combined_text = " ".join(parts).strip()
                            if not combined_text: continue
                            
                            final_tags = []
                            
                            # 1. Приоритет ручному текстовому комментарию (если он есть)
                            if corr_col and pd.notna(row[corr_col]):
                                val = str(row[corr_col]).strip()
                                # Разбиваем строку по точке с запятой или запятой
                                split_vals = [v.strip() for v in re.split(r'[;,]', val) if v.strip()]
                                for v in split_vals:
                                    # Проверяем, есть ли такое значение в нашем словаре
                                    if v in CATEGORIES.values():
                                        final_tags.append(v)
                                    
                            # 2. Если ручного текста нет, ищем отметки 1/+/Да в колонках
                            if not final_tags:
                                for col in df_import.columns:
                                    col_str = str(col).lower()
                                    cat_id = None
                                    
                                    # Обрабатываем колонки, которые просто названы цифрами (1.0, 2.0)
                                    if isinstance(col, (int, float)):
                                        cat_id = int(col)
                                    else:
                                        # Ищем слово Кат/Cat и цифру рядом
                                        match = re.search(r'(?:кат|cat|категория)?\s*(\d+)', col_str)
                                        if match:
                                            cat_id = int(match.group(1))
                                            
                                    if cat_id and cat_id in CATEGORIES:
                                        cell_val = str(row[col]).strip().lower()
                                        if cell_val in ['1', '1.0', '+', 'true', 'да', 'v']:
                                            final_tags.append(CATEGORIES[cat_id])
                                            
                            if final_tags: 
                                new_memory_dict[combined_text] = "; ".join(final_tags)

                        if new_memory_dict:
                            try:
                                # Сохраняем в SQL через UPSERT (если текст такой же - обновим тег)
                                with engine.begin() as conn:
                                    for txt, tgs in new_memory_dict.items():
                                        conn.execute(text("""
                                            INSERT INTO ai_knowledge_base (content, tags, source) 
                                            VALUES (:txt, :tgs, 'training')
                                            ON CONFLICT (content) DO UPDATE SET tags = EXCLUDED.tags
                                        """), {"txt": txt, "tgs": tgs})
                                
                                st.success(f":material/check_circle: Опыт успешно записан в SQL-базу! Всего добавлено: {len(new_memory_dict)} примеров.")
                            except Exception as e: 
                                st.error(f"Ошибка записи в SQL: {e}")
                        else: 
                            st.warning("⚠️ Не найдено валидных тегов в файле. Проверьте, стоят ли '1' в колонках категорий.")
        else: 
            st.warning("Пожалуйста, загрузите файл.")

elif page == "ИИ Тегирование":
    st.title(":material/biotech: ИИ Тегирование и Проверка")
    
    if engine:
        # ПУНКТ 4: Исключаем те, что ИИ не смог распознать
        query_unprocessed = """
            SELECT srid, supplier_article, user_comment 
            FROM wb_claims 
            WHERE NOT (cat_1 OR cat_2 OR cat_3 OR cat_4 OR cat_5 OR cat_6 OR cat_7 OR cat_8 OR cat_9 OR cat_10 OR cat_11 OR cat_12 OR cat_13)
            AND (audit_status IS NULL OR audit_status != 'Пропущено ИИ')
        """
        df_unprocessed = pd.read_sql(query_unprocessed, engine)
        
        query_audit = """
            SELECT srid, user_comment, cat_1, cat_2, cat_3, cat_4, cat_5, cat_6, cat_7, cat_8, cat_9, cat_10, cat_11, cat_12, cat_13 
            FROM wb_claims 
            WHERE (cat_1 OR cat_2 OR cat_3 OR cat_4 OR cat_5 OR cat_6 OR cat_7 OR cat_8 OR cat_9 OR cat_10 OR cat_11 OR cat_12 OR cat_13)
            AND (audit_status IS NULL OR audit_status = '')
        """
        df_audit = pd.read_sql(query_audit, engine)
        
        t1, t2 = st.tabs([":material/label: Первичная разметка", ":material/rule: Перекрестная проверка (Grok)"])
        
        with t1:
            st.subheader("Разметка новых заявок (Только ID)")
            if not df_unprocessed.empty:
                total_rows = len(df_unprocessed)
                col1, col2 = st.columns(2)
                batch_size = col1.slider("Размер пачки", 5, 50, 10, key="batch_tag")
                model_choice = col2.radio("Модель:", ["YandexGPT Lite (Дешево)", "YandexGPT Pro (Умнее)", "Grok (xAI)"], key="mod_tag")
                
                model_key = "yandex-lite" if "Lite" in model_choice else "yandex-pro" if "Pro" in model_choice else "grok"
                est_cost = total_rows * (0.08 if model_key == "yandex-lite" else 0.40 if model_key == "yandex-pro" else 0.50)
                st.info(f"📊 **Аналитика:** Найдено **{total_rows}** строк без тегов.\n💰 **Предварительный расход:** ~{est_cost:.2f} руб.")
                
                if st.button("🚀 ЗАПУСТИТЬ ТЕГИРОВАНИЕ", type="primary"):
                    st.cache_data.clear() 
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    log_container = st.container()
                    add_system_log("Запуск тегирования", "INFO", f"Строк: {total_rows}")
                    
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    
                    for i in range(0, total_rows, batch_size):
                        chunk = df_unprocessed.iloc[i:i+batch_size].copy()
                        
                        # 1. Готовим "шпаргалку" из БД для пачки
                        memory_list = []
                        for _, row in chunk.iterrows():
                            t_text = str(row.get('user_comment',''))
                            if t_text:
                                mem = find_similar_examples_sql(t_text, engine, top_n=2)
                                if mem and "Прямых совпадений" not in mem:
                                    memory_list.append(mem)
                        chunk_memory = "\n".join(set(memory_list)) if memory_list else "Опыта пока нет."
                        
                        # 2. ВОЗВРАЩАЕМ СТАРУЮ ХИТРОСТЬ: Делаем короткие REF_ID
                        # Это защитит нас от того, что ИИ перепутает сложный SRID
                        chunk['temp_id'] = [f"REF_{x}" for x in range(len(chunk))]
                        srid_map = dict(zip(chunk['temp_id'], chunk['srid'])) # Словарь REF -> настоящий SRID
                        
                        # Формируем данные для отправки с короткими ID
                        batch_to_send = []
                        for _, row in chunk.iterrows():
                            batch_to_send.append({
                                "id": row['temp_id'], 
                                "text": f"Артикул: {row.get('supplier_article','')}. Текст: {row.get('user_comment','')}"
                            })
                        
                        # 3. Отправляем в ИИ
                        async def send_batch():
                            async with aiohttp.ClientSession() as session:
                                return await fetch_ai_tags(session, batch_to_send, chunk_memory, model_key)
                        
                        results = loop.run_until_complete(send_batch())
                        
                        # 4. Сохраняем в SQL и пишем подробности в SQL-журнал
                        if results:
                            saved_count = 0
                            batch_details = [] # Список для формирования TEXT в колонку details
                            
                            for res in results:
                                if "error" in res:
                                    batch_details.append(f"❌ Ошибка ИИ: {res.get('error')}")
                                    continue
                                
                                # Очистка ID
                                raw_id = str(res.get('id', '')).upper()
                                num_match = re.search(r'\d+', raw_id)
                                
                                if not num_match: 
                                    batch_details.append(f"⚠️ Сбой ID: ИИ вернул '{raw_id}' без цифр")
                                    continue 
                                    
                                clean_temp_id = f"REF_{num_match.group()}"
                                cats_array = res.get('category_ids', [])
                                real_srid = srid_map.get(clean_temp_id)
                                
                                if real_srid:
                                    if cats_array:
                                        updates = {f"cat_{re.search(r'\d+', str(c)).group()}": True for c in cats_array if re.search(r'\d+', str(c))}
                                        if updates:
                                            if update_db_row(real_srid, updates):
                                                saved_count += 1
                                                batch_details.append(f"✅ SRID {real_srid}: теги {cats_array}")
                                    else:
                                        # ПУНКТ 4: Если тегов нет, помечаем как "Пропущено", чтобы не зацикливалось!
                                        update_db_row(real_srid, {"audit_status": "Пропущено ИИ"})
                                        batch_details.append(f"⚠️ SRID {real_srid}: ИИ не смог определить теги (Пропущено)")
                                else:
                                    batch_details.append(f"❓ ПРОПУСК: SRID для {clean_temp_id} не найден в маппинге")
                            
                            # Формируем итоговый текст для колонки details в system_logs
                            full_log_text = f"Пачка обработана: {saved_count} из {len(chunk)} сохранены.\n" + "\n".join(batch_details)
                            
                            # Пишем в SQL через твой add_system_log
                            if saved_count > 0:
                                add_system_log(f"Пачка {i}", "SUCCESS", full_log_text)
                            else:
                                add_system_log(f"Пачка {i}", "WARNING", full_log_text)
                        else:
                            add_system_log(f"Пачка {i}", "ERROR", "ИИ вернул пустой результат (результаты отсутствуют)")

                        # ==========================================
                        # === ВОТ ОНО: ОБНОВЛЕНИЕ ПРОГРЕСС-БАРА ===
                        # ==========================================
                        current_processed = min(total_rows, i + len(chunk))
                        progress_percent = current_processed / total_rows
                        progress_bar.progress(progress_percent)
                        status_text.text(f"⏳ Прогресс: {int(progress_percent * 100)}% ({current_processed} из {total_rows})")
                        # ==========================================

                    st.success("✅ Тегирование успешно завершено! Проверьте отчет производства.")
                    st.rerun()
            else:
                st.success("🎉 Все заявки в базе имеют первичную разметку!")

        with t2:
            st.subheader("Глубокая проверка (Аудит)")
            if not df_audit.empty:
                total_audit_rows = len(df_audit)
                col1, col2 = st.columns(2)
                batch_size_audit = col1.slider("Размер пачки для аудита", 5, 50, 10, key="batch_audit")
                # ПУНКТ 5: Выбор ИИ
                model_audit = col2.radio("Модель аудитора:", ["YandexGPT Lite", "YandexGPT Pro", "Grok (xAI)"], key="mod_audit")
                model_audit_key = "yandex-lite" if "Lite" in model_audit else "yandex-pro" if "Pro" in model_audit else "grok"
                
                if st.button("🕵️‍♂️ ЗАПУСТИТЬ АУДИТ", type="primary"):
                    progress_bar_audit = st.progress(0)
                    status_text_audit = st.empty()
                    
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    
                    for i in range(0, total_audit_rows, batch_size_audit):
                        chunk = df_audit.iloc[i:i+batch_size_audit]
                        status_text_audit.text(f"⏳ Аудит: {int(((i + len(chunk)) / total_audit_rows) * 100)}%")
                        
                        # Передаем выбранную модель!
                        results = loop.run_until_complete(run_ai_batch_processing(chunk, model_audit_key, mode="crosscheck"))
                        
                        for res in results:
                            if "error" in res: continue
                            srid = res.get('id')
                            audit_status = str(res.get('audit', '')).strip()
                            comment_text = str(res.get('comment', '')).strip()
                            cats_array = res.get('category_ids', [])
                            
                            updates = {'audit_status': audit_status, 'audit_comment': comment_text}
                            
                            if audit_status.upper() != "ОК" and cats_array:
                                # Сбрасываем старые теги и ставим новые
                                for c in range(1, 14): updates[f'cat_{c}'] = False
                                for cat_val in cats_array:
                                    cat_num_match = re.search(r'\d+', str(cat_val))
                                    if cat_num_match: updates[f"cat_{cat_num_match.group()}"] = True
                            
                            if srid: update_db_row(srid, updates)
                            
                        progress_bar_audit.progress(min(1.0, (i + len(chunk)) / total_audit_rows))
                    st.success("✅ Аудит завершен! Ошибки исправлены.")
                    st.rerun()
            else:
                st.success("🎉 Все размеченные заявки проверены аудитором!")

elif page == "Модерация":
    st.title(":material/fact_check: Модерация (Ручная проверка)")

    @st.dialog("Просмотр видео")
    def play_video_modal(url): 
        st.video(url)

    def render_pagination(total_pages, key_prefix):
        if total_pages <= 1: return
        curr = st.session_state.mod_page
        if total_pages <= 7: window = list(range(1, total_pages + 1))
        elif curr <= 4: window = [1, 2, 3, 4, 5, "...", total_pages]
        elif curr >= total_pages - 3: window = [1, "...", total_pages - 4, total_pages - 3, total_pages - 2, total_pages - 1, total_pages]
        else: window = [1, "...", curr - 1, curr, curr + 1, "...", total_pages]

        _, center_col, _ = st.columns([1, 2, 1])
        with center_col:
            cols = st.columns(len(window) + 2, gap="small")
            with cols[0]:
                if st.button("«", key=f"{key_prefix}_prev", disabled=(curr == 1), use_container_width=True): 
                    st.session_state.mod_page -= 1
                    st.rerun()
            for i, p in enumerate(window):
                with cols[i+1]:
                    if p == "...": 
                        st.markdown("<div style='text-align: center; color: #888; padding-top: 5px; font-weight: bold;'>...</div>", unsafe_allow_html=True)
                    else:
                        if st.button(str(p), key=f"{key_prefix}_page_{p}", type="primary" if p == curr else "secondary", use_container_width=True): 
                            st.session_state.mod_page = p
                            st.rerun()
            with cols[-1]:
                if st.button("»", key=f"{key_prefix}_next", disabled=(curr == total_pages), use_container_width=True): 
                    st.session_state.mod_page += 1
                    st.rerun()

    if engine:
        st.markdown("### 🔍 Фильтры обращений")
        c1, c2 = st.columns(2)
        filter_mode = c1.radio("Показать обращения:", ["Все ожидающие модерации", "С замечаниями от Аудитора"], horizontal=True)
        
        # ПУНКТ 1: Добавляем фильтр по категориям
        cat_options = ["Все категории"] + list(CATEGORIES.values())
        selected_cat_filter = c2.selectbox("Фильтр по категории ИИ:", cat_options)
        
        query = """
            SELECT *
            FROM view_cx_dashboard
            WHERE ("1" OR "2" OR "3" OR "4" OR "5" OR "6" OR "7" OR "8" OR "9" OR "10" OR "11" OR "12" OR "13")
            AND ("Корректировка" IS NULL OR "Корректировка" = '')
        """
        if filter_mode == "С замечаниями от Аудитора": query += " AND UPPER(\"Аудит\") LIKE '%ОШИБКА%'"
        
        to_review = pd.read_sql(query, engine)
        reverse_cats = {v.strip().lower(): k for k, v in CATEGORIES.items()}
        
        # ПРИМЕНЯЕМ ФИЛЬТР КАТЕГОРИИ
        if selected_cat_filter != "Все категории" and not to_review.empty:
            cat_id_str = str(reverse_cats.get(selected_cat_filter.lower()))
            if cat_id_str in to_review.columns:
                to_review = to_review[to_review[cat_id_str].astype(str).str.lower().isin(['1', '1.0', '+', 'true', 'да'])]
        
        if not to_review.empty:
            total_items = len(to_review)
            ITEMS_PER_PAGE = 20
            total_pages = max(1, (total_items - 1) // ITEMS_PER_PAGE + 1)
            
            if 'mod_page' not in st.session_state: st.session_state.mod_page = 1
            if st.session_state.mod_page > total_pages: st.session_state.mod_page = 1
            
            st.success(f"Найдено обращений в очереди: **{total_items}**")
            render_pagination(total_pages, key_prefix="top")
            
            start_idx = (st.session_state.mod_page - 1) * ITEMS_PER_PAGE
            current_page_df = to_review.iloc[start_idx:start_idx + ITEMS_PER_PAGE]
            
            cats_list = list(CATEGORIES.values())
            reverse_cats = {v.strip().lower(): k for k, v in CATEGORIES.items()}
            
            for idx, row in current_page_df.iterrows():
                srid = str(row['SRID']).strip()
                st.markdown("---")
                col_info, col_media = st.columns([1.2, 1])
                
                with col_info:
                    st.markdown(f"**Артикул:** {row.get('Артикул продавца', '---')} | **Дата заявки:** {row.get('Дата и время оформления заявки на возврат', '')}")
                    st.info(row.get('Комментарий покупателя', 'Нет текста'))
                    
                    ai_selected = [CATEGORIES[i] for i in range(1, 14) if row.get(str(i)) == True]
                    ai_text = ", ".join(ai_selected) if ai_selected else "Категории не определены"
                    st.markdown(f'<div class="ai-tags-box">🤖 <b>Выбор ИИ:</b> {ai_text}</div>', unsafe_allow_html=True)
                    
                    selected_cats = st.multiselect("Выберите правильные категории (если ИИ ошибся):", options=cats_list, default=[], key=f"ms_{srid}")
                    
                    btn_col1, btn_col2 = st.columns(2)
                    with btn_col1:
                        if st.button("💾 Исправить", key=f"btn_{srid}", type="primary"):
                            if selected_cats:
                                updates = {f"cat_{i}": False for i in range(1, 14)}
                                for cat_name in selected_cats:
                                    cat_num = reverse_cats.get(cat_name.strip().lower())
                                    if cat_num: updates[f"cat_{cat_num}"] = True
                                updates["correction"] = "; ".join(selected_cats)
                                if update_db_row(srid, updates):
                                    st.rerun()

                    with btn_col2:
                        if st.button("✅ Подтвердить (Без изменений)", key=f"ok_{srid}"):
                            if update_db_row(srid, {"correction": "Подтверждено"}):
                                st.rerun()

                with col_media:
                    # НОВАЯ ЛОГИКА: Запрашиваем медиа и разделяем пары превью|оригинал
                    raw_photos, raw_videos = get_media_for_srid(srid)
                    
                    # Обработка фото (делим строку на группы "превью|оригинал")
                    photo_groups = raw_photos.split()
                    # Обработка видео
                    video_urls = re.findall(r'(?:https?:)?//[^\s"\'\;\]\[,<>]+', raw_videos)
                    
                    if photo_groups:
                        # Стили пишем без лишних отступов в начале строк
                        html_imgs = '<style>.mod-zoom { transition: transform 0.2s ease; cursor: pointer; border-radius: 8px; object-fit: cover; } .mod-zoom:hover { transform: scale(2.5); z-index: 9999; position: relative; border-radius: 0px; box-shadow: 0 10px 30px rgba(0,0,0,0.5); }</style>'
                        html_imgs += '<div style="display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 15px;">'
                        
                        for group in photo_groups[:6]:
                            if "|" in group:
                                s3_url, wb_url = group.split("|", 1)
                            else:
                                s3_url = wb_url = group
                            
                            # ВАЖНО: Вся f-строка должна быть либо в одну линию, либо без пробелов в начале каждой строки
                            html_imgs += f'<div style="text-align: center; width: 130px;"><a href="{wb_url}" target="_blank" rel="noreferrer noopener"><img src="{s3_url}" class="mod-zoom" style="width: 130px; height: 130px;" referrerpolicy="no-referrer"></a><a href="{wb_url}" download target="_blank" style="text-decoration:none; font-size:10px; color:#3498db; display:block; margin-top:5px;">📥 Оригинал</a></div>'
                        
                        html_imgs += '</div>'
                        st.markdown(html_imgs, unsafe_allow_html=True)
                    
                    if video_urls:
                        for v_idx, v_url in enumerate(video_urls):
                            if st.button("🎥 Видео", key=f"vid_{srid}_{v_idx}"): 
                                play_video_modal(v_url)
                
            st.markdown("---")
            render_pagination(total_pages, key_prefix="bottom")
        else: 
            st.success("🎉 Очередь пуста! Все обращения проверены.")

elif page == "Отчет производства":
    st.title(":material/insights: Отчет производства")
    
    import altair as alt 
    import time
    from collections import defaultdict
    
    @st.dialog("Детализация пересечения", width="large")
    def show_matrix_details(sku, reason_name, filtered_df, reason_id):
        st.subheader(f"📦 Артикул: {sku} | 🛠 Причина: {reason_name}")
        
        # ИСПРАВЛЕНИЕ 1: Приводим Артикулы к тому же виду, что и в матрице, чтобы избежать сбоя при "Без артикула"
        clean_skus = filtered_df['Артикул продавца'].astype(str).str.strip().replace('', 'Без артикула')
        details = filtered_df[(clean_skus == sku) & (filtered_df[str(reason_id)].astype(str).str.strip().str.lower().isin(['1', '1.0', '+', 'true', 'да']))]
        
        if not details.empty:
            all_original_photos = []
            for _, r in details.iterrows():
                p_raw, _ = get_media_for_srid(r['SRID'])
                for group in p_raw.split():
                    wb_url = group.split("|")[-1]
                    if wb_url.startswith("//"): wb_url = "https:" + wb_url
                    all_original_photos.append(wb_url)
            
            if all_original_photos:
                if st.button(f"📥 Скачать ВСЕ ОРИГИНАЛЫ ({len(all_original_photos)} шт.)", type="primary", key=f"dl_all_{sku}_{reason_id}"):
                    with st.spinner("Сбор оригинальных фото..."):
                        zip_all = create_images_zip(all_original_photos)
                        b64 = base64.b64encode(zip_all).decode()
                        components.html(f'<a id="dl" href="data:application/zip;base64,{b64}" download="{sku}_{reason_id}_ORIGINALS.zip"></a><script>document.getElementById("dl").click();</script>', width=0, height=0)
            
            st.markdown("---")
            for _, r in details.iterrows():
                with st.container():
                    st.markdown('<div class="detail-card">', unsafe_allow_html=True)
                    c1, media_col = st.columns([1.2, 1])
                    
                    p_raw, v_raw = get_media_for_srid(r['SRID'])
                    photo_groups = p_raw.split()
                    video_urls = re.findall(r'(?:https?:)?//[^\s"\'\;\]\[,]+', v_raw)

                    with c1:
                        st.write(f"💬 **Текст клиента:**\n{r.get('Комментарий покупателя', '---')}")
                        st.write(f"🧾 **Инвойс:** {r.get('Инвойс', '---')} | **Поставка:** {r.get('Номер поставки_ОРИГИНАЛ', '---')}")
                    
                    with media_col:
                        if photo_groups:
                            html_imgs = '<div class="media-row">'
                            for group in photo_groups[:6]:
                                if "|" in group:
                                    s3_url, wb_url = group.split("|", 1)
                                else:
                                    s3_url = wb_url = group
                                html_imgs += f'<a href="{wb_url}" target="_blank"><img src="{s3_url}" class="photo-zoom"></a>'
                            html_imgs += '</div>'
                            st.markdown(html_imgs, unsafe_allow_html=True)
                            
                        if video_urls:
                            for v_idx, v_url in enumerate(video_urls): 
                                st.markdown(f'<a href="{v_url}" target="_blank" class="video-link-btn">🎥 Видео {v_idx+1}</a>', unsafe_allow_html=True)
                    st.markdown('</div>', unsafe_allow_html=True)
        else: 
            st.write("Нет данных по этому пересечению.")

        if st.button("Закрыть детализацию"):
            st.session_state.matrix_key = int(time.time())
            st.rerun()

    # --- ДИАЛОГ ДЛЯ ИНВОЙСОВ ---
    @st.dialog("Детализация Инвойса", width="medium")
    def show_invoice_details(invoice, filtered_df):
        st.subheader(f"🧾 Инвойс: {invoice}")
        
        # ИСПРАВЛЕНИЕ 2: Приводим Инвойсы к тому же виду, что на графике, для защиты от 'nan' / 'None'
        clean_inv = filtered_df['Инвойс'].astype(str).str.strip().replace(['nan', 'None', ''], 'Не указан')
        inv_details = filtered_df[clean_inv == invoice]
        
        if not inv_details.empty:
            st.write(f"**Всего дефектных заявок в инвойсе:** {len(inv_details)}")
            st.markdown("---")
            
            sku_stats = defaultdict(lambda: defaultdict(int))
            valid_tag_vals = ['1', '1.0', '+', 'true', 'да']
            
            for _, r in inv_details.iterrows():
                sku = str(r.get('Артикул продавца', 'Без артикула')).strip()
                if not sku: sku = 'Без артикула'
                for i in range(1, 14):
                    cat_col = str(i)
                    if cat_col in r and str(r.get(cat_col, '')).strip().lower() in valid_tag_vals:
                        sku_stats[sku][CATEGORIES[i]] += 1
            
            for sku, problems in sku_stats.items():
                st.markdown(f"#### 📦 {sku}")
                for prob, count in problems.items():
                    st.markdown(f"- 🛠 **{prob}**: {count} шт.")
                st.markdown("<br>", unsafe_allow_html=True)
        else:
            st.info("Нет данных по этому инвойсу.")

        if st.button("Закрыть окно", type="primary"):
            st.session_state.invoice_key = int(time.time()) 
            st.rerun()

    # ==========================================
    # ТЯЖЕЛАЯ ЧАСТЬ
    # ==========================================
    try:
        with st.spinner("📊 Загрузка и анализ данных..."):
            df = load_cached_hybrid_data()

        if not df.empty:
            if 'Инвойс' not in df.columns: df['Инвойс'] = 'Не указан'
            df['Инвойс'] = df['Инвойс'].fillna('Не указан')
            df['Номер поставки_ОРИГИНАЛ'] = df.get('Номер поставки_ОРИГИНАЛ', 'Не указан').fillna('Не указан')
            
            valid_tag_vals = ['1', '1.0', '+', 'true', 'да']
            tag_cols = [str(i) for i in range(1, 14) if str(i) in df.columns]
            if tag_cols:
                df_tags = df[tag_cols].fillna('').astype(str).apply(lambda x: x.str.strip().str.lower())
                df['Размечено'] = df_tags.isin(valid_tag_vals).any(axis=1)
            else:
                df['Размечено'] = False
            
            inv_list = ['Все'] + sorted(list(set([str(x) for x in df['Инвойс'] if str(x).strip() and str(x) != 'Не указан'])))
            sku_list = ['Все'] + sorted(list(set([str(x) for x in df['Артикул продавца'] if str(x).strip()])))

            @st.fragment
            def render_production_dashboard(df_full):
                # 1. Объявляем константы в самом начале фрагмента
                valid_tag_vals = ['1', '1.0', '+', 'true', 'да']
                
                # --- Глобальные фильтры (ваша логика) ---
                st.markdown("### 🔍 Глобальные фильтры")
                f_col1, f_col2, f_col3 = st.columns(3)
                with f_col1:
                    today = datetime.now().date()
                    start_month = today.replace(day=1)
                    date_range = st.date_input("Период анализа:", [start_month, today], format="DD.MM.YYYY", key="prod_date_filter")
                
                selected_sku = f_col2.selectbox("Артикул:", sku_list)
                selected_inv = f_col3.selectbox("Инвойс / Поставка:", inv_list)
                
                # Фильтрация для статистики и хитмапа
                df_filtered = df_full.copy()
                if len(date_range) == 2:
                    df_filtered = df_filtered[(df_filtered['Дата_ДТ'].dt.date >= date_range[0]) & (df_filtered['Дата_ДТ'].dt.date <= date_range[1])]
                if selected_sku != 'Все': df_filtered = df_filtered[df_filtered['Артикул продавца'].astype(str) == selected_sku]
                if selected_inv != 'Все': df_filtered = df_filtered[df_filtered['Инвойс'].astype(str) == selected_inv]

                total_rows = len(df_filtered)
                tagged_rows = df_filtered['Размечено'].sum()
                
                if 'Корректировка' in df_filtered.columns:
                    corr_col = df_filtered['Корректировка'].fillna('')
                    corr_clean = corr_col.astype(str).str.strip().str.lower()
                    corrected_rows = len(df_filtered[(corr_clean != '') & (~corr_clean.isin(['nan', 'none', 'null', 'подтверждено', 'нет тегов']))])
                else:
                    corrected_rows = 0
                
                accuracy = round((1 - (corrected_rows / tagged_rows)) * 100, 1) if tagged_rows > 0 else 0
                processed_percent = round((tagged_rows / total_rows) * 100, 1) if total_rows > 0 else 0
                
                st.markdown("### 📈 Общая статистика (ТОЛЬКО 'Одобрено')")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Всего заявок", total_rows)
                c2.metric("Размечено", f"{tagged_rows} ({processed_percent}%)")
                c3.metric("Изменено вручную", corrected_rows)
                c4.metric("Точность ИИ", f"{accuracy}%")
                st.markdown("---")
                
                st.info("💡 **Кликните на любой цветной квадрат (или столбец инвойса ниже) для мгновенной детализации!**")
                
                matrix_list = []
                if tag_cols:
                    df_tags_filt = df_filtered[tag_cols].fillna('').astype(str).apply(lambda x: x.str.strip().str.lower())
                    for i in range(1, 14):
                        cat_col = str(i)
                        if cat_col in df_tags_filt.columns:
                            valid_rows = df_tags_filt[cat_col].isin(valid_tag_vals)
                            if valid_rows.any():
                                temp = df_filtered[valid_rows]
                                temp_matrix = pd.DataFrame({
                                    'Артикул продавца': temp['Артикул продавца'].astype(str).str.strip().replace('', 'Без артикула'),
                                    'Причина': f"{i}. {CATEGORIES[i]}",
                                    'ID': i,
                                    'Инвойс': temp['Инвойс'].astype(str).str.strip().replace(['nan', 'None', ''], 'Не указан'),
                                    'Номер поставки': temp.get('Номер поставки_ОРИГИНАЛ', temp['Инвойс']).astype(str).str.strip()
                                })
                                matrix_list.extend(temp_matrix.to_dict('records'))

                if matrix_list:
                    df_matrix = pd.DataFrame(matrix_list)
                    pivot = pd.crosstab(df_matrix['Причина'], df_matrix['Артикул продавца']).fillna(0).astype(int)
                    pivot['ID'] = [int(x.split('.')[0]) for x in pivot.index]
                    reason_totals = pivot.drop(columns=['ID']).sum(axis=1).to_dict()
                    df_melt = pivot.reset_index().melt(id_vars=['Причина', 'ID'], var_name='Артикул продавца', value_name='Дефекты')
                    df_melt['Артикул_Метка'] = df_melt['Артикул продавца'] 
                    df_melt['Причина_Метка'] = df_melt['Причина'].apply(lambda x: f"{x} [{reason_totals.get(x, 0)}]")
                    df_melt['Текст'] = df_melt['Дефекты'].apply(lambda x: str(x) if x > 0 else "")

                    click_selector = alt.selection_point(name='cell_click', fields=['Артикул_Метка', 'Причина_Метка'])
                    base = alt.Chart(df_melt).encode(x=alt.X('Артикул_Метка:N', title=None, axis=alt.Axis(labelAngle=-90, labelLimit=1000, orient='bottom')), y=alt.Y('Причина_Метка:N', title=None, axis=alt.Axis(labelLimit=1000), sort=alt.EncodingSortField(field='ID', order='ascending')))
                    rects = base.mark_rect(stroke='white', strokeWidth=1).encode(color=alt.Color('Дефекты:Q', scale=alt.Scale(scheme='blues'), legend=None), tooltip=[alt.Tooltip('Артикул продавца:N', title='Артикул'), alt.Tooltip('Причина:N', title='Причина'), alt.Tooltip('Дефекты:Q', title='Кол-во')])
                    
                    text_marks = base.mark_text(baseline='middle', fontSize=11).encode(text='Текст:N', color=alt.condition(alt.datum.Дефекты > (df_melt['Дефекты'].max() / 2), alt.value('white'), alt.value('black')))
                    final_chart = alt.layer(rects, text_marks).properties(height=max(400, len(pivot) * 35 + 100)).add_params(click_selector)
                    
                    current_matrix_key = st.session_state.get('matrix_key', 0)
                    chart_key = f"prod_matrix_{current_matrix_key}"
                    event = st.altair_chart(final_chart, use_container_width=True, on_select="rerun", key=chart_key)
                    
                    if event and hasattr(event, "selection"):
                        sel = event.selection.get("cell_click", [])
                        if sel and len(sel) > 0:
                            sku_clicked = sel[0].get('Артикул_Метка')
                            reason_clicked = sel[0].get('Причина_Метка')
                            if sku_clicked and reason_clicked:
                                clean_sku = sku_clicked.split(' [')[0]
                                clean_reason = reason_clicked.split(' [')[0]
                                reason_id = int(reason_clicked.split('.')[0])
                                show_matrix_details(clean_sku, clean_reason, df_filtered, reason_id)

                    # ==========================================================
                    # ОБНОВЛЕННЫЙ БЛОК: ДИНАМИКА SKU (Plotly Trend)
                    # ==========================================================
                    st.markdown("---")
                    st.markdown("### 📈 Детальная динамика и Исторический тренд")
                    
                    all_sku_options = [s for s in sku_list if s != 'Все']
                    
                    c_sku, _ = st.columns([1.5, 2])
                    sku_dyn_target = c_sku.selectbox("Выберите SKU для анализа тренда:", all_sku_options, key="sku_trend_select")

                    if sku_dyn_target:
                        with st.spinner("Сбор истории и системных данных..."):
                            plot_data = []
                            
                            # 1. Системные данные
                            df_sku_sys = df_full[df_full['Артикул продавца'].astype(str).str.strip() == sku_dyn_target].copy()
                            if not df_sku_sys.empty:
                                for i in range(1, 14):
                                    cat_col = str(i)
                                    if cat_col in df_sku_sys.columns:
                                        mask = df_sku_sys[cat_col].astype(str).str.strip().str.lower().isin(valid_tag_vals)
                                        temp = df_sku_sys[mask].copy()
                                        if not temp.empty:
                                            temp['Месяц'] = pd.to_datetime(temp['Дата_ДТ']).dt.to_period('M').dt.to_timestamp()
                                            monthly = temp.groupby('Месяц').size().reset_index(name='Количество')
                                            monthly['Источник'] = f"{i}. {CATEGORIES.get(i, f'Категория {i}')}"
                                            plot_data.append(monthly)

                            # 2. Исторические данные (Корректировка названия)
                            try:
                                hist_query = text("SELECT month_date, defects FROM historical_ppm WHERE article = :sku")
                                with engine.connect() as conn:
                                    df_h = pd.read_sql(hist_query, conn, params={"sku": sku_dyn_target})
                                if not df_h.empty:
                                    df_h['Месяц'] = pd.to_datetime(df_h['month_date']).dt.to_period('M').dt.to_timestamp()
                                    df_h = df_h.groupby('Месяц')['defects'].sum().reset_index(name='Количество')
                                    df_h['Источник'] = "Общий брак" # Переименовано по запросу
                                    plot_data.append(df_h)
                            except: pass

                            if plot_data:
                                df_plot = pd.concat(plot_data).sort_values('Месяц')
                                
                                # Сортировка категорий (1-13 и Общий брак в конце)
                                category_order = [f"{i}. {CATEGORIES.get(i, f'Категория {i}')}" for i in range(1, 14)]
                                category_order.append("Общий брак")
                                
                                import plotly.express as px
                                fig_dyn = px.bar(
                                    df_plot,
                                    x='Месяц',
                                    y='Количество',
                                    color='Источник',
                                    title=f"Анализ дефектов по месяцам: {sku_dyn_target}",
                                    text_auto=True,
                                    template='plotly_white',
                                    color_discrete_sequence=px.colors.qualitative.Pastel + px.colors.qualitative.Bold,
                                    # КОРРЕКТИРОВКА 2: Фиксированный порядок легенды и слоев
                                    category_orders={"Источник": category_order}
                                )

                                # КОРРЕКТИРОВКА 1: Убираем жирный шрифт и настраиваем формат "Количество: ***"
                                fig_dyn.update_traces(
                                    hovertemplate="%{fullData.name}<br>Количество: %{y}<extra></extra>"
                                )

                                fig_dyn.update_layout(
                                    hovermode="closest",
                                    height=850, # Еще немного увеличили для удобства навигации
                                    xaxis_title=None,
                                    yaxis_title="Кол-во дефектов (заявки)",
                                    legend=dict(
                                        orientation="h", 
                                        yanchor="bottom", 
                                        y=-0.35, 
                                        xanchor="center", 
                                        x=0.5,
                                        title=None
                                    ),
                                    bargap=0.4
                                )
                                
                                fig_dyn.update_xaxes(dtick="M1", tickformat="%b %Y", tickangle=-45)
                                
                                st.plotly_chart(fig_dyn, use_container_width=True)
                            else:
                                st.info(f"Данных по артикулу {sku_dyn_target} не найдено.")
                    
                    # ==========================================================
                    # ПРОБЛЕМНЫЕ ИНВОЙСЫ
                    # ==========================================================
                    st.markdown("---")
                    st.markdown("### 📦 Проблемные инвойсы (Топ-15)")
                    
                    invoice_grouped = []
                    df_matrix_data = pd.DataFrame(matrix_list)
                    
                    if 'Инвойс' in df_matrix_data.columns:
                        for invoice, group in df_matrix_data.groupby('Инвойс'):
                            clean_invoice = str(invoice).strip()
                            if clean_invoice in ['Не указан', '', '0', '0.0'] or pd.isna(invoice): 
                                continue
                            
                            supplies = ", ".join(sorted(list(set([str(x) for x in group['Номер поставки'] if str(x) != 'Не указан' and str(x).strip()]))))
                            all_skus = " • ".join([f"{k} ({v} шт.)" for k, v in group['Артикул продавца'].value_counts().items()])
                            
                            invoice_grouped.append({
                                'Инвойс': clean_invoice, 
                                'Дефекты': len(group), 
                                'Поставки': supplies if supplies else "Не указаны", 
                                'Список Артикулов': all_skus
                            })
                        
                        if invoice_grouped:
                            df_invoice = pd.DataFrame(invoice_grouped).sort_values('Дефекты', ascending=False).head(15)
                            
                            inv_click_selector = alt.selection_point(name='inv_click', fields=['Инвойс'])
                            
                            invoice_chart = alt.Chart(df_invoice).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                                x=alt.X('Инвойс:N', sort='-y', title=None, axis=alt.Axis(labelAngle=-45, labelLimit=500)), 
                                y=alt.Y('Дефекты:Q', title='Количество дефектов'), 
                                color=alt.Color('Дефекты:Q', scale=alt.Scale(scheme='oranges'), legend=None), 
                                tooltip=[
                                    alt.Tooltip('Инвойс:N', title='Инвойс'), 
                                    alt.Tooltip('Дефекты:Q', title='Всего дефектов'), 
                                    alt.Tooltip('Поставки:N', title='Связанные поставки'), 
                                    alt.Tooltip('Список Артикулов:N', title='Артикулы')
                                ]
                            ).properties(height=350).add_params(inv_click_selector)
                            
                            current_inv_key = st.session_state.get('invoice_key', 0)
                            inv_event = st.altair_chart(invoice_chart, use_container_width=True, on_select="rerun", key=f"inv_chart_{current_inv_key}")
                            
                            if inv_event and hasattr(inv_event, "selection"):
                                sel_inv = inv_event.selection.get("inv_click", [])
                                if sel_inv and len(sel_inv) > 0:
                                    clicked_invoice = sel_inv[0].get('Инвойс')
                                    if clicked_invoice:
                                        show_invoice_details(clicked_invoice, df_filtered)
                        else: 
                            st.info("Нет данных по инвойсам (или у всех дефектов не указан инвойс).")
                    else:
                        st.info("Невозможно построить график: отсутствует колонка 'Инвойс'.")
                else: 
                    st.info("Данных для матрицы пока нет.")
                    
            # Вызываем фрагмент
            render_production_dashboard(df)

    except Exception as e: 
        st.error(f"Ошибка Отчета: {e}")

elif page == "Уровень PPM":
    st.title(":material/report_problem: Уровень PPM и Классификация")
    
    import numpy as np
    import plotly.graph_objects as go
    from datetime import datetime
    from sqlalchemy import text

    try:
        def update_abc_in_sql(df_to_upload):
            if engine and not df_to_upload.empty:
                try:
                    with engine.begin() as conn:
                        conn.execute(text("TRUNCATE TABLE product_classification"))
                        df_to_upload = df_to_upload[['Артикул', 'Класс ABC', 'Класс XYZ']].copy()
                        df_to_upload.columns = ['article', 'class_abc', 'class_xyz']
                        df_to_upload.to_sql('product_classification', conn, if_exists='append', index=False)
                    return True
                except Exception as e:
                    st.error(f"Ошибка сохранения в БД: {e}")
                    return False
            return False

      # ==========================================
        # 1. ТЯЖЕЛАЯ ЧАСТЬ (Выполняется мгновенно из кеша)
        # ==========================================
        with st.spinner("Синхронизация данных..."):
            df_total = build_ppm_base_dataset()
            sku_options = sorted(df_total['Артикул'].unique().tolist()) if not df_total.empty else []
            
            # ==========================================
            # 2. ИНТЕРАКТИВНЫЙ ФРАГМЕНТ (Твой UI + Генератор)
            # ==========================================
            @st.fragment
            def render_ppm_dashboard(df_full, skus):
                st.markdown("### :material/query_stats: Настройки отображения")
                f1, f2 = st.columns(2)
                
                with f1:
                    today = datetime.now().date()
                    start_current_month = today.replace(day=1)
                    date_range = st.date_input("📅 Период анализа:", [start_current_month, today], format="DD.MM.YYYY", key="ppm_date_filter")

                all_options = ['[Все артикулы]', '[Вся Группа A]', '[Вся Группа B]', '[Вся Группа C]'] + skus
                sel_skus = f2.multiselect("📦 Объекты (Таблица + График):", options=all_options, default=['[Все артикулы]'])

                active_skus = set()
                if not sel_skus or '[Все артикулы]' in sel_skus:
                    active_skus.update(skus)
                else:
                    for item in sel_skus:
                        if item.startswith('[Вся Группа'):
                            g = item.replace('[Вся Группа ', '').replace(']', '')
                            active_skus.update(df_full[df_full['ABC_Группа'] == g]['Артикул'].tolist())
                        else:
                            active_skus.add(item)
                
                filtered_sku_df = df_full[df_full['Артикул'].isin(active_skus)].copy()
                
                if len(date_range) == 2:
                    table_df = filtered_sku_df[
                        (filtered_sku_df['Месяц_ДТ'].dt.date >= date_range[0].replace(day=1)) & 
                        (filtered_sku_df['Месяц_ДТ'].dt.date <= date_range[1])
                    ].copy()
                else:
                    table_df = filtered_sku_df.copy()

                table_agg = table_df.groupby(['Артикул', 'ABC_Группа', 'Класс XYZ']).agg({'Брак':'sum', 'Заказы':'sum'}).reset_index()
                table_agg['PPM'] = np.where(table_agg['Заказы'] > 0, (table_agg['Брак'] / table_agg['Заказы']) * 1000000, 0).astype(int)
                table_agg['%'] = np.where(table_agg['Заказы'] > 0, (table_agg['Брак'] / table_agg['Заказы']) * 100, 0)
                table_agg = table_agg.sort_values(by=['ABC_Группа', 'PPM'], ascending=[True, False])

                st.markdown("### :material/query_stats: Сводка по группам")
                m1, m2, m3 = st.columns(3)

                def get_group_metrics(group_name):
                    group_data = table_agg[table_agg['ABC_Группа'] == group_name]
                    total_sku = len(group_data)
                    problem_sku = len(group_data[group_data['PPM'] > 10000])
                    sum_brk = group_data['Брак'].sum()
                    sum_ord = group_data['Заказы'].sum()
                    avg_ppm = int((sum_brk / sum_ord * 1000000)) if sum_ord > 0 else 0
                    return total_sku, problem_sku, avg_ppm

                for col, grp in zip([m1, m2, m3], ['A', 'B', 'C']):
                    t_sku, p_sku, a_ppm = get_group_metrics(grp)
                    with col:
                        st.metric(label=f"Группа {grp}", value=f"{t_sku} SKU", delta=f"{p_sku} проблемных", delta_color="inverse")
                        st.markdown(f"**Средний PPM: {a_ppm:,}**".replace(',', ' '))
                        st.caption(f"Доля проблемных SKU: {round(p_sku/t_sku*100 if t_sku > 0 else 0)}%")

                st.divider()

                col_table, col_chart = st.columns([2.5, 2], gap="large") 

                with col_table:
                    st.markdown("#### :material/table_rows: Артикулы")
                    def highlight(row): return ['background-color: #fee2e2; color: #991b1b' if row.get('PPM',0) > 10000 else ''] * len(row)
                    selection = st.dataframe(
                        table_agg.style.apply(highlight, axis=1), 
                        use_container_width=True, hide_index=True, height=450,
                        on_select="rerun", selection_mode="single-row",
                        column_config={
                            "Артикул": st.column_config.TextColumn("Артикул", width="medium"),
                            "ABC_Группа": st.column_config.TextColumn("ABC", width="small"),
                            "Класс XYZ": st.column_config.TextColumn("XYZ", width="small"),
                            "Заказы": st.column_config.NumberColumn("Заказы", format="%d", width="small"),
                            "Брак": st.column_config.NumberColumn("Брак", format="%d", width="small"),
                            "%": st.column_config.NumberColumn("%", format="%.2f", width="small"),
                            "PPM": st.column_config.NumberColumn("PPM", format="%d", width="small")
                        }
                    )
                
                # Твоя логика определения клика
                selected_indices = selection.selection.rows
                clicked_sku = table_agg.iloc[selected_indices[0]]['Артикул'] if selected_indices else None

                with col_chart:
                    chart_base_df = filtered_sku_df[filtered_sku_df['Артикул'] == clicked_sku].copy() if clicked_sku else filtered_sku_df.copy()
                    chart_title = f"Динамика: {clicked_sku}" if clicked_sku else "Динамика"
                    
                    if not chart_base_df.empty:
                        latest = chart_base_df['Месяц_ДТ'].max()
                        start = latest - pd.DateOffset(months=11)
                        chart_agg = chart_base_df[chart_base_df['Месяц_ДТ'] >= start].groupby(['Месяц_ДТ', 'Месяц_Стр', 'Source']).agg({'Брак':'sum', 'Заказы':'sum'}).reset_index()
                        chart_agg = chart_agg[chart_agg['Месяц_ДТ'] >= '2026-01-01']
                        
                        mask_hist = (chart_agg['Source'] == 'External') & (chart_agg['Месяц_ДТ'] < '2026-04-01')
                        mask_sys = (chart_agg['Source'] == 'System') & (chart_agg['Месяц_ДТ'] >= '2026-04-01')
                        chart_agg = chart_agg[mask_hist | mask_sys]
                        
                        chart_agg['PPM'] = np.where(chart_agg['Заказы'] > 0, (chart_agg['Брак'] / chart_agg['Заказы']) * 1000000, 0).astype(int)
                        chart_agg = chart_agg.sort_values('Месяц_ДТ')

                        max_val = chart_agg['PPM'].max() if not chart_agg.empty else 0
                        y_limit = max(20000, max_val * 1.15)

                        st.markdown(f"#### :material/monitoring: {chart_title}")
                        fig = go.Figure()
                        for src, clr, nm in [('External', '#f39c12', 'История'), ('System', '#3b82f6', 'Система')]:
                            curr = chart_agg[chart_agg['Source'] == src]
                            if not curr.empty:
                                fig.add_trace(go.Bar(x=curr['Месяц_Стр'], y=curr['PPM'], name=nm, marker_color=clr, text=curr['PPM'], textposition='outside'))
                        
                        fig.add_hline(y=10000, line_dash="dash", line_color="#e74c3c", annotation_text="Limit 1%")
                        
                        # ОСТАВЛЯЕМ ТОЛЬКО ЛИНИЮ БРАКА (без заказов)
                        fig.add_trace(go.Scatter(
                            x=chart_agg['Месяц_Стр'], 
                            y=chart_agg['Брак'], 
                            name='Кол-во брака', 
                            line=dict(color='#e74c3c', width=3, dash='dot'), 
                            yaxis='y2'
                        ))
                        
                        fig.update_layout(
                            barmode='overlay', 
                            height=420, 
                            margin=dict(l=0, r=0, t=20, b=0),
                            xaxis=dict(type='category', categoryorder='array', categoryarray=chart_agg['Месяц_Стр'].unique(), showgrid=False),
                            legend=dict(orientation="h", y=1.15), 
                            yaxis=dict(title="PPM", range=[0, y_limit], side='left', showgrid=False),
                            yaxis2=dict(title="Кол-во брака", overlaying='y', side='right', showgrid=True, rangemode='tozero'), 
                            hovermode="x unified"
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    else: st.info("Нет данных для графика")

              # --- ОБНОВЛЕННЫЙ БЛОК: ДЕТАЛИЗАЦИЯ И ГЕНЕРАТОР ---
                if selected_indices:
                    selected_row = table_agg.iloc[selected_indices[0]]
                    current_sku = selected_row['Артикул']
                    
                    # 1. ЗАГРУЗКА ДАННЫХ И ИНВОЙСОВ
                    try:
                        df_sys_detail = load_cached_hybrid_data()
                        if len(date_range) == 2:
                            mask_date = (df_sys_detail['Дата_ДТ'].dt.date >= date_range[0]) & \
                                        (df_sys_detail['Дата_ДТ'].dt.date <= date_range[1])
                            sku_details = df_sys_detail[(df_sys_detail['Артикул продавца'] == current_sku) & mask_date].copy()
                        else:
                            sku_details = df_sys_detail[df_sys_detail['Артикул продавца'] == current_sku].copy()
                        
                        valid_vals = ['1', '1.0', '+', 'true', 'да']
                        defect_mask = pd.Series(False, index=sku_details.index)
                        for i in range(1, 14):
                            cat_col = str(i)
                            if cat_col in sku_details.columns:
                                defect_mask |= sku_details[cat_col].astype(str).str.strip().str.lower().isin(valid_vals)
                        
                        defect_invs = sku_details.loc[defect_mask, 'Инвойс'].dropna().unique()
                        valid_invs = [str(inv).strip() for inv in defect_invs if str(inv).strip() not in ['nan', 'None', '', 'Не указан', '0', '0.0']]
                        auto_invoices = ", ".join(valid_invs)
                    except:
                        sku_details = pd.DataFrame()
                        auto_invoices = ""

                    # 2. ЕДИНАЯ ПРЕДЗАГРУЗКА ФОТОГРАФИЙ (Для UI и Экселя одновременно)
                    prefetched_photos = {} # Словарь для хранения найденных фото
                    seen_photos_global = set() # Чтобы фото не дублировались
                    
                    if not sku_details.empty:
                        for i in range(1, 14):
                            cat_col = str(i)
                            if cat_col in sku_details.columns:
                                cat_mask = sku_details[cat_col].astype(str).str.strip().str.lower().isin(valid_vals)
                                if cat_mask.any():
                                    srids = sku_details.loc[cat_mask, 'SRID'].dropna().unique().tolist()
                                    if srids:
                                        import random
                                        random.shuffle(srids)
                                        urls = []
                                        for srid in srids:
                                            try:
                                                p_raw, _ = get_media_for_srid(srid)
                                                if p_raw:
                                                    groups = p_raw.split()
                                                    if groups:
                                                        for group in groups:
                                                            if "|" in group:
                                                                s3_url, wb_url = group.split("|", 1)
                                                            else:
                                                                s3_url = wb_url = group
                                                            
                                                            if s3_url.startswith("//"): s3_url = "https:" + s3_url
                                                            if wb_url.startswith("//"): wb_url = "https:" + wb_url
                                                            
                                                            if wb_url not in seen_photos_global:
                                                                seen_photos_global.add(wb_url)
                                                                # Сохраняем и превью (для скорости UI) и оригинал (для экселя)
                                                                urls.append((s3_url, wb_url))
                                                                break 
                                            except: continue
                                            if len(urls) >= 6: break # Берем до 6 фото для UI
                                        prefetched_photos[i] = urls

                    # 3. ПОДГОТОВКА ДАННЫХ ДЛЯ ЭКСЕЛЯ (Берем из предзагруженного кэша)
                    combined_issues = []
                    photo_payload = {}
                    
                    if not sku_details.empty and 'CLAIM_CATEGORIES_LOGIC' in globals():
                        for group_key, config in CLAIM_CATEGORIES_LOGIC.items():
                            group_mask = pd.Series(False, index=sku_details.index)
                            group_urls = []
                            for cid in config['ids']:
                                if str(cid) in sku_details.columns:
                                    group_mask |= sku_details[str(cid)].astype(str).str.strip().str.lower().isin(valid_vals)
                                
                                # Забираем уже найденные фото из кэша
                                for _, wb_url in prefetched_photos.get(cid, []):
                                    if wb_url not in group_urls:
                                        group_urls.append(wb_url)
                            
                            group_count = sku_details[group_mask].shape[0]
                            if group_count > 0:
                                combined_issues.append(f"{config['ru']} ({group_count})")
                                # Передаем только первые 3 фотки в Excel
                                photo_payload[config['ru']] = group_urls[:3]

                    # 4. ВИЗУАЛЬНАЯ ДЕТАЛИЗАЦИЯ (Интерфейс)
                    st.markdown("<hr style='margin: 2em 0; border: none; border-bottom: 1px solid #cbd5e1;'/>", unsafe_allow_html=True)
                    st.subheader(f":material/troubleshoot: Визуальная детализация брака: {current_sku}")
                    
                    if not sku_details.empty:
                        has_data = False
                        for i in range(1, 14):
                            cat_col = str(i)
                            if cat_col in sku_details.columns:
                                cat_mask = sku_details[cat_col].astype(str).str.strip().str.lower().isin(valid_vals)
                                count = sku_details[cat_mask].shape[0]
                                if count > 0:
                                    has_data = True
                                    cat_name = CATEGORIES.get(i, f"Категория {i}") if 'CATEGORIES' in globals() else f"Категория {i}"
                                    
                                    with st.container():
                                        st.markdown("<hr style='margin: 1em 0; border: none; border-bottom: 1px solid #e2e8f0;'/>", unsafe_allow_html=True)
                                        c_text, c_media = st.columns([1.2, 1])
                                        
                                        cat_specific_df = sku_details[cat_mask]
                                        invs = cat_specific_df['Инвойс'].dropna().unique()
                                        invs_clean = [str(inv).strip() for inv in invs if str(inv).strip() not in ['nan', 'None', '', 'Не указан']]
                                        inv_str = ", ".join(invs_clean) if invs_clean else "Не указан"

                                        with c_text:
                                            st.markdown(f"#### :material/report: {cat_name}")
                                            st.markdown(f"**Количество:** {count} шт.")
                                            st.markdown(f"**Инвойсы:** {inv_str}")
                                        
                                        with c_media:
                                            # Отрисовываем фото мгновенно из кэша
                                            urls = prefetched_photos.get(i, [])
                                            if urls:
                                                html_imgs = '<div style="display: flex; flex-wrap: wrap; gap: 8px;">'
                                                for s3_url, wb_url in urls:
                                                    html_imgs += f'<a href="{wb_url}" target="_blank"><img src="{s3_url}" class="photo-zoom" style="width: 70px; height: 70px; object-fit: cover; border-radius: 6px; border: 1px solid #e2e8f0; transition: transform 0.2s;"></a>'
                                                html_imgs += '</div>'
                                                st.markdown(html_imgs, unsafe_allow_html=True)
                                            else:
                                                st.caption("Нет уникального фото")
                        if not has_data:
                            st.info("Нет детализированных данных по категориям за выбранный период.")
                    else:
                        st.warning("За выбранный период данных по этому артикулу не найдено.")

                    # 5. ФОРМА ПРЕТЕНЗИИ И СКАЧИВАНИЕ
                    st.markdown("<hr style='margin: 2em 0; border: none; border-bottom: 1px solid #cbd5e1;'/>", unsafe_allow_html=True)
                    st.subheader(f":material/edit_document: Формирование рекламации: {current_sku}")
                    
                    auto_desc_ru = " / ".join(combined_issues) if combined_issues else "Дефекты не обнаружены"
                    auto_desc_cn = " / ".join([CLAIM_CATEGORIES_LOGIC[k]['cn'] for k in CLAIM_CATEGORIES_LOGIC if CLAIM_CATEGORIES_LOGIC[k]['ru'] in [x.split(' (')[0] for x in combined_issues]]) if 'CLAIM_CATEGORIES_LOGIC' in globals() and combined_issues else ""

                    cl1, cl2, cl3 = st.columns(3)
                    with cl1:
                        sup = st.text_input("Завод", value="Уточняется", key="cl_sup")
                        num = st.text_input("Номер Рекламационного Акта", value="", placeholder="Введите номер...", key="cl_num")
                    with cl2:
                        inv_val = st.text_input("Инвойс (Invoice)", value=auto_invoices, key="cl_inv")
                        period_val = f"{date_range[0].strftime('%d.%m.%Y')} - {date_range[1].strftime('%d.%m.%Y')}" if len(date_range)==2 else "01.04.2026 - 30.04.2026"
                        per = st.text_input("Период (Period)", value=period_val, key="cl_per")
                    with cl3:
                        d_ru = st.text_area("Описание дефектов (RU)", value=auto_desc_ru, key="cl_d_ru")
                        d_cn = st.text_area("Описание дефектов (CN)", value=auto_desc_cn, key="cl_d_cn")

                    c_data = {
                        "number": num, 
                        "date": datetime.now().strftime("%Y-%m-%d"), 
                        "supplier": sup,
                        "period": per, 
                        "invoice": inv_val, 
                        "sku": current_sku,
                        "name": "Наименование товара", 
                        "name_cn": "产品", 
                        "defects": selected_row['Брак'],
                        "ppm_pct": round(selected_row['%'], 2), 
                        "desc_ru": d_ru, 
                        "desc_cn": d_cn,
                        "cause_ru": CLAIM_CATEGORIES_LOGIC.get('Damage', {}).get('cause_ru', 'Нарушение при производстве') if 'CLAIM_CATEGORIES_LOGIC' in globals() else 'Нарушение', 
                        "cause_cn": CLAIM_CATEGORIES_LOGIC.get('Damage', {}).get('cause_cn', '生产过程异常') if 'CLAIM_CATEGORIES_LOGIC' in globals() else '异常',
                        "photo_groups": photo_payload 
                    }
                    
                    st.download_button(
                        label=f"📥 Скачать Рекламацию для {current_sku}", 
                        data=generate_claim_from_template(c_data, chart_fig=fig), 
                        file_name=f"RA_{num if num else 'draft'}_{current_sku}.xlsx", 
                        type="primary", 
                        use_container_width=True
                    )

            render_ppm_dashboard(df_total, sku_options)

    except Exception as e:
        st.error(f"Ошибка PPM: {e}")
        
elif page == "Рейтинг товаров":
    st.title(":material/star_rate: Управление качеством и рейтингами")
    
    if engine:
        try:
            # Загрузка данных
            query = """
                SELECT 
                    r.date as "Дата", 
                    r.supplier_article as "Артикул", 
                    ROUND(CAST(r.average_rating AS NUMERIC), 1) as "Рейтинг", 
                    r.review_count as "Отзывы"
                FROM wb_ratings r
                ORDER BY r.date ASC
            """
            df_ratings = pd.read_sql(query, engine)
            
            if not df_ratings.empty:
                df_ratings['Дата'] = pd.to_datetime(df_ratings['Дата'])
                latest_overall_date = df_ratings['Дата'].max()

                # --- 1. БЛОК ФИЛЬТРАЦИИ ---
                st.markdown("### :material/filter_alt: Настройки")
                col_f1, col_f2, col_f3 = st.columns([1.5, 2, 1])
                
                with col_f1:
                    start_default = (latest_overall_date - pd.Timedelta(days=7)).date()
                    date_range = st.date_input(
                        "Период анализа:", 
                        [start_default, latest_overall_date.date()],
                        format="DD.MM.YYYY"
                    )

                # Шаг 1: Первичная фильтрация по дате
                if len(date_range) == 2:
                    df_date_filtered = df_ratings[
                        (df_ratings['Дата'].dt.date >= date_range[0]) & 
                        (df_ratings['Дата'].dt.date <= date_range[1])
                    ].copy()
                else:
                    df_date_filtered = df_ratings.copy()

                # Находим последний день в выбранном периоде
                max_date_in_period = df_date_filtered['Дата'].max() if not df_date_filtered.empty else latest_overall_date
                last_day_data = df_date_filtered[df_date_filtered['Дата'] == max_date_in_period]

                # --- ПРЕДВАРИТЕЛЬНЫЙ РАСЧЕТ ДЛЯ СВОДКИ И ГРУППИРОВОК ---
                total_skus = len(last_day_data)
                avg_rating = round(last_day_data['Рейтинг'].mean(), 1) if not last_day_data.empty else 0.0
                
                # Собираем список артикулов "Ниже среднего"
                below_avg_skus = last_day_data[last_day_data['Рейтинг'] < avg_rating]['Артикул'].tolist() if not last_day_data.empty else []
                critical_count = len(below_avg_skus)

                # Собираем список артикулов "Упали в рейтинге" (Сравнение с предыдущим днем)
                all_dates = sorted(df_ratings['Дата'].dt.date.unique())
                current_date_idx = all_dates.index(max_date_in_period.date()) if max_date_in_period.date() in all_dates else -1
                
                dropped_skus = []
                if current_date_idx > 0:
                    prev_date = all_dates[current_date_idx - 1]
                    prev_day_data = df_ratings[df_ratings['Дата'].dt.date == prev_date]
                    
                    merged = pd.merge(last_day_data, prev_day_data, on='Артикул', suffixes=('_curr', '_prev'))
                    dropped_skus = merged[merged['Рейтинг_curr'] < merged['Рейтинг_prev']]['Артикул'].tolist()
                
                dropped_count = len(dropped_skus)

                with col_f2:
                    all_available_skus = sorted(df_ratings['Артикул'].unique().tolist())
                    
                    # Добавляем новые группировки в начало списка
                    sku_options = [
                        "[ВСЕ АРТИКУЛЫ]", 
                        "[НИЖЕ СРЕДНЕГО]", 
                        "[УПАЛИ В РЕЙТИНГЕ]"
                    ] + all_available_skus
                    
                    selected_skus = st.multiselect(
                        "Артикулы:", 
                        sku_options, 
                        default=["[ВСЕ АРТИКУЛЫ]"],
                        key="sku_selector_main"
                    )

                with col_f3:
                    rating_options = ["Все", "5.0", "4.9", "4.8", "4.7", "4.6", "4.5", "Ниже 4.5"]
                    sel_rating = st.selectbox("Рейтинг на конец периода:", rating_options)

                # --- 2. СВОДКА СОСТОЯНИЯ (ОТВЯЗАНА ОТ ФИЛЬТРОВ) ---
                st.markdown("### :material/analytics: Сводка")
                
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("📦 Всего товаров", f"{total_skus} SKU")
                m2.metric("⚠️ Ниже среднего", f"{critical_count} SKU", help=f"Товары с рейтингом ниже {avg_rating:.1f}")
                m3.metric("⭐ Средний балл", f"{avg_rating:.1f}")
                m4.metric("📉 Упали в рейтинге", f"{dropped_count} SKU", help="Количество товаров, чей рейтинг снизился по сравнению с предыдущим днем")

                st.divider()

                # --- 3. ПРИМЕНЕНИЕ ФИЛЬТРОВ ДЛЯ МАТРИЦЫ И ГРАФИКОВ ---
                
                # Обрабатываем выбор артикулов и группировок
                if "[ВСЕ АРТИКУЛЫ]" in selected_skus or not selected_skus: 
                    active_skus = all_available_skus
                else:
                    active_skus_set = set()
                    if "[НИЖЕ СРЕДНЕГО]" in selected_skus:
                        active_skus_set.update(below_avg_skus)
                    if "[УПАЛИ В РЕЙТИНГЕ]" in selected_skus:
                        active_skus_set.update(dropped_skus)
                    
                    # Добавляем индивидуально выбранные артикулы (исключая наши теги в скобках)
                    individual_skus = [s for s in selected_skus if not s.startswith("[")]
                    active_skus_set.update(individual_skus)
                    
                    active_skus = list(active_skus_set)
                
                # Применяем фильтр по конкретной оценке
                if sel_rating != "Все":
                    if sel_rating == "Ниже 4.5":
                        valid_skus = last_day_data[last_day_data['Рейтинг'] < 4.5]['Артикул'].tolist()
                    else:
                        target_val = float(sel_rating)
                        valid_skus = last_day_data[last_day_data['Рейтинг'] == target_val]['Артикул'].tolist()
                    
                    active_skus = [s for s in active_skus if s in valid_skus]

                # Формируем финальный датафрейм
                df_filtered = df_date_filtered[df_date_filtered['Артикул'].isin(active_skus)].copy()

                if not df_filtered.empty:
                    # --- 4. ТЕПЛОВАЯ МАТРИЦА ---
                    st.markdown("#### :material/grid_view: Матрица состояния")
                    
                    df_pivot = df_filtered.pivot(index="Артикул", columns="Дата", values="Рейтинг").sort_index()
                    df_pivot.columns = [d.strftime('%d.%m.%Y') for d in df_pivot.columns]

                    import plotly.graph_objects as go
                    
                    custom_colorscale = [
                        [0, '#ef4444'],    # Красный (до 3.5)
                        [0.8, '#fef08a'],  # Желтый начинается с 4.7
                        [1, '#22c55e']     # Зеленый (5.0)
                    ]

                    fig_heat = go.Figure(data=go.Heatmap(
                        z=df_pivot.values,
                        x=df_pivot.columns,
                        y=df_pivot.index,
                        colorscale=custom_colorscale,
                        zmin=3.5, zmax=5.0,
                        xgap=2, ygap=2,
                        colorbar=dict(title="Рейтинг"),
                        hovertemplate="Артикул: %{y}<br>Дата: %{x}<br>Рейтинг: %{z:.1f}<extra></extra>"
                    ))
                    
                    fig_heat.update_layout(
                        height=max(350, len(df_pivot) * 25),
                        margin=dict(l=0, r=0, t=10, b=0),
                        xaxis=dict(tickangle=-45, showgrid=False),
                        yaxis=dict(showgrid=False)
                    )
                    st.plotly_chart(fig_heat, use_container_width=True)

                    st.divider()

                    # --- 5. ДЕТАЛЬНАЯ ДИНАМИКА ---
                    st.markdown("#### :material/stacks: Индивидуальные графики")
                    
                    import plotly.express as px
                    
                    num_skus = len(df_filtered['Артикул'].unique())
                    num_rows = (num_skus - 1) // 3 + 1
                    
                    safe_spacing = min(0.04, 0.9 / (num_rows - 1)) if num_rows > 1 else 0.0
                    
                    fig_facet = px.line(
                        df_filtered, 
                        x="Дата", 
                        y="Рейтинг", 
                        facet_col="Артикул", 
                        facet_col_wrap=3,
                        markers=True,
                        color="Артикул",
                        template="plotly_white",
                        facet_row_spacing=safe_spacing,
                        hover_data={"Дата": "|%d.%m.%Y", "Рейтинг": ":.1f"}
                    )
                    
                    fig_facet.update_yaxes(range=[3.0, 5.1], dtick=0.5, showgrid=True, gridcolor="#f8fafc")
                    fig_facet.update_xaxes(title=None, tickformat="%d.%m")
                    
                    fig_facet.update_layout(
                        height=max(300, 250 * num_rows),
                        showlegend=False,
                        margin=dict(l=40, r=20, t=40, b=40)
                    )
                    
                    fig_facet.for_each_annotation(lambda a: a.update(text=f"📦 {a.text.split('=')[-1]}"))
                    
                    st.plotly_chart(fig_facet, use_container_width=True)
                else:
                    st.info("Нет данных по выбранным фильтрам (либо ни один товар не попал в выбранную группу).")
        except Exception as e:
            st.error(f"Ошибка блока Рейтингов: {e}")
        
elif page == "Системный Журнал":
    st.title(":material/receipt_long: Системный Журнал (SQL Edition)")
    
    if st.button("🔄 Обновить журнал"): 
        st.rerun()

    if engine:
        try:
            # Читаем последние 200 записей
            query = "SELECT created_at as \"Дата\", action as \"Действие\", status as \"Статус\", details as \"Детали\" FROM system_logs ORDER BY created_at DESC LIMIT 200"
            df_logs = pd.read_sql(query, engine)

            if not df_logs.empty:
                # Форматируем дату для удобства
                df_logs['Дата'] = pd.to_datetime(df_logs['Дата']).dt.strftime('%d.%m.%Y %H:%M:%S')

                def color_status(val):
                    color = '#22c55e' if val == 'INFO' or val == 'SUCCESS' else '#ef4444' if 'ERR' in str(val) else '#f59e0b'
                    return f'color: {color}; font-weight: bold;'

                st.dataframe(
                    df_logs.style.map(color_status, subset=['Статус']), 
                    width="stretch", 
                    height=600,
                    hide_index=True
                )
            else:
                st.info("Журнал пуст. Все события будут появляться здесь после запуска тегирования.")
        except Exception as e:
            st.error(f"Ошибка чтения логов из SQL: {e}")
    else:
        st.warning("База данных не подключена. Проверьте DB_URL.")
