"""
Яндекс Маркет воркер — фоновая синхронизация данных.
Поля и структура проверены через прямые curl-тесты 2026-06-09.

Запуск (PYTHONPATH=/root/cx-dashboard):
  python3 ym_worker.py hourly   — заказы, возвраты, отзывы, вопросы  (cron: 0 */2 * * *)
  python3 ym_worker.py daily    — ассортимент + цены + запуск фин-отчётов (cron: 0 3 * * *)
  python3 ym_worker.py reports  — опрос статуса и скачивание отчётов (cron: 0 5,9,13 * * *)

Переменные в .env:
  YM_API_KEY, YM_CAMPAIGN_ID, YM_BUSINESS_ID, DATABASE_URL_LOCAL
"""

import csv
import io
import logging
import os
import re
import sys
import time
import zipfile
from datetime import date, datetime, timedelta, timezone

# Директория для хранения фото возвратов ЯМ (nginx раздаёт /static/)
YM_PHOTO_DIR = "/root/cx-dashboard/static/ym_returns"

# Только возвраты начиная с этой даты
YM_RETURNS_FROM = datetime(2026, 1, 1, tzinfo=timezone.utc)

import requests
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env'))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [YM] %(levelname)s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("ym_worker")

YM_API_KEY     = os.getenv("YM_API_KEY", "").strip()
YM_CAMPAIGN_ID = os.getenv("YM_CAMPAIGN_ID", "").strip()
YM_BUSINESS_ID = os.getenv("YM_BUSINESS_ID", "").strip()
DATABASE_URL   = os.getenv(
    "DATABASE_URL_LOCAL",
    "postgresql://db_user:RDB_r6o_BA0qSlVVGjb_2026@127.0.0.1:5432/cx_dashboard",
)

BASE   = "https://api.partner.market.yandex.ru"
engine = create_engine(DATABASE_URL, pool_pre_ping=True)

# Маппинг refundStatus ЯМ → читаемый статус
YM_REFUND_STATUS_MAP = {
    "WAITING_FOR_DECISION":  "На рассмотрении",
    "REFUND_IN_PROGRESS":    "Одобрено",
    "REFUNDED":              "Одобрено",
    "WAITING_FOR_RETURN":    "На рассмотрении",
    "RETURNED":              "На рассмотрении",
    "NOT_RETURNED":          "Отказ",
    "FAILED":                "Отказ",
}


def _hdr():
    return {"Api-Key": YM_API_KEY, "Content-Type": "application/json"}


class RateLimitError(Exception):
    pass


def _request(method, path, **kwargs):
    """HTTP-запрос. При 420 — сразу RateLimitError (caller пропускает задачу)."""
    r = getattr(requests, method)(f"{BASE}{path}", headers=_hdr(), timeout=30, **kwargs)
    if r.status_code == 420:
        raise RateLimitError(f"420 rate limit на {path}. Попробуем на следующем cron-запуске.")
    r.raise_for_status()
    return r.json()


def _get(path, params=None):
    return _request("get", path, params=params or {})


def _post(path, body):
    return _request("post", path, json=body)


def _log_db(action, status, details=""):
    try:
        with engine.begin() as conn:
            conn.execute(
                text("INSERT INTO system_logs (action, status, details) VALUES (:a,:s,:d)"),
                {"a": action, "s": status, "d": str(details)[:2000]},
            )
    except Exception:
        pass


def _parse_ym_date(s):
    """Парсит дату ЯМ: ISO ('2025-05-29T22:54:05+03:00' или '2025-05-29 22:54:05')
    и нестандартный формат API ('29-05-2025 22:54:05')."""
    if not s:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:len(fmt)+5].strip(), fmt).isoformat()
        except ValueError:
            continue
    return s  # вернём как есть если не распарсилось


# ── 1. Заказы ────────────────────────────────────────────────────────────────
# ym_orders: UNIQUE (order_id, supplier_article) — один заказ может иметь несколько SKU
def sync_orders():
    log.info("Синхронизация заказов ЯМ...")
    since = (datetime.now(timezone.utc) - timedelta(hours=48)).strftime("%d-%m-%Y")
    page_token = None
    total = 0

    while True:
        params = {"fromDate": since, "limit": 50}
        if page_token:
            params["page_token"] = page_token

        data   = _get(f"/v2/campaigns/{YM_CAMPAIGN_ID}/orders", params=params)
        orders = data.get("orders", [])
        if not orders:
            break

        with engine.begin() as conn:
            for o in orders:
                for item in (o.get("items") or [{}]):
                    conn.execute(text("""
                        INSERT INTO ym_orders
                            (order_id, status, supplier_article, sku_name,
                             created_at, price, buyer_total, currency, region_name)
                        VALUES
                            (:order_id, :status, :article, :name,
                             :created_at, :price, :buyer_total, :currency, :region)
                        ON CONFLICT (order_id, supplier_article) DO UPDATE SET
                            status    = EXCLUDED.status,
                            synced_at = NOW()
                    """), {
                        "order_id":    o["id"],
                        "status":      o.get("status"),
                        "article":     item.get("offerId"),
                        "name":        item.get("offerName"),
                        "created_at":  _parse_ym_date(o.get("creationDate")),
                        "price":       item.get("price"),
                        "buyer_total": o.get("buyerTotal"),
                        "currency":    o.get("currency", "RUR"),
                        "region":      o.get("deliveryRegion", {}).get("name"),
                    })
                    total += 1

        # paging.nextPageToken (не pager!)
        page_token = (data.get("paging") or {}).get("nextPageToken")
        if not page_token:
            break
        time.sleep(0.3)

    log.info(f"Заказы ЯМ: {total} строк")
    _log_db("ym_sync_orders", "success", f"{total} строк заказов")


def sync_orders_backfill(from_date_str: str = "2025-01-01"):
    """Исторический сбор заказов ЯМ по 30-дневным окнам начиная с from_date_str."""
    log.info(f"Бэкфилл заказов ЯМ с {from_date_str}...")
    start = datetime.strptime(from_date_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    end   = datetime.now(timezone.utc)
    total = 0
    window_start = start

    while window_start < end:
        window_end = min(window_start + timedelta(days=29), end)
        from_str = window_start.strftime("%d-%m-%Y")
        to_str   = window_end.strftime("%d-%m-%Y")
        log.info(f"  Окно: {from_str} → {to_str}")

        page_token = None
        while True:
            params = {"fromDate": from_str, "toDate": to_str, "limit": 50}
            if page_token:
                params["page_token"] = page_token

            data   = _get(f"/v2/campaigns/{YM_CAMPAIGN_ID}/orders", params=params)
            orders = data.get("orders", [])
            if not orders:
                break

            with engine.begin() as conn:
                for o in orders:
                    for item in (o.get("items") or [{}]):
                        conn.execute(text("""
                            INSERT INTO ym_orders
                                (order_id, status, supplier_article, sku_name,
                                 created_at, price, buyer_total, currency, region_name)
                            VALUES
                                (:order_id, :status, :article, :name,
                                 :created_at, :price, :buyer_total, :currency, :region)
                            ON CONFLICT (order_id, supplier_article) DO UPDATE SET
                                status    = EXCLUDED.status,
                                synced_at = NOW()
                        """), {
                            "order_id":    o["id"],
                            "status":      o.get("status"),
                            "article":     item.get("offerId"),
                            "name":        item.get("offerName"),
                            "created_at":  _parse_ym_date(o.get("creationDate")),
                            "price":       item.get("price"),
                            "buyer_total": o.get("buyerTotal"),
                            "currency":    o.get("currency", "RUR"),
                            "region":      o.get("deliveryRegion", {}).get("name"),
                        })
                        total += 1

            page_token = (data.get("paging") or {}).get("nextPageToken")
            if not page_token:
                break
            time.sleep(0.3)

        time.sleep(0.5)
        window_start = window_end + timedelta(days=1)

    log.info(f"Бэкфилл заказов ЯМ завершён: {total} строк")
    _log_db("ym_orders_backfill", "success", f"{total} строк за период {from_date_str}→сегодня")


# ── 2. Возвраты ──────────────────────────────────────────────────────────────
# Структура (проверено curl 2026-06-09):
#   returnType: "RETURN" (реальный) или "UNREDEEMED" (не выкупил из ПВЗ)
#   items[].shopSku                      — артикул продавца
#   items[].decisions[].returnItemId     — itemId для endpoint фото
#   items[].decisions[].comment          — текст от покупателя ✓
#   items[].decisions[].images[]         — хэши фотографий ✓
#   refundStatus                         — статус решения по возврату
#
# Логика фильтрации:
#   • только с 01.01.2026
#   • пропускаем UNREDEEMED (не выкупы) — нет текста/фото
#   • сохраняем только те, у кого ЕСТЬ комментарий ИЛИ фотографии
#   • cat_1..cat_13 оставляем NULL — заполнит ИИ-тегирование

def _download_ym_photo(order_id, return_id, item_id, image_hash):
    """
    Скачивает фото возврата ЯМ и сохраняет в статичную папку.
    Возвращает URL-путь для фронтенда или None при ошибке.
    Если файл уже скачан — возвращает путь без повторной загрузки.

    Ретрай: до 3 попыток с нарастающей паузой на 420 (rate limit) и сетевые
    ошибки — иначе при массовой синхронизации фото молча терялись.
    404 — фото действительно нет, ретрай не нужен.
    """
    os.makedirs(YM_PHOTO_DIR, exist_ok=True)
    safe_hash = str(image_hash)[:40]
    filename  = f"{return_id}_{safe_hash[:16]}.jpg"
    save_path = os.path.join(YM_PHOTO_DIR, filename)

    # Уже скачано — сразу возвращаем путь
    if os.path.exists(save_path):
        return f"/static/ym_returns/{filename}"

    url = (
        f"{BASE}/v2/campaigns/{YM_CAMPAIGN_ID}"
        f"/orders/{order_id}/returns/{return_id}"
        f"/decision/{item_id}/image/{image_hash}"
    )
    for attempt in range(3):
        try:
            r = requests.get(url, headers=_hdr(), timeout=20)
            if r.status_code == 404:
                return None
            if r.status_code == 420:
                # rate limit — ждём и пробуем снова
                wait = 2 * (attempt + 1)
                log.warning(f"    420 на фото {safe_hash[:8]}…, ждём {wait}с (попытка {attempt+1}/3)")
                time.sleep(wait)
                continue
            r.raise_for_status()
            with open(save_path, "wb") as f:
                f.write(r.content)
            log.debug(f"    Фото сохранено: {filename}")
            return f"/static/ym_returns/{filename}"
        except Exception as e:
            log.warning(f"    Ошибка скачивания фото {safe_hash[:8]}… (попытка {attempt+1}/3): {e}")
            time.sleep(1.5 * (attempt + 1))
    return None


def sync_returns():
    log.info("Синхронизация возвратов ЯМ...")
    page_token = None
    scanned = 0
    new_rows = 0   # новые записи (INSERT)
    updated  = 0   # обновлённые записи (UPDATE)
    skipped_unredeemed = 0
    skipped_no_content = 0
    skipped_old        = 0

    while True:
        params = {"limit": 50}
        if page_token:
            params["page_token"] = page_token

        data    = _get(f"/v2/campaigns/{YM_CAMPAIGN_ID}/returns", params=params)
        result  = data.get("result", {})
        returns = result.get("returns", [])
        if not returns:
            break

        for r in returns:
            scanned += 1
            return_id   = r.get("id")
            order_id    = r.get("orderId")
            return_type = r.get("returnType", "")

            created_str = r.get("creationDate") or ""
            try:
                created_dt  = datetime.fromisoformat(created_str)
                created_utc = created_dt.astimezone(timezone.utc)
            except Exception:
                created_dt  = None
                created_utc = None

            if created_utc and created_utc < YM_RETURNS_FROM:
                skipped_old += 1
                continue

            if return_type == "UNREDEEMED":
                skipped_unredeemed += 1
                continue

            items = r.get("items") or []
            if not items:
                skipped_no_content += 1
                continue

            article        = ""
            comments       = []
            photo_urls     = []
            reason_type    = None
            subreason_type = None

            for item in items:
                if not article:
                    article = (item.get("shopSku") or "").strip()
                for decision in (item.get("decisions") or []):
                    if reason_type is None:
                        reason_type    = decision.get("reasonType")
                        subreason_type = decision.get("subreasonType")
                    c = (decision.get("comment") or "").strip()
                    if c and c not in comments:
                        comments.append(c)
                    for img_hash in (decision.get("images") or []):
                        url = _download_ym_photo(order_id, return_id,
                                                 decision.get("returnItemId"), img_hash)
                        if url:
                            photo_urls.append(url)
                        time.sleep(0.15)

            comment = "\n\n".join(comments)

            # Сохраняем даже без текста/фото — reason_type сам по себе информативен.
            # Пустые без причины всё равно пропускаем.
            if not comment and not photo_urls and not reason_type:
                skipped_no_content += 1
                continue

            photos_str    = " ".join(photo_urls) if photo_urls else None
            refund_status = r.get("refundStatus") or r.get("shipmentStatus") or ""
            status_ru     = YM_REFUND_STATUS_MAP.get(refund_status, "На рассмотрении")

            try:
                with engine.begin() as conn:
                    # xmax = 0 → новая строка (INSERT), иначе UPDATE
                    row = conn.execute(text("""
                        INSERT INTO ym_returns
                            (return_id, order_id, supplier_article,
                             return_type, status, status_ru,
                             created_at, updated_at,
                             return_comment, photos,
                             reason_type, subreason_type)
                        VALUES
                            (:rid, :oid, :article,
                             :rtype, :status, :status_ru,
                             :created, :updated,
                             :comment, :photos,
                             :reason, :subreason)
                        ON CONFLICT (return_id) DO UPDATE SET
                            status         = EXCLUDED.status,
                            status_ru      = EXCLUDED.status_ru,
                            updated_at     = EXCLUDED.updated_at,
                            return_comment = COALESCE(EXCLUDED.return_comment, ym_returns.return_comment),
                            photos         = COALESCE(EXCLUDED.photos, ym_returns.photos),
                            reason_type    = COALESCE(EXCLUDED.reason_type, ym_returns.reason_type),
                            subreason_type = COALESCE(EXCLUDED.subreason_type, ym_returns.subreason_type),
                            synced_at      = NOW()
                        RETURNING (xmax = 0) AS is_new
                    """), {
                        "rid":       return_id,
                        "oid":       order_id,
                        "article":   article,
                        "rtype":     return_type,
                        "status":    refund_status,
                        "status_ru": status_ru,
                        "created":   created_dt,
                        "updated":   r.get("updateDate"),
                        "comment":   comment or None,
                        "photos":    photos_str,
                        "reason":    reason_type,
                        "subreason": subreason_type,
                    }).fetchone()

                if row and row[0]:
                    new_rows += 1
                    log.info(f"  NEW {return_id} | {article} | текст: {'✓' if comment else '–'} | фото: {len(photo_urls)} | {status_ru}")
                else:
                    updated += 1

            except Exception as e:
                log.error(f"  Ошибка записи возврата {return_id}: {e}")

        page_token = result.get("paging", {}).get("nextPageToken")
        if not page_token:
            break
        time.sleep(0.5)

    summary = (
        f"просмотрено {scanned}, новых {new_rows}, обновлено {updated} | "
        f"пропущено: до2026={skipped_old} unredeemed={skipped_unredeemed} без-контента={skipped_no_content}"
    )
    log.info(f"Возвраты ЯМ: {summary}")
    _log_db("ym_sync_returns", "success", summary)


# ── 3. Отзывы (VOC) ─────────────────────────────────────────────────────────
# Реальная структура (проверено curl):
#   feedbackId  (не id!)
#   identifiers.offerId — артикул
#   statistics.rating   — оценка (не grades.product!)
#   description.advantages / description.disadvantages
#   Пагинация: тело {"pageToken": ..., "pageSize": 100} — НЕ вложено в "pagination"!
def sync_feedbacks():
    log.info("Синхронизация отзывов ЯМ...")
    page_token = None
    total = 0

    while True:
        body = {"pageSize": 100}
        if page_token:
            body["pageToken"] = page_token

        data  = _post(f"/v2/businesses/{YM_BUSINESS_ID}/goods-feedback", body)
        items = data.get("result", {}).get("feedbacks", [])
        if not items:
            break

        with engine.begin() as conn:
            for fb in items:
                desc    = fb.get("description") or {}
                stats   = fb.get("statistics") or {}
                idents  = fb.get("identifiers") or {}
                conn.execute(text("""
                    INSERT INTO ym_feedbacks
                        (feedback_id, supplier_article, created_date, valuation,
                         pro_text, contra_text, comment)
                    VALUES
                        (:fid, :article, :created_date, :val,
                         :pro, :contra, :comment)
                    ON CONFLICT (feedback_id) DO NOTHING
                """), {
                    "fid":         str(fb.get("feedbackId", "")),
                    "article":     str(idents.get("offerId") or ""),
                    "created_date": fb.get("createdAt"),
                    "val":         stats.get("rating"),
                    "pro":         desc.get("advantages", ""),
                    "contra":      desc.get("disadvantages", ""),
                    "comment":     desc.get("comment", ""),
                })
                total += 1

        page_token = data.get("result", {}).get("paging", {}).get("nextPageToken")
        if not page_token:
            break
        time.sleep(2)  # пауза между страницами — осторожно с rate limit

    log.info(f"Отзывы ЯМ: {total}")
    _log_db("ym_sync_feedbacks", "success", f"{total} отзывов")


# ── 4. Вопросы ───────────────────────────────────────────────────────────────
# Метод POST (не GET!)
# questionIdentifiers.id и questionIdentifiers.offerId — не просто id и offerId
# Ответы: questions[].answers[].text (если есть)
def sync_questions():
    log.info("Синхронизация вопросов ЯМ...")
    page_token = None
    total = 0

    while True:
        body = {"pageSize": 100}
        if page_token:
            body["pageToken"] = page_token

        data      = _post(f"/v1/businesses/{YM_BUSINESS_ID}/goods-questions", body)
        result    = data.get("result", {})
        questions = result.get("questions", [])
        if not questions:
            break

        with engine.begin() as conn:
            for q in questions:
                qid     = q.get("questionIdentifiers", {})
                answers = q.get("answers") or []
                answer_text = answers[0].get("text") if answers else None
                state   = "ANSWERED" if answer_text else "UNANSWERED"

                conn.execute(text("""
                    INSERT INTO ym_questions
                        (question_id, supplier_article, created_date, text, answer_text, state)
                    VALUES
                        (:qid, :article, :created_date, :text, :answer, :state)
                    ON CONFLICT (question_id) DO UPDATE SET
                        answer_text = EXCLUDED.answer_text,
                        state       = EXCLUDED.state
                """), {
                    "qid":         str(qid.get("id", "")),
                    "article":     str(qid.get("offerId") or ""),
                    "created_date": q.get("createdAt"),
                    "text":        q.get("text", ""),
                    "answer":      answer_text,
                    "state":       state,
                })
                total += 1

        page_token = result.get("paging", {}).get("nextPageToken")
        if not page_token:
            break
        time.sleep(2)

    log.info(f"Вопросы ЯМ: {total}")
    _log_db("ym_sync_questions", "success", f"{total} вопросов")


# ── 5. Ассортимент ───────────────────────────────────────────────────────────
def sync_assortment():
    log.info("Синхронизация ассортимента ЯМ...")
    page_token = None
    total = 0

    while True:
        body = {"limit": 200}
        if page_token:
            body["page_token"] = page_token

        data   = _post(f"/v2/businesses/{YM_BUSINESS_ID}/offer-mappings", body)
        result = data.get("result", {})
        offers = result.get("offerMappings", [])
        if not offers:
            break

        with engine.begin() as conn:
            for o in offers:
                offer   = o.get("offer", {})
                mapping = o.get("mapping", {})
                conn.execute(text("""
                    INSERT INTO ym_assortment
                        (offer_id, supplier_article, name, category, market_sku)
                    VALUES
                        (:oid, :article, :name, :cat, :msku)
                    ON CONFLICT (offer_id) DO UPDATE SET
                        name       = EXCLUDED.name,
                        category   = EXCLUDED.category,
                        market_sku = EXCLUDED.market_sku,
                        synced_at  = NOW()
                """), {
                    "oid":     offer.get("offerId", ""),
                    "article": offer.get("offerId", ""),
                    "name":    offer.get("name"),
                    "cat":     offer.get("category"),
                    "msku":    mapping.get("marketSku"),
                })
                total += 1

        page_token = result.get("paging", {}).get("nextPageToken")
        if not page_token:
            break
        time.sleep(0.3)

    log.info(f"Ассортимент ЯМ: {total}")
    _log_db("ym_sync_assortment", "success", f"{total} позиций")


# ── 6. Цены ──────────────────────────────────────────────────────────────────
# price.value — в рублях (не копейках!)
def sync_prices():
    log.info("Синхронизация цен ЯМ...")
    page_token = None
    total = 0

    while True:
        body = {"limit": 200}
        if page_token:
            body["page_token"] = page_token

        data   = _post(f"/v2/businesses/{YM_BUSINESS_ID}/offer-prices", body)
        result = data.get("result", {})
        offers = result.get("offers", [])
        if not offers:
            break

        with engine.begin() as conn:
            for o in offers:
                p = o.get("price", {})
                conn.execute(text("""
                    INSERT INTO ym_prices
                        (supplier_article, price, discount_base, currency)
                    VALUES
                        (:article, :price, :base, :cur)
                    ON CONFLICT (supplier_article) DO UPDATE SET
                        price         = EXCLUDED.price,
                        discount_base = EXCLUDED.discount_base,
                        updated_at    = NOW()
                """), {
                    "article": o.get("offerId", ""),
                    "price":   p.get("value"),           # уже в рублях
                    "base":    p.get("discountBase"),
                    "cur":     p.get("currencyId", "RUR"),
                })
                total += 1

        page_token = result.get("paging", {}).get("nextPageToken")
        if not page_token:
            break
        time.sleep(0.3)

    log.info(f"Цены ЯМ: {total}")
    _log_db("ym_sync_prices", "success", f"{total} позиций")


# ── 7. Финансовые отчёты (асинхронные) ──────────────────────────────────────
def generate_finance_reports():
    yesterday = date.today() - timedelta(days=1)
    month_ago = yesterday   - timedelta(days=30)

    for rtype, endpoint in [
        ("realization", "/v2/reports/goods-realization/generate"),
        ("services",    "/v2/reports/united-marketplace-services/generate"),
    ]:
        try:
            resp = _post(endpoint, {
                "campaignId": int(YM_CAMPAIGN_ID),
                "dateFrom":   month_ago.isoformat(),
                "dateTo":     yesterday.isoformat(),
            })
            report_id = resp.get("result", {}).get("reportId")
            if report_id:
                with engine.begin() as conn:
                    conn.execute(text("""
                        INSERT INTO ym_finance_reports
                            (report_type, report_id, status, period_from, period_to)
                        VALUES (:rtype, :rid, 'pending', :pfrom, :pto)
                    """), {"rtype": rtype, "rid": report_id,
                            "pfrom": month_ago, "pto": yesterday})
                log.info(f"Отчёт {rtype} запущен: {report_id}")
                _log_db(f"ym_report_{rtype}_generate", "success", report_id)
        except Exception as e:
            log.error(f"Ошибка запуска отчёта {rtype}: {e}")
            _log_db(f"ym_report_{rtype}_generate", "error", str(e))


def poll_finance_reports():
    with engine.connect() as conn:
        pending = conn.execute(
            text("SELECT id, report_id, report_type FROM ym_finance_reports WHERE status='pending'")
        ).mappings().all()

    if not pending:
        log.info("Нет ожидающих финансовых отчётов.")
        return

    for row in pending:
        try:
            resp   = _get(f"/v2/reports/info/{row['report_id']}")
            info   = resp.get("result", {})
            status = info.get("status")

            if status == "DONE":
                url = info.get("file")
                revenue, commissions = _parse_report(row["report_type"], url)
                with engine.begin() as conn:
                    conn.execute(text("""
                        UPDATE ym_finance_reports
                           SET status='downloaded', download_url=:url,
                               revenue=:rev, commissions=:comm, downloaded_at=NOW()
                         WHERE id=:id
                    """), {"url": url, "rev": revenue, "comm": commissions, "id": row["id"]})
                log.info(f"Отчёт {row['report_type']} скачан: выручка={revenue}, комиссии={commissions}")
            elif status == "FAILED":
                with engine.begin() as conn:
                    conn.execute(text("UPDATE ym_finance_reports SET status='failed' WHERE id=:id"),
                                  {"id": row["id"]})
                log.warning(f"Отчёт {row['report_id']} завершился ошибкой на стороне ЯМ.")
            else:
                log.info(f"Отчёт {row['report_id']} ещё не готов (status={status}).")
        except Exception as e:
            log.error(f"Ошибка опроса отчёта {row['report_id']}: {e}")


def _parse_report(report_type, url):
    if not url:
        return None, None
    try:
        r  = requests.get(url, headers=_hdr(), timeout=120)
        r.raise_for_status()
        ct = r.headers.get("content-type", "")
        if "zip" in ct or url.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(r.content)) as z:
                content = z.read(z.namelist()[0]).decode("utf-8-sig")
        else:
            content = r.content.decode("utf-8-sig")

        reader = csv.DictReader(io.StringIO(content), delimiter="\t")
        revenue = commissions = 0.0
        for row in reader:
            def _num(k):
                v = str(row.get(k, "") or "").replace(",", ".").replace("\xa0", "").replace(" ", "")
                try:   return float(v)
                except ValueError: return 0.0
            if report_type == "realization":
                revenue     += _num("Стоимость продаж")
                commissions += _num("Агентское вознаграждение")
            else:
                commissions += _num("Сумма")
        return round(revenue, 2), round(commissions, 2)
    except Exception as e:
        log.warning(f"Парсинг отчёта {report_type} упал: {e}")
        return None, None



# ── 8. Поставки (supply requests) ────────────────────────────────────────────
# Тянем все заявки на поставку ЯМ + товарный состав.
# Используется для связи возврат → поставка → инвойс (эвристика: SKU + склад + дата).
#
# Форматы в Google Sheet "Номер поставки" для МП=Яндекс:
#   • "30725002"     → marketplaceRequestId (номер с акта приёмки)
#   • "ВРЦ-8299161"  → ВРЦ- + request_id (VDC-родительская заявка)
# Оба формата обрабатываются в запросе production.py.
def sync_supplies():
    log.info("Синхронизация поставок ЯМ...")
    saved_supplies = 0
    saved_items    = 0

    # Берём большой диапазон — всё что есть; limit=200 чтобы не пагинировать лишний раз
    data = _post(f"/v2/campaigns/{YM_CAMPAIGN_ID}/supply-requests", {
        "limit": 200,
        "dateFrom": "2024-01-01",
        "dateTo": (datetime.now(timezone.utc) + timedelta(days=180)).strftime("%Y-%m-%d"),
    })

    requests_list = data.get("result", {}).get("requests", [])
    if not requests_list:
        log.info("Поставки ЯМ: нет данных")
        return

    for req in requests_list:
        rid_obj = req.get("id", {})
        request_id = rid_obj.get("id")
        if not request_id:
            continue

        marketplace_request_id = rid_obj.get("marketplaceRequestId")
        warehouse_request_id   = rid_obj.get("warehouseRequestId")

        parent_link = req.get("parentLink") or {}
        parent_id   = (parent_link.get("id") or {}).get("id")

        target         = req.get("targetLocation") or {}
        warehouse_id   = target.get("serviceId")
        warehouse_name = target.get("name")
        requested_date = target.get("requestedDate")

        try:
            with engine.begin() as conn:
                conn.execute(text("""
                    INSERT INTO ym_supplies
                        (request_id, marketplace_request_id, warehouse_request_id,
                         parent_request_id, warehouse_id, warehouse_name,
                         requested_date, status, subtype)
                    VALUES
                        (:rid, :mprid, :whrid, :parent,
                         :whid, :whnm, :rdate, :status, :subtype)
                    ON CONFLICT (request_id) DO UPDATE SET
                        marketplace_request_id = COALESCE(EXCLUDED.marketplace_request_id, ym_supplies.marketplace_request_id),
                        warehouse_request_id   = COALESCE(EXCLUDED.warehouse_request_id,   ym_supplies.warehouse_request_id),
                        parent_request_id      = COALESCE(EXCLUDED.parent_request_id,      ym_supplies.parent_request_id),
                        status    = EXCLUDED.status,
                        synced_at = NOW()
                """), {
                    "rid":    request_id,
                    "mprid":  marketplace_request_id,
                    "whrid":  warehouse_request_id,
                    "parent": parent_id,
                    "whid":   warehouse_id,
                    "whnm":   warehouse_name,
                    "rdate":  requested_date,
                    "status": req.get("status"),
                    "subtype":req.get("subtype"),
                })
            saved_supplies += 1
        except Exception as e:
            log.warning(f"  Ошибка записи поставки {request_id}: {e}")
            continue

        # Товарный состав — только у заявок с marketplaceRequestId (VDC-родители пустые)
        if not marketplace_request_id:
            continue
        try:
            items_data = _post(
                f"/v2/campaigns/{YM_CAMPAIGN_ID}/supply-requests/items",
                {"requestId": request_id, "limit": 200}
            )
            items = items_data.get("result", {}).get("items", [])
            for item in items:
                offer_id = item.get("offerId")
                if not offer_id:
                    continue
                cnt = item.get("counters") or {}
                with engine.begin() as conn:
                    conn.execute(text("""
                        INSERT INTO ym_supply_items
                            (request_id, offer_id, plan_count, fact_count, defect_count)
                        VALUES (:rid, :oid, :plan, :fact, :defect)
                        ON CONFLICT (request_id, offer_id) DO UPDATE SET
                            plan_count   = EXCLUDED.plan_count,
                            fact_count   = COALESCE(EXCLUDED.fact_count,   ym_supply_items.fact_count),
                            defect_count = COALESCE(EXCLUDED.defect_count, ym_supply_items.defect_count),
                            synced_at    = NOW()
                    """), {
                        "rid":    request_id,
                        "oid":    offer_id,
                        "plan":   cnt.get("planCount"),
                        "fact":   cnt.get("factCount"),
                        "defect": cnt.get("defectCount"),
                    })
                saved_items += 1
            time.sleep(0.3)
        except Exception as e:
            log.warning(f"  Ошибка items поставки {request_id}: {e}")

    log.info(f"Поставки ЯМ: {saved_supplies} поставок, {saved_items} позиций")
    _log_db("ym_sync_supplies", "success", f"{saved_supplies} поставок, {saved_items} позиций")



def _parse_ratings_xlsx(content_bytes, report_date):
    """
    Парсит 'business_rating_report_*.xlsx' (лист 'Рейтинг товаров').
    Структура: строки-группы (col0 заполнен, col2 пустой) → рейтинг витрины + дельта.
               строки-SKU (col0 пустой, col2 заполнен) → артикул + звёзды + кол-во.
    Столбцы: 0=Название, 1=Ссылка, 2=SKU, 3=Рейтинг, 4=Δнед, 5=Всего, 6=5★…10=1★
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)

    if "Рейтинг товаров" not in wb.sheetnames:
        log.warning(f"Лист 'Рейтинг товаров' не найден. Листы: {wb.sheetnames}")
        return []

    ws = wb["Рейтинг товаров"]
    log.info(f"Парсим лист 'Рейтинг товаров' ({ws.max_row} строк)")

    current_group_rating = None
    current_weekly_delta = None
    current_product_name = None
    rows = []

    for row_vals in ws.iter_rows(min_row=3, values_only=True):
        name  = row_vals[0] if len(row_vals) > 0 else None
        sku   = row_vals[2] if len(row_vals) > 2 else None
        rat   = row_vals[3] if len(row_vals) > 3 else None
        delta = row_vals[4] if len(row_vals) > 4 else None
        total = row_vals[5] if len(row_vals) > 5 else None
        s5    = row_vals[6] if len(row_vals) > 6 else None
        s4    = row_vals[7] if len(row_vals) > 7 else None
        s3    = row_vals[8] if len(row_vals) > 8 else None
        s2    = row_vals[9] if len(row_vals) > 9 else None
        s1    = row_vals[10] if len(row_vals) > 10 else None

        if name and str(name).strip():
            current_group_rating = float(rat)   if isinstance(rat,   (int, float)) else None
            current_weekly_delta = float(delta) if isinstance(delta, (int, float)) else None
            current_product_name = str(name).strip()
        elif sku and str(sku).strip():
            rows.append({
                "report_date":      report_date,
                "supplier_article": str(sku).strip(),
                "product_name":     current_product_name,
                "group_rating":     current_group_rating,
                "weekly_delta":     current_weekly_delta,
                "review_count":     int(total) if isinstance(total, (int, float)) else None,
                "stars_5":          int(s5)    if isinstance(s5,    (int, float)) else None,
                "stars_4":          int(s4)    if isinstance(s4,    (int, float)) else None,
                "stars_3":          int(s3)    if isinstance(s3,    (int, float)) else None,
                "stars_2":          int(s2)    if isinstance(s2,    (int, float)) else None,
                "stars_1":          int(s1)    if isinstance(s1,    (int, float)) else None,
            })

    return rows


# ── 10. Автоматическая загрузка рейтингов из кабинета ЯМ ────────────────────
# Использует веб-API partner.market.yandex.ru через Playwright (автологин).
SESSION_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ym_session.json")


def _ym_login_get_session():
    """Открывает кабинет ЯМ через Playwright. Использует сохранённую сессию если есть."""
    from playwright.sync_api import sync_playwright

    bid  = int(YM_BUSINESS_ID)
    path = f"/business/{bid}/reviews-new?campaignId={YM_CAMPAIGN_ID}&activeTab=products&ratingDrawer=opened"
    ua   = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=[
            "--no-sandbox", "--disable-dev-shm-usage",
            "--disable-blink-features=AutomationControlled",
        ])

        # Загружаем сохранённую сессию если файл есть
        session_exists = os.path.exists(SESSION_FILE)
        ctx_kwargs = dict(user_agent=ua, locale="ru-RU", timezone_id="Europe/Moscow")
        if session_exists:
            ctx_kwargs["storage_state"] = SESSION_FILE
            log.info("Playwright: загружаем сохранённую сессию ЯМ...")

        context = browser.new_context(**ctx_kwargs)
        page    = context.new_page()

        log.info("Playwright: переходим в кабинет ЯМ...")
        page.goto(f"https://partner.market.yandex.ru{path}", wait_until="domcontentloaded", timeout=60000)
        time.sleep(6)

        content = page.content()
        sk_m = re.search(r'"sk":"(u[a-f0-9]+)"', content)

        if not sk_m and session_exists:
            # Сессия протухла — пробуем свежий логин через логин/пароль
            log.warning("Сохранённая сессия устарела, выполняем повторный логин...")
            browser.close()
            # Удаляем протухший файл и вызываем себя рекурсивно (один раз)
            os.remove(SESSION_FILE)
            return _ym_login_get_session()

        cookies = context.cookies()
        cookie_str = "; ".join(
            f"{c['name']}={c['value']}" for c in cookies
            if any(d in c["domain"] for d in ["yandex", "market"])
        )

        if sk_m:
            # Обновляем сохранённую сессию
            context.storage_state(path=SESSION_FILE)
            log.info("Playwright: сессия обновлена")

        browser.close()

    if not sk_m:
        snippet = re.sub(r'<[^>]+>', ' ', content)[:400].strip()
        raise RuntimeError(f"sk-токен не найден. Начало страницы: {snippet}")
    return cookie_str, sk_m.group(1)


def sync_business_rating_report():
    """Скачивает BUSINESS_RATING_REPORT через автологин Playwright."""
    today = date.today()
    bid   = int(YM_BUSINESS_ID)
    ua    = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36"
    partner_base = "https://partner.market.yandex.ru"
    path  = f"/business/{bid}/reviews-new?campaignId={YM_CAMPAIGN_ID}&activeTab=products&ratingDrawer=opened"

    try:
        cookie, sk = _ym_login_get_session()
    except Exception as e:
        log.error(f"Не удалось залогиниться в ЯМ: {e}")
        _log_db("ym_sync_ratings", "error", f"login failed: {e}")
        # Переносим данные с предыдущего дня чтобы не было пропусков в дашборде
        filled = _fill_missing_skus(today)
        if filled:
            log.info(f"Рейтинги ЯМ: перенесено {filled} SKU с предыдущего дня (логин недоступен)")
        return

    hdrs = {
        "Content-Type": "application/json",
        "Cookie": cookie,
        "Accept": "*/*",
        "Origin": partner_base,
        "Referer": f"{partner_base}{path}",
        "User-Agent": ua,
        "sk": sk,
    }

    # Шаг 2: запросить генерацию (обязателен reportParams с language!)
    log.info("Запрашиваем генерацию BUSINESS_RATING_REPORT...")
    gen_resp = requests.post(
        f"{partner_base}/api/resolve/?r=mbiPartner/asyncReports/generateReport:resolveGenerateReport",
        headers=hdrs,
        json={
            "params": [{"businessId": bid, "reportType": "BUSINESS_RATING_REPORT",
                        "reportParams": {"businessId": bid, "language": "ru"}}],
            "path": path,
        },
        timeout=30,
    )
    gen_resp.raise_for_status()
    gen_data = gen_resp.json()

    try:
        report_id = gen_data["results"][0]["data"][0]["id"]
    except (KeyError, IndexError, TypeError):
        log.error(f"Не удалось получить reportId: {str(gen_data)[:300]}")
        _log_db("ym_sync_ratings", "error", f"no reportId: {str(gen_data)[:200]}")
        return

    log.info(f"Отчёт запущен, reportId={report_id}. Ждём готовности...")

    # Шаг 3: опрашиваем статус (раз в 10 сек, до 10 минут)
    ready = False
    for attempt in range(60):
        time.sleep(10)
        poll_resp = requests.post(
            f"{partner_base}/api/resolve/?r=mbiPartner/asyncReports/getReport:resolveReport",
            headers=hdrs,
            json={"params": [{"businessId": bid, "reportId": report_id}], "path": path},
            timeout=30,
        )
        try:
            state = poll_resp.json()["results"][0]["data"][0].get("state", "")
        except (KeyError, IndexError, TypeError):
            state = ""
        log.info(f"  попытка {attempt+1}: state={state}")
        if state == "DONE":
            ready = True
            break
        if state in ("ERROR", "FAILED", "FAILURE"):
            log.error(f"Отчёт завершился ошибкой: {poll_resp.text[:200]}")
            _log_db("ym_sync_ratings", "error", f"report state={state}")
            return

    if not ready:
        log.error("Отчёт не готов за 10 минут — таймаут")
        _log_db("ym_sync_ratings", "error", "timeout")
        return

    # Шаг 4: скачиваем файл (ждём 10 сек — данные заполняются чуть позже DONE)
    time.sleep(10)
    download_url = f"{partner_base}/api/files/stat/async-reports/{report_id}?businessId={bid}"
    log.info("Скачиваем отчёт...")
    dl = requests.get(download_url, headers={"Cookie": cookie, "User-Agent": ua,
                                             "Referer": f"{partner_base}/"}, timeout=120)
    dl.raise_for_status()

    # Шаг 5: парсим и сохраняем
    rows = _parse_ratings_xlsx(dl.content, today)
    if not rows:
        log.error(f"Парсинг вернул 0 строк (размер файла: {len(dl.content)} байт)")
        _log_db("ym_sync_ratings", "error", "0 rows parsed")
        return

    _ingest_ratings_rows(rows)
    # Заполняем пропущенные SKU данными с предыдущего дня
    filled = _fill_missing_skus(today)
    log.info(f"Рейтинги ЯМ обновлены: {len(rows)} SKU из отчёта + {filled} SKU перенесены с предыдущего дня")
    _log_db("ym_sync_ratings", "success", f"{len(rows)} SKU за {today}, filled={filled}")


def _ingest_ratings_rows(rows):
    """Upsert строк в ym_ratings_report по (report_date, supplier_article)."""
    if not rows:
        return
    with engine.begin() as conn:
        for r in rows:
            conn.execute(text("""
                INSERT INTO ym_ratings_report
                    (report_date, supplier_article, product_name, group_rating,
                     weekly_delta, review_count, stars_5, stars_4, stars_3, stars_2, stars_1)
                VALUES
                    (:report_date, :supplier_article, :product_name, :group_rating,
                     :weekly_delta, :review_count, :stars_5, :stars_4, :stars_3, :stars_2, :stars_1)
                ON CONFLICT (report_date, supplier_article) DO UPDATE SET
                    product_name = EXCLUDED.product_name,
                    group_rating = EXCLUDED.group_rating,
                    weekly_delta = EXCLUDED.weekly_delta,
                    review_count = EXCLUDED.review_count,
                    stars_5 = EXCLUDED.stars_5, stars_4 = EXCLUDED.stars_4,
                    stars_3 = EXCLUDED.stars_3, stars_2 = EXCLUDED.stars_2,
                    stars_1 = EXCLUDED.stars_1,
                    synced_at = NOW()
            """), r)


def _fill_missing_skus(report_date):
    """Для SKU из ym_prices которых нет в отчёте — переносим данные с ближайшего предыдущего дня."""
    with engine.begin() as conn:
        missing = conn.execute(text("""
            SELECT p.supplier_article
            FROM ym_prices p
            WHERE NOT EXISTS (
                SELECT 1 FROM ym_ratings_report r
                WHERE r.supplier_article = p.supplier_article
                  AND r.report_date = :dt
            )
        """), {"dt": report_date}).fetchall()

        filled = 0
        for (sku,) in missing:
            prev = conn.execute(text("""
                SELECT report_date, product_name, group_rating, weekly_delta,
                       review_count, stars_5, stars_4, stars_3, stars_2, stars_1
                FROM ym_ratings_report
                WHERE supplier_article = :sku AND report_date < :dt
                ORDER BY report_date DESC LIMIT 1
            """), {"sku": sku, "dt": report_date}).fetchone()

            if prev:
                conn.execute(text("""
                    INSERT INTO ym_ratings_report
                        (report_date, supplier_article, product_name, group_rating,
                         weekly_delta, review_count, stars_5, stars_4, stars_3, stars_2, stars_1)
                    VALUES (:dt, :sku, :product_name, :group_rating, 0,
                            :review_count, :stars_5, :stars_4, :stars_3, :stars_2, :stars_1)
                    ON CONFLICT (report_date, supplier_article) DO NOTHING
                """), {"dt": report_date, "sku": sku, "product_name": prev[1],
                       "group_rating": prev[2], "review_count": prev[4],
                       "stars_5": prev[5], "stars_4": prev[6], "stars_3": prev[7],
                       "stars_2": prev[8], "stars_1": prev[9]})
                filled += 1
            else:
                # Новый SKU без истории — вставляем нулевую запись
                conn.execute(text("""
                    INSERT INTO ym_ratings_report (report_date, supplier_article, group_rating, review_count)
                    VALUES (:dt, :sku, 0, 0)
                    ON CONFLICT (report_date, supplier_article) DO NOTHING
                """), {"dt": report_date, "sku": sku})
                filled += 1

    return filled


# ── Исторический сбор отзывов (backfill) ────────────────────────────────────
# Запускается вручную один раз: python3 ym_worker.py backfill
# Загружает все отзывы с начала года с повторными попытками при 420.
def sync_feedbacks_backfill():
    log.info("Исторический сбор отзывов ЯМ (backfill)...")
    page_token = None
    total = 0
    max_attempts = 12

    while True:
        body = {"pageSize": 100}
        if page_token:
            body["pageToken"] = page_token

        attempt = 0
        data = None
        while attempt < max_attempts:
            try:
                data = _post(f"/v2/businesses/{YM_BUSINESS_ID}/goods-feedback", body)
                break
            except RateLimitError:
                attempt += 1
                wait = 60 * attempt
                log.warning(f"Rate limit на backfill, ждём {wait}с (попытка {attempt}/{max_attempts})...")
                time.sleep(wait)

        if data is None:
            log.error("Backfill прерван: исчерпаны попытки после rate limit.")
            break

        items = data.get("result", {}).get("feedbacks", [])
        if not items:
            break

        page_new = 0
        with engine.begin() as conn:
            for fb in items:
                desc   = fb.get("description") or {}
                stats  = fb.get("statistics") or {}
                idents = fb.get("identifiers") or {}
                r = conn.execute(text("""
                    INSERT INTO ym_feedbacks
                        (feedback_id, supplier_article, created_date, valuation,
                         pro_text, contra_text, comment)
                    VALUES
                        (:fid, :article, :created_date, :val,
                         :pro, :contra, :comment)
                    ON CONFLICT (feedback_id) DO NOTHING
                """), {
                    "fid":          str(fb.get("feedbackId", "")),
                    "article":      str(idents.get("offerId") or ""),
                    "created_date": fb.get("createdAt"),
                    "val":          stats.get("rating"),
                    "pro":          desc.get("advantages", ""),
                    "contra":       desc.get("disadvantages", ""),
                    "comment":      desc.get("comment", ""),
                })
                page_new += r.rowcount

        total += page_new

        # Если страница не дала новых записей — ЯМ закончил уникальные данные
        if page_new == 0:
            log.info("Backfill: страница без новых записей — завершаем.")
            break

        page_token = data.get("result", {}).get("paging", {}).get("nextPageToken")
        if not page_token:
            break
        log.info(f"Backfill: +{page_new} новых ({total} всего), продолжаем...")
        time.sleep(3)

    log.info(f"Backfill завершён: {total} новых отзывов добавлено")
    _log_db("ym_backfill_feedbacks", "success", f"{total} новых отзывов")


# ── Точка входа ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    task = sys.argv[1] if len(sys.argv) > 1 else "hourly"

    if not YM_API_KEY or not YM_CAMPAIGN_ID or not YM_BUSINESS_ID:
        log.error("Не заданы YM_API_KEY / YM_CAMPAIGN_ID / YM_BUSINESS_ID в .env")
        sys.exit(1)

    def run(fn):
        try:
            fn()
        except RateLimitError as e:
            log.warning(f"Пропуск задачи из-за rate limit: {e}")
            _log_db(fn.__name__, "rate_limit", str(e))
        except Exception as e:
            log.error(f"Ошибка в {fn.__name__}: {e}")
            _log_db(fn.__name__, "error", str(e))

    if task == "hourly":
        run(sync_orders)
        run(sync_returns)
        run(sync_feedbacks)
        run(sync_questions)
        # Автоматическое ИИ-тегирование новых возвратов ЯМ после синхронизации
        try:
            import subprocess
            script_dir = os.path.dirname(os.path.abspath(__file__))
            subprocess.run(
                [sys.executable, os.path.join(script_dir, "ai_tagger.py"), "ym"],
                timeout=3600
            )
        except Exception as e:
            _log_db("ai_tagger_ym", "error", f"Сбой запуска: {e}")
    elif task == "daily":
        run(sync_assortment)
        run(sync_prices)
        run(generate_finance_reports)
        run(sync_supplies)
        run(sync_business_rating_report)
    elif task == "reports":
        run(poll_finance_reports)
    elif task == "backfill":
        run(sync_feedbacks_backfill)
    elif task == "orders-backfill":
        from_date = sys.argv[2] if len(sys.argv) > 2 else "2025-01-01"
        sync_orders_backfill(from_date)
    else:
        log.error(f"Неизвестная задача: '{task}'. Используй: hourly | daily | reports | backfill | orders-backfill [YYYY-MM-DD]")
        sys.exit(1)
