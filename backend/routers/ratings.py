import io
import re
from datetime import date, datetime

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session
from sqlalchemy import text

from ..database import get_db
from .auth import get_current_user

router = APIRouter(prefix="/api/ratings", tags=["Ratings"], dependencies=[Depends(get_current_user)])


# ── YM-рейтинги: гибридный источник ─────────────────────────────────────────
# Основа: ym_ratings_report — снимки из XLSX (загружаются через /ym/upload).
#         Содержит накопительные итоги и официальный рейтинг витрины.
# Дополнение: ym_feedbacks — почасовая синхронизация новых отзывов.
#         Заполняет даты без снимков, считает daily_new и дельту звёзд.
_YM_REPORT_SQL = text("""
    WITH fb_day AS (
        SELECT
            TRIM(supplier_article)                            AS sku,
            created_date::date                               AS dt,
            COUNT(*)                                         AS new_cnt,
            COUNT(CASE WHEN valuation = 5 THEN 1 END)        AS new_s5,
            COUNT(CASE WHEN valuation = 4 THEN 1 END)        AS new_s4,
            COUNT(CASE WHEN valuation = 3 THEN 1 END)        AS new_s3,
            COUNT(CASE WHEN valuation = 2 THEN 1 END)        AS new_s2,
            COUNT(CASE WHEN valuation = 1 THEN 1 END)        AS new_s1
        FROM ym_feedbacks
        WHERE supplier_article IS NOT NULL
        GROUP BY TRIM(supplier_article), created_date::date
    )
    SELECT
        d.report_date::text       AS date,
        d.supplier_article,
        COALESCE(d.group_rating, 0) AS average_rating,

        CASE
            WHEN COALESCE(d.review_count, 0) > 0 THEN d.review_count
            ELSE COALESCE(prev.review_count, 0) + COALESCE(fb.new_cnt, 0)
        END AS review_count,

        CASE
            WHEN COALESCE(d.review_count, 0) > 0 AND COALESCE(prev.review_count, 0) > 0
                THEN GREATEST(0, d.review_count - prev.review_count)
            ELSE COALESCE(fb.new_cnt, 0)
        END AS daily_new,

        COALESCE(d.stars_5, COALESCE(prev.stars_5, 0) + COALESCE(fb.new_s5, 0)) AS stars_5,
        COALESCE(d.stars_4, COALESCE(prev.stars_4, 0) + COALESCE(fb.new_s4, 0)) AS stars_4,
        COALESCE(d.stars_3, COALESCE(prev.stars_3, 0) + COALESCE(fb.new_s3, 0)) AS stars_3,
        COALESCE(d.stars_2, COALESCE(prev.stars_2, 0) + COALESCE(fb.new_s2, 0)) AS stars_2,
        COALESCE(d.stars_1, COALESCE(prev.stars_1, 0) + COALESCE(fb.new_s1, 0)) AS stars_1,

        COALESCE(a.is_new, FALSE)        AS is_new,
        COALESCE(a.category_1, 'Прочее') AS category
    FROM ym_ratings_report d
    LEFT JOIN LATERAL (
        SELECT review_count, stars_5, stars_4, stars_3, stars_2, stars_1
        FROM ym_ratings_report p
        WHERE p.supplier_article = d.supplier_article AND p.report_date < d.report_date
        ORDER BY p.report_date DESC
        LIMIT 1
    ) prev ON TRUE
    LEFT JOIN fb_day fb ON fb.sku = d.supplier_article AND fb.dt = d.report_date
    LEFT JOIN wb_assortment a ON a.supplier_article = d.supplier_article
    WHERE d.supplier_article IS NOT NULL
    ORDER BY d.report_date ASC
""")


def _ym_rows_to_dicts(rows):
    return [
        {
            "date": str(r.date),
            "supplier_article": r.supplier_article,
            "average_rating": float(r.average_rating or 0),
            "review_count": r.review_count,
            "daily_new": int(r.daily_new or 0),
            "stars_1": r.stars_1, "stars_2": r.stars_2, "stars_3": r.stars_3,
            "stars_4": r.stars_4, "stars_5": r.stars_5,
            "is_new": bool(getattr(r, "is_new", False)),
            "category": getattr(r, "category", "Прочее") or "Прочее",
        }
        for r in rows if r.supplier_article
    ]


@router.get("/wb/summary")
def get_wb_ratings_summary(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Сводная таблица рейтингов ВБ: накопительные итоги + новые отзывы вчера."""
    try:
        rows = db.execute(text("""
            SELECT
                t.supplier_article,
                t.average_rating,
                t.review_count          AS total_reviews,
                COALESCE(r."feedbackCount_current", 0) AS new_today,
                t.stars_5, t.stars_4, t.stars_3, t.stars_2, t.stars_1,
                t.refreshed_at::text    AS snapshot_date,
                COALESCE(a.is_new, FALSE) AS is_new
            FROM wb_ratings_totals t
            LEFT JOIN wb_ratings_raw r
                ON r.sys_article = t.supplier_article
                AND r.sys_date = (SELECT MAX(sys_date) FROM wb_ratings_raw)
            LEFT JOIN wb_assortment a ON a.supplier_article = t.supplier_article
            ORDER BY t.review_count DESC NULLS LAST
        """)).fetchall()
        return [
            {
                "supplier_article": r.supplier_article,
                "average_rating":   float(r.average_rating or 0),
                "total_reviews":    r.total_reviews or 0,
                "new_today":        int(r.new_today or 0),
                "stars_5": r.stars_5 or 0, "stars_4": r.stars_4 or 0,
                "stars_3": r.stars_3 or 0, "stars_2": r.stars_2 or 0,
                "stars_1": r.stars_1 or 0,
                "snapshot_date": str(r.snapshot_date or ""),
                "is_new": bool(r.is_new),
            }
            for r in rows
        ]
    except Exception as e:
        print(f"Ошибка WB summary: {e}")
        return []


@router.get("/ym/summary")
def get_ym_ratings_summary(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Сводка рейтингов ЯМ — считаем из ym_feedbacks (синхр. каждые 2ч)."""
    try:
        rows = db.execute(text("""
            WITH totals AS (
                SELECT
                    TRIM(supplier_article)                              AS sku,
                    COUNT(*)                                            AS total_reviews,
                    ROUND(AVG(valuation)::numeric, 2)                  AS average_rating,
                    COUNT(CASE WHEN valuation = 5 THEN 1 END)          AS stars_5,
                    COUNT(CASE WHEN valuation = 4 THEN 1 END)          AS stars_4,
                    COUNT(CASE WHEN valuation = 3 THEN 1 END)          AS stars_3,
                    COUNT(CASE WHEN valuation = 2 THEN 1 END)          AS stars_2,
                    COUNT(CASE WHEN valuation = 1 THEN 1 END)          AS stars_1,
                    MAX(created_date::date)::text                       AS snapshot_date
                FROM ym_feedbacks
                WHERE supplier_article IS NOT NULL
                  AND TRIM(supplier_article) != ''
                  AND valuation IS NOT NULL
                GROUP BY TRIM(supplier_article)
            ),
            new_today AS (
                SELECT
                    TRIM(supplier_article) AS sku,
                    COUNT(*) AS cnt
                FROM ym_feedbacks
                WHERE supplier_article IS NOT NULL
                  AND created_date::date = CURRENT_DATE
                GROUP BY TRIM(supplier_article)
            )
            SELECT
                t.sku               AS supplier_article,
                t.average_rating,
                t.total_reviews,
                COALESCE(n.cnt, 0)  AS new_today,
                t.stars_5, t.stars_4, t.stars_3, t.stars_2, t.stars_1,
                t.snapshot_date,
                COALESCE(a.is_new, FALSE) AS is_new
            FROM totals t
            LEFT JOIN new_today n ON n.sku = t.sku
            LEFT JOIN wb_assortment a ON a.supplier_article = t.sku
            ORDER BY t.total_reviews DESC
        """)).fetchall()
        return [
            {
                "supplier_article": r.supplier_article,
                "product_name":     "",
                "average_rating":   float(r.average_rating or 0),
                "weekly_delta":     0.0,
                "total_reviews":    r.total_reviews or 0,
                "new_today":        int(r.new_today or 0),
                "stars_5": r.stars_5 or 0, "stars_4": r.stars_4 or 0,
                "stars_3": r.stars_3 or 0, "stars_2": r.stars_2 or 0,
                "stars_1": r.stars_1 or 0,
                "snapshot_date":    str(r.snapshot_date or ""),
                "is_new": bool(r.is_new),
            }
            for r in rows
        ]
    except Exception as e:
        print(f"Ошибка YM summary: {e}")
        return []


@router.get("/")
def get_ratings(
    platform: str = Query("wb", pattern="^(wb|ym|ozon|all)$"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    columns_query = text("SELECT column_name FROM information_schema.columns WHERE table_name = 'wb_ratings_raw'")
    tables_query = text("SELECT table_name FROM information_schema.tables WHERE table_name = 'wb_assortment'")

    try:
        if platform == "ym":
            rows = db.execute(_YM_REPORT_SQL).fetchall()
            return _ym_rows_to_dicts(rows)

        # ── WB ────────────────────────────────────────────────────────────────
        existing_cols = [r[0] for r in db.execute(columns_query).fetchall()]
        if not existing_cols:
            return []
        has_assortment = db.execute(tables_query).scalar() is not None

        def find_col(names):
            lower_map = {c.lower(): c for c in existing_cols}
            for n in names:
                if n.lower() in lower_map:
                    return lower_map[n.lower()]
            return "0"

        rating_col = find_col(["feedbackRating_current", "statistics_valuation", "valuation", "rating_current"])
        count_col  = find_col(["feedbackCount_current", "statistics_feedbacksCount", "feedbacksCount_current"])
        s1 = find_col(["oneStar_current", "statistics_star1", "star1", "onestar"])
        s2 = find_col(["twoStar_current", "statistics_star2", "star2", "twostar"])
        s3 = find_col(["threeStar_current", "statistics_star3", "star3", "threestar"])
        s4 = find_col(["fourStar_current", "statistics_star4", "star4", "fourstar"])
        s5 = find_col(["fiveStar_current", "statistics_star5", "star5", "fivestar"])

        join_clause      = "LEFT JOIN wb_assortment a ON r.sys_article = a.supplier_article" if has_assortment else ""
        is_new_select    = "COALESCE(a.is_new, FALSE) as is_new" if has_assortment else "FALSE as is_new"
        category_select  = "COALESCE(a.category_1, 'Прочее') as category" if has_assortment else "'Прочее' as category"

        wb_query = text(f"""
            SELECT
                r.sys_date as date,
                r.sys_article as supplier_article,
                CAST(COALESCE(r."{rating_col}", 0) AS FLOAT)   as average_rating,
                CAST(COALESCE(r."{count_col}",  0) AS INTEGER) as review_count,
                CAST(COALESCE(r."{s1}", 0) AS INTEGER) as stars_1,
                CAST(COALESCE(r."{s2}", 0) AS INTEGER) as stars_2,
                CAST(COALESCE(r."{s3}", 0) AS INTEGER) as stars_3,
                CAST(COALESCE(r."{s4}", 0) AS INTEGER) as stars_4,
                CAST(COALESCE(r."{s5}", 0) AS INTEGER) as stars_5,
                {is_new_select},
                {category_select}
            FROM wb_ratings_raw r
            {join_clause}
            WHERE r.sys_article IS NOT NULL
                AND r.sys_article NOT IN ('Unknown', 'nan', '')
            ORDER BY r.sys_date ASC
        """)

        result = db.execute(wb_query).fetchall()
        wb_result = [
            {
                "date": str(r.date),
                "supplier_article": r.supplier_article,
                "average_rating": round(r.average_rating, 2) if r.average_rating <= 5.0 else round(r.average_rating / 10.0, 2),
                "review_count": r.review_count,
                "daily_new": 0,
                "stars_1": r.stars_1, "stars_2": r.stars_2, "stars_3": r.stars_3,
                "stars_4": r.stars_4, "stars_5": r.stars_5,
                "is_new": r.is_new,
                "category": getattr(r, "category", "Прочее") or "Прочее",
            }
            for r in result
        ]

        if platform == "wb":
            return wb_result

        # platform == "all": WB + YM из официального отчёта
        ym_rows = db.execute(_YM_REPORT_SQL).fetchall()
        return wb_result + _ym_rows_to_dicts(ym_rows)

    except Exception as e:
        print(f"Ошибка чтения рейтингов: {e}")
        return []


def _parse_business_rating_xlsx(content: bytes, report_date: date) -> list[dict]:
    """
    Парсит 'business_rating_report_*.xlsx' (лист 'Рейтинг товаров').
    Структура: строки-группы (col0 заполнен, col2 пустой) → рейтинг витрины + дельта за неделю.
               строки-SKU (col0 пустой, col2 заполнен)   → артикул + кол-во отзывов + звёзды.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    if "Рейтинг товаров" not in wb.sheetnames:
        raise ValueError(f"Лист 'Рейтинг товаров' не найден. Листы: {wb.sheetnames}")

    ws = wb["Рейтинг товаров"]

    current_group_rating = None
    current_weekly_delta = None
    current_product_name = None
    rows = []

    for row_vals in ws.iter_rows(min_row=3, values_only=True):
        name   = row_vals[0]
        sku    = row_vals[2] if len(row_vals) > 2 else None
        rating = row_vals[3] if len(row_vals) > 3 else None
        delta  = row_vals[4] if len(row_vals) > 4 else None
        total  = row_vals[5] if len(row_vals) > 5 else None
        s5     = row_vals[6] if len(row_vals) > 6 else None
        s4     = row_vals[7] if len(row_vals) > 7 else None
        s3     = row_vals[8] if len(row_vals) > 8 else None
        s2     = row_vals[9] if len(row_vals) > 9 else None
        s1     = row_vals[10] if len(row_vals) > 10 else None

        if name and str(name).strip():
            # Строка-группа: запоминаем рейтинг для дочерних SKU
            current_group_rating = float(rating) if isinstance(rating, (int, float)) else None
            current_weekly_delta = float(delta)  if isinstance(delta,  (int, float)) else None
            current_product_name = str(name).strip()

        elif sku and str(sku).strip():
            # Строка-SKU
            sku_str = str(sku).strip()
            rows.append({
                "report_date":      report_date,
                "supplier_article": sku_str,
                "product_name":     current_product_name,
                "group_rating":     current_group_rating,
                "weekly_delta":     current_weekly_delta,
                "review_count":     int(total) if isinstance(total, (int, float)) else None,
                "stars_5":          int(s5)    if isinstance(s5,    (int, float)) else None,
                "stars_4":          int(s4)    if isinstance(s4,    (int, float)) else None,
                "stars_3":          int(s3)    if isinstance(s3,    (int, float)) else None,
                "stars_2":          int(s2)    if isinstance(s2,    (int, float)) else None,
                "stars_1":          int(s1)    if isinstance(s1,    (int, float)) else None,
            })

    return rows


def _ingest_rating_rows(rows: list[dict], conn) -> int:
    """Записывает строки рейтингов в ym_ratings_report. Возвращает кол-во вставленных."""
    inserted = 0
    for r in rows:
        conn.execute(text("""
            INSERT INTO ym_ratings_report
                (report_date, supplier_article, product_name, group_rating, weekly_delta,
                 review_count, stars_5, stars_4, stars_3, stars_2, stars_1)
            VALUES
                (:report_date, :supplier_article, :product_name, :group_rating, :weekly_delta,
                 :review_count, :stars_5, :stars_4, :stars_3, :stars_2, :stars_1)
            ON CONFLICT (report_date, supplier_article) DO UPDATE SET
                product_name  = EXCLUDED.product_name,
                group_rating  = EXCLUDED.group_rating,
                weekly_delta  = EXCLUDED.weekly_delta,
                review_count  = EXCLUDED.review_count,
                stars_5       = EXCLUDED.stars_5,
                stars_4       = EXCLUDED.stars_4,
                stars_3       = EXCLUDED.stars_3,
                stars_2       = EXCLUDED.stars_2,
                stars_1       = EXCLUDED.stars_1
        """), r)
        inserted += 1
    return inserted


@router.post("/ym/upload")
async def upload_ym_ratings_report(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Загрузка официального отчёта ЯМ 'Аналитика отзывов' (business_rating_report_*.xlsx).
    Дата берётся из имени файла (DD-MM-YYYY) или из сегодняшней даты.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Ожидается файл .xlsx")

    # Извлекаем дату из имени: business_rating_report_203987989_11-06-2026.xlsx
    report_date = date.today()
    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", file.filename or "")
    if m:
        try:
            report_date = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass

    content = await file.read()
    try:
        rows = _parse_business_rating_xlsx(content, report_date)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка парсинга: {e}")

    if not rows:
        raise HTTPException(status_code=422, detail="Файл не содержит данных (0 SKU)")

    with db.bind.connect() as conn:
        inserted = _ingest_rating_rows(rows, conn)
        conn.commit()

    return {
        "status": "ok",
        "report_date": report_date.isoformat(),
        "skus_imported": inserted,
    }
