import secrets
import os
import json
import uuid
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File, Form
from fastapi.responses import JSONResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel
from typing import Optional
from ..database import get_db
from .auth import get_current_user

try:
    from slowapi import Limiter
    from slowapi.util import get_remote_address
    _limiter = Limiter(key_func=get_remote_address)
    _rate_limit = _limiter.limit("10/hour")
    HAS_LIMITER = True
except ImportError:
    HAS_LIMITER = False
    _rate_limit = lambda f: f

UPLOAD_DIR = Path("/root/cx-dashboard/uploads/reshipment")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"}
MAX_FILE_SIZE_MB = 15

# ── Публичный роутер (без авторизации) ──────────────────────────────────────
public_router = APIRouter(prefix="/api/v1/reshipment", tags=["Reshipment / Доотправка"])

# ── Внутренний роутер (с авторизацией) ──────────────────────────────────────
router = APIRouter(
    prefix="/api/v1/reshipment",
    tags=["Reshipment / Доотправка"],
    dependencies=[Depends(get_current_user)]
)


# ── Схемы ────────────────────────────────────────────────────────────────────

class ApproveRequest(BaseModel):
    matched_srid: Optional[str] = None
    match_notes: Optional[str] = None
    moderator_comment: Optional[str] = None
    processed_by: Optional[str] = None


class RejectRequest(BaseModel):
    moderator_comment: str
    processed_by: Optional[str] = None


class WarehouseRejectRequest(BaseModel):
    rejection_reason: str
    processed_by: Optional[str] = None


class BulkConfirmRequest(BaseModel):
    ids: list[int]
    processed_by: Optional[str] = None


class ShipYandexRequest(BaseModel):
    track_number: str
    shipping_cost: Optional[float] = None
    processed_by: Optional[str] = None


# ── Публичные эндпоинты ───────────────────────────────────────────────────────

@public_router.get("/categories")
def get_categories(db: Session = Depends(get_db)):
    """Список категорий товаров для формы"""
    rows = db.execute(text("""
        SELECT DISTINCT category_1 FROM wb_assortment
        WHERE matrix_status IN ('В матрице', 'Новинка')
          AND category_1 IS NOT NULL AND category_1 != ''
        ORDER BY category_1
    """)).fetchall()

    if rows:
        return {"categories": [r[0] for r in rows]}

    rows = db.execute(text("""
        SELECT DISTINCT category FROM wb_logistics
        WHERE category IS NOT NULL AND category != ''
        ORDER BY category
    """)).fetchall()
    return {"categories": [r[0] for r in rows]}


@public_router.get("/pvz")
def get_pvz_list(city: str = Query(..., min_length=2)):
    """Список ПВЗ СДЭК по городу — для формы клиента."""
    try:
        from ..cdek import get_pvz_list
        points = get_pvz_list(city)
        return {"status": "success", "data": points}
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Ошибка получения ПВЗ: {e}")


@public_router.post("/upload-photo")
async def upload_photo(request: Request, file: UploadFile = File(...)):
    """Загрузка фото от клиента"""
    ext = Path(file.filename or "").suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Допустимые форматы: JPG, PNG, WEBP, HEIC")

    contents = await file.read()
    if len(contents) > MAX_FILE_SIZE_MB * 1024 * 1024:
        raise HTTPException(status_code=400, detail=f"Файл слишком большой (максимум {MAX_FILE_SIZE_MB} МБ)")

    safe_name = f"{uuid.uuid4().hex}{ext}"
    file_path = UPLOAD_DIR / safe_name
    file_path.write_bytes(contents)
    return {"filename": safe_name, "url": f"/uploads/reshipment/{safe_name}"}


@public_router.post("/submit")
def submit_reshipment(
    request: Request,
    customer_name: str = Form(...),
    customer_phone: str = Form(...),
    customer_email: str = Form(None),
    order_number: str = Form(None),
    product_category: str = Form(None),
    problem_type: str = Form(...),
    items_to_send: str = Form(...),
    # ПВЗ — новые поля
    client_pvz_code: str = Form(None),
    client_pvz_address: str = Form(None),
    client_pvz_city: str = Form(None),
    # Старые адресные поля (оставляем для совместимости)
    address_postal: str = Form(None),
    address_region: str = Form(None),
    address_city: str = Form(None),
    address_street: str = Form(None),
    address_house: str = Form(None),
    personal_data_consent: bool = Form(...),
    photo_files: str = Form(None),
    honeypot: str = Form(""),
    # WB чат — привязка если форма открыта по ссылке из чата
    wb_chat_id: str = Form(None),
    # Salebot — ID подписчика, если форма открыта по ссылке из бота (?sbc=...)
    salebot_client_id: str = Form(None),
    db: Session = Depends(get_db)
):
    """Отправка публичной формы клиентом"""
    if honeypot:
        return JSONResponse({"status": "success", "id": 0, "message": "Заявка принята."})

    if not personal_data_consent:
        raise HTTPException(status_code=400, detail="Необходимо согласие на обработку персональных данных")
    if not customer_name.strip():
        raise HTTPException(status_code=400, detail="Укажите имя")
    if not customer_phone.strip():
        raise HTTPException(status_code=400, detail="Укажите телефон")
    if not problem_type.strip():
        raise HTTPException(status_code=400, detail="Выберите тип проблемы")
    if not items_to_send.strip():
        raise HTTPException(status_code=400, detail="Укажите, что нужно отправить")
    if not address_city and not client_pvz_city:
        raise HTTPException(status_code=400, detail="Укажите город доставки")

    shipping_address = client_pvz_address or ", ".join(filter(None, [
        address_postal, address_region, address_city, address_street, address_house
    ])) or None

    # Salebot client ID — чистим значение из формы
    clean_salebot_id = (salebot_client_id or "").strip() or None

    # Если форма пришла из WB чата — достаём reply_sign и nmId из pending
    wb_reply_sign  = None
    wb_nm_id       = None
    wb_client_id   = None
    clean_chat_id  = (wb_chat_id or "").strip() or None

    if clean_chat_id:
        pending = db.execute(text("""
            SELECT reply_sign, nm_id, client_id FROM wb_chat_pending WHERE chat_id = :cid
        """), {"cid": clean_chat_id}).mappings().first()
        if pending:
            wb_reply_sign = pending["reply_sign"]
            wb_nm_id      = pending["nm_id"]
            wb_client_id  = pending["client_id"]

    try:
        result = db.execute(text("""
            INSERT INTO reshipment_requests
                (customer_name, customer_email, customer_phone, order_number,
                 product_category, problem_type, items_to_send,
                 shipping_address, address_postal, address_region,
                 address_city, address_street, address_house,
                 client_pvz_code, client_pvz_address, client_pvz_city,
                 photo_files, privacy_consent,
                 wb_chat_id, wb_reply_sign, wb_nm_id, wb_client_id,
                 salebot_client_id)
            VALUES
                (:name, :email, :phone, :order_num,
                 :category, :problem_type, :items,
                 :address, :postal, :region,
                 :city, :street, :house,
                 :pvz_code, :pvz_address, :pvz_city,
                 :photos, :consent,
                 :wb_chat_id, :wb_reply_sign, :wb_nm_id, :wb_client_id,
                 :salebot_client_id)
            RETURNING id
        """), {
            "name":          customer_name.strip(),
            "email":         customer_email,
            "phone":         customer_phone.strip(),
            "order_num":     order_number,
            "category":      product_category,
            "problem_type":  problem_type,
            "items":         items_to_send.strip(),
            "address":       shipping_address,
            "postal":        address_postal,
            "region":        address_region,
            "city":          address_city or client_pvz_city,
            "street":        address_street,
            "house":         address_house,
            "pvz_code":      client_pvz_code,
            "pvz_address":   client_pvz_address,
            "pvz_city":      client_pvz_city,
            "photos":        photo_files,
            "consent":       personal_data_consent,
            "wb_chat_id":        clean_chat_id,
            "wb_reply_sign":     wb_reply_sign,
            "wb_nm_id":          wb_nm_id,
            "wb_client_id":      wb_client_id,
            "salebot_client_id": clean_salebot_id,
        })
        db.commit()
        new_id = result.scalar()

        # Обновляем статус pending записи
        if clean_chat_id:
            db.execute(text("""
                UPDATE wb_chat_pending
                SET status = 'form_submitted', request_id = :req_id, updated_at = NOW()
                WHERE chat_id = :cid
            """), {"req_id": new_id, "cid": clean_chat_id})
            db.commit()

        return {"status": "success", "id": new_id, "message": "Заявка принята. Мы свяжемся с вами."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при сохранении: {e}")


@public_router.get("/confirm/{token}")
def confirm_delivery(token: str, db: Session = Depends(get_db)):
    """Клиент подтверждает получение по ссылке с токеном"""
    row = db.execute(text("""
        SELECT id, status FROM reshipment_requests WHERE confirmation_token = :token
    """), {"token": token}).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Ссылка недействительна")
    if row["status"] == "delivered":
        return {"status": "already_confirmed", "message": "Вы уже подтвердили получение. Спасибо!"}
    if row["status"] != "shipped":
        raise HTTPException(status_code=400, detail="Заявка ещё не отправлена")

    db.execute(text("""
        UPDATE reshipment_requests SET status = 'delivered', confirmed_at = NOW()
        WHERE confirmation_token = :token
    """), {"token": token})
    db.commit()
    return {"status": "success", "message": "Получение подтверждено. Спасибо!"}


@public_router.post("/link-salebot")
async def link_salebot(request: Request, db: Session = Depends(get_db)):
    """Salebot вызывает этот webhook когда клиент нажимает /start box{ID} в боте."""
    import logging
    logger = logging.getLogger(__name__)

    payload = {}
    try:
        payload = await request.json()
    except Exception:
        try:
            form = await request.form()
            payload = dict(form)
        except Exception:
            pass

    logger.warning("link-salebot payload: %s", payload)

    client_id   = str(payload.get("telegram_id") or payload.get("client_id", "")).strip()
    start_param = str(payload.get("start_param", "")).strip()

    if not client_id or not start_param.startswith("box"):
        return {"ok": True}

    try:
        req_id = int(start_param[3:])
    except ValueError:
        return {"ok": True}

    db.execute(text("""
        UPDATE reshipment_requests
        SET salebot_client_id = :cid
        WHERE id = :id AND salebot_client_id IS NULL
    """), {"cid": client_id, "id": req_id})
    db.commit()
    logger.warning("link-salebot: привязан telegram_id=%s к заявке #%s", client_id, req_id)
    return {"ok": True}


@public_router.post("/cdek-webhook")
def cdek_webhook(payload: dict, db: Session = Depends(get_db)):
    """Webhook от СДЭК: обновление статуса доставки."""
    attributes = payload.get("attributes", {})
    cdek_number = attributes.get("cdek_number", "")
    status_code = attributes.get("status_code", "")

    if not cdek_number:
        return {"ok": True}

    # Статус DELIVERED — клиент получил посылку
    if status_code in ("DELIVERED", "ACCEPTED_AT_PICK_UP_POINT"):
        db.execute(text("""
            UPDATE reshipment_requests
            SET status = 'delivered', confirmed_at = NOW()
            WHERE cdek_number = :num AND status = 'shipped'
        """), {"num": cdek_number})
        db.commit()

    return {"ok": True}


# ── Внутренние эндпоинты ──────────────────────────────────────────────────────

@router.get("/requests")
def get_requests(status: Optional[str] = Query(None), db: Session = Depends(get_db)):
    where = "WHERE 1=1"
    params: dict = {}
    if status:
        where += " AND status = :status"
        params["status"] = status

    rows = db.execute(text(f"""
        SELECT id, created_at, updated_at, customer_name, customer_email, customer_phone,
               order_number, product_category, problem_type, items_to_send,
               shipping_address, address_postal, address_region,
               address_city, address_street, address_house,
               client_pvz_code, client_pvz_address, client_pvz_city,
               photo_files, status,
               matched_srid, match_notes, moderator_comment, processed_by, processed_at,
               rejected_by, rejection_reason,
               track_number, shipping_cost, cdek_uuid, cdek_number, cdek_cost,
               shipped_at, confirmed_at, review_requested,
               wb_chat_id, wb_nm_id, wb_client_id, delivery_method,
               salebot_client_id
        FROM reshipment_requests {where}
        ORDER BY created_at DESC
    """), params).mappings().all()

    return {"status": "success", "count": len(rows), "data": [dict(r) for r in rows]}


@router.get("/requests/{req_id}")
def get_request(req_id: int, db: Session = Depends(get_db)):
    row = db.execute(text("SELECT * FROM reshipment_requests WHERE id = :id"), {"id": req_id}).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    return {"status": "success", "data": dict(row)}


@router.get("/requests/{req_id}/cost")
def get_shipping_cost(req_id: int, db: Session = Depends(get_db)):
    """Рассчитать стоимость доставки через СДЭК (для показа КС при одобрении)."""
    row = db.execute(text("""
        SELECT address_city, client_pvz_city, cdek_cost FROM reshipment_requests WHERE id = :id
    """), {"id": req_id}).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Заявка не найдена")

    if row["cdek_cost"]:
        return {"status": "success", "cost": float(row["cdek_cost"]), "cached": True}

    city = row["address_city"] or row["client_pvz_city"]
    if not city:
        return {"status": "no_city", "cost": None, "message": "Город клиента не указан"}

    try:
        from ..cdek import calculate_cost_by_city
        cost = calculate_cost_by_city(city)
        return {"status": "success", "cost": cost, "city": city, "cached": False}
    except Exception as e:
        return {"status": "error", "cost": None, "message": str(e)}


@router.post("/requests/{req_id}/approve")
def approve_request(req_id: int, payload: ApproveRequest, db: Session = Depends(get_db)):
    row = _check_exists(req_id, db)

    # 1. Создаём заказ СДЭК автоматически
    city = row.get("address_city") or row.get("client_pvz_city")
    cdek_uuid  = row.get("cdek_uuid") or ""
    cost       = row.get("cdek_cost")
    pvz_code   = row.get("client_pvz_code") or ""
    cdek_error = None

    if city and not cdek_uuid:
        try:
            from ..cdek import create_order, calculate_cost_by_city, find_nearest_pvz
            if not cost:
                cost = calculate_cost_by_city(city)
            pvz_code = pvz_code or find_nearest_pvz(city) or ""
            if pvz_code:
                cdek_result = create_order(
                    req_id=req_id,
                    customer_name=row["customer_name"],
                    customer_phone=row["customer_phone"],
                    to_pvz_code=pvz_code,
                    items_description=row.get("items_to_send", "Комплектующие"),
                )
                cdek_uuid = cdek_result["uuid"]
        except Exception as e:
            cdek_error = str(e)  # не блокируем одобрение

    db.execute(text("""
        UPDATE reshipment_requests SET
            status = 'approved', matched_srid = :srid, match_notes = :notes,
            moderator_comment = :comment, processed_by = :by, processed_at = NOW(),
            cdek_cost = COALESCE(:cost, cdek_cost),
            cdek_uuid = COALESCE(NULLIF(:uuid,''), cdek_uuid),
            client_pvz_code = COALESCE(NULLIF(:pvz,''), client_pvz_code)
        WHERE id = :id
    """), {
        "srid":    payload.matched_srid,
        "notes":   payload.match_notes,
        "comment": payload.moderator_comment,
        "by":      payload.processed_by,
        "cost":    cost,
        "uuid":    cdek_uuid,
        "pvz":     pvz_code,
        "id":      req_id,
    })
    db.commit()

    # 2. Уведомить клиента через Salebot
    salebot_notified = False
    salebot_id = row.get("salebot_client_id")
    if salebot_id:
        try:
            from ..salebot import notify_approved, is_enabled
            if is_enabled():
                salebot_notified = notify_approved(salebot_id, req_id)
        except Exception:
            pass

    result = {
        "status":           "success",
        "message":          f"Заявка #{req_id} одобрена",
        "cdek_uuid":        cdek_uuid,
        "salebot_notified": salebot_notified,
    }
    if cost is not None:
        result["cdek_cost"] = cost
    if cdek_error:
        result["cdek_warning"] = f"Заказ СДЭК не создан: {cdek_error}"
    return result


@router.post("/requests/{req_id}/reject")
def reject_request(req_id: int, payload: RejectRequest, db: Session = Depends(get_db)):
    """Отклонение заявки менеджером КС (без уведомления клиенту)."""
    _check_exists(req_id, db)
    db.execute(text("""
        UPDATE reshipment_requests SET
            status = 'rejected', moderator_comment = :comment,
            processed_by = :by, processed_at = NOW(), rejected_by = 'cs'
        WHERE id = :id
    """), {"comment": payload.moderator_comment, "by": payload.processed_by, "id": req_id})
    db.commit()
    return {"status": "success", "message": f"Заявка #{req_id} отклонена"}


@router.post("/requests/{req_id}/warehouse-reject")
def warehouse_reject(req_id: int, payload: WarehouseRejectRequest, db: Session = Depends(get_db)):
    """Отклонение заявки менеджером склада — с уведомлением клиенту."""
    row = _check_exists(req_id, db)

    db.execute(text("""
        UPDATE reshipment_requests SET
            status = 'rejected', rejection_reason = :reason,
            processed_by = :by, processed_at = NOW(), rejected_by = 'warehouse'
        WHERE id = :id
    """), {"reason": payload.rejection_reason, "by": payload.processed_by, "id": req_id})
    db.commit()

    # Отправляем email клиенту если есть адрес
    email_sent = False
    if row.get("customer_email"):
        try:
            from ..cdek import send_rejection_email
            email_sent = send_rejection_email(
                to_email=row["customer_email"],
                customer_name=row["customer_name"],
                rejection_reason=payload.rejection_reason,
                req_id=req_id,
            )
        except Exception:
            pass

    # Уведомить покупателя в WB чате (если заявка пришла через чат)
    wb_notified = False
    wb_reply_sign = row.get("wb_reply_sign")
    if wb_reply_sign:
        try:
            from ..wb_chat import send_rejection_notification, is_enabled
            if is_enabled():
                wb_notified = send_rejection_notification(wb_reply_sign, payload.rejection_reason)
        except Exception:
            pass

    # Уведомить клиента через Salebot
    salebot_notified = False
    salebot_id = row.get("salebot_client_id")
    if salebot_id:
        try:
            from ..salebot import notify_rejected, is_enabled
            if is_enabled():
                salebot_notified = notify_rejected(salebot_id, req_id, payload.rejection_reason)
        except Exception:
            pass

    return {
        "status":          "success",
        "message":         f"Заявка #{req_id} отклонена",
        "email_sent":      email_sent,
        "wb_notified":     wb_notified,
        "salebot_notified": salebot_notified,
        "customer_phone":  row.get("customer_phone"),
        "customer_name":   row.get("customer_name"),
    }


@router.post("/requests/{req_id}/warehouse-confirm")
def warehouse_confirm(req_id: int, processed_by: Optional[str] = None, db: Session = Depends(get_db)):
    """Подтверждение одиночной заявки складом — создаёт заказ в СДЭК."""
    return _do_warehouse_confirm(req_id, processed_by, db)


@router.post("/requests/warehouse-confirm-bulk")
def warehouse_confirm_bulk(payload: BulkConfirmRequest, db: Session = Depends(get_db)):
    """Массовое подтверждение заявок складом — создаёт заказы в СДЭК."""
    results = []
    for req_id in payload.ids:
        try:
            r = _do_warehouse_confirm(req_id, payload.processed_by, db)
            results.append({"id": req_id, "ok": True, "cdek_number": r.get("cdek_number", "")})
        except HTTPException as e:
            results.append({"id": req_id, "ok": False, "error": e.detail})
        except Exception as e:
            results.append({"id": req_id, "ok": False, "error": str(e)})
    return {"status": "success", "results": results}


def _do_warehouse_confirm(req_id: int, processed_by: Optional[str], db: Session) -> dict:
    """Внутренняя функция: подтвердить заявку и создать заказ в СДЭК."""
    row = _check_exists(req_id, db)

    if row["status"] != "approved":
        raise HTTPException(status_code=400, detail="Заявка должна быть в статусе 'одобрена'")

    # Определяем город и находим ближайший ПВЗ автоматически
    city = row.get("address_city") or row.get("client_pvz_city")
    if not city:
        raise HTTPException(status_code=400, detail="У заявки не указан город клиента — невозможно найти ПВЗ СДЭК")

    try:
        from ..cdek import create_order, calculate_cost_by_city, find_nearest_pvz
        pvz_code = row.get("client_pvz_code") or find_nearest_pvz(city)
        if not pvz_code:
            raise HTTPException(status_code=502, detail=f"Не найдены ПВЗ СДЭК в городе {city}")

        cdek_result = create_order(
            req_id=req_id,
            customer_name=row["customer_name"],
            customer_phone=row["customer_phone"],
            to_pvz_code=pvz_code,
            items_description=row.get("items_to_send", "Комплектующие"),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Ошибка СДЭК API: {e}")

    # Пересчитываем стоимость если не сохранена
    cost = row.get("cdek_cost")
    if not cost:
        try:
            cost = calculate_cost_by_city(city)
        except Exception:
            pass

    token = secrets.token_hex(32)
    db.execute(text("""
        UPDATE reshipment_requests SET
            status = 'shipped',
            delivery_method = 'cdek',
            cdek_uuid = :uuid,
            cdek_number = :cdek_num,
            cdek_cost = COALESCE(:cost, cdek_cost),
            client_pvz_code = COALESCE(client_pvz_code, :pvz),
            track_number = COALESCE(:cdek_num, track_number),
            shipped_at = NOW(),
            confirmation_token = :token,
            processed_by = COALESCE(:by, processed_by)
        WHERE id = :id
    """), {
        "uuid":    cdek_result["uuid"],
        "cdek_num": cdek_result["cdek_number"] or None,
        "cost":    cost,
        "pvz":     pvz_code,
        "token":   token,
        "by":      processed_by,
        "id":      req_id,
    })
    db.commit()

    confirm_url = f"https://box.vidovito.com/reshipment/confirm/{token}"

    # Уведомить покупателя в WB чате (если заявка пришла через чат)
    wb_notified = False
    wb_reply_sign = row.get("wb_reply_sign")
    if wb_reply_sign and cdek_result.get("cdek_number"):
        try:
            from ..wb_chat import send_tracking_notification, is_enabled
            if is_enabled():
                wb_notified = send_tracking_notification(
                    reply_sign=wb_reply_sign,
                    cdek_number=cdek_result["cdek_number"],
                    confirm_url=confirm_url,
                )
        except Exception:
            pass

    # Уведомить клиента через Salebot
    salebot_notified = False
    salebot_id = row.get("salebot_client_id")
    if salebot_id and cdek_result.get("cdek_number"):
        try:
            from ..salebot import notify_shipped, is_enabled
            if is_enabled():
                salebot_notified = notify_shipped(
                    salebot_id, req_id,
                    track_number=cdek_result["cdek_number"],
                    delivery_method="cdek",
                    confirm_url=confirm_url,
                )
        except Exception:
            pass

    return {
        "status":            "success",
        "message":           f"Заказ СДЭК создан для заявки #{req_id}",
        "cdek_number":       cdek_result["cdek_number"],
        "cdek_uuid":         cdek_result["uuid"],
        "cdek_cost":         cost,
        "pvz_code":          pvz_code,
        "confirmation_url":  confirm_url,
        "wb_notified":       wb_notified,
        "salebot_notified":  salebot_notified,
    }


@router.post("/requests/{req_id}/ship-yandex")
def ship_yandex(req_id: int, payload: ShipYandexRequest, db: Session = Depends(get_db)):
    """Отправка через Яндекс Доставку — ручной ввод трека и стоимости."""
    row = _check_exists(req_id, db)
    token = secrets.token_hex(32)
    db.execute(text("""
        UPDATE reshipment_requests SET
            status = 'shipped',
            delivery_method = 'yandex',
            track_number = :track,
            shipping_cost = :cost,
            shipped_at = NOW(),
            confirmation_token = :token,
            processed_by = COALESCE(:by, processed_by)
        WHERE id = :id
    """), {
        "track": payload.track_number.strip(),
        "cost":  payload.shipping_cost,
        "token": token,
        "by":    payload.processed_by,
        "id":    req_id,
    })
    db.commit()
    confirm_url = f"https://box.vidovito.com/reshipment/confirm/{token}"

    # Уведомить клиента через Salebot
    salebot_notified = False
    salebot_id = row.get("salebot_client_id")
    if salebot_id:
        try:
            from ..salebot import notify_shipped, is_enabled
            if is_enabled():
                salebot_notified = notify_shipped(
                    salebot_id, req_id,
                    track_number=payload.track_number.strip(),
                    delivery_method="yandex",
                    confirm_url=confirm_url,
                )
        except Exception:
            pass

    return {
        "status":           "success",
        "message":          f"Заявка #{req_id} отправлена через Яндекс",
        "confirmation_url": confirm_url,
        "salebot_notified": salebot_notified,
    }


@router.post("/requests/{req_id}/ship-yandex-auto")
def ship_yandex_auto(req_id: int, processed_by: Optional[str] = None, db: Session = Depends(get_db)):
    """Создать заявку Яндекс Доставки автоматически через API."""
    row = _check_exists(req_id, db)

    if row["status"] != "approved":
        raise HTTPException(status_code=400, detail="Заявка должна быть в статусе 'одобрена'")

    # Собираем адрес клиента
    city   = row.get("address_city") or row.get("client_pvz_city") or ""
    street = row.get("address_street") or ""
    house  = row.get("address_house") or ""
    if not city:
        raise HTTPException(status_code=400, detail="В заявке не указан город — невозможно создать заявку Яндекс")

    to_fullname = ", ".join(filter(None, [city, street, house]))

    try:
        from ..yandex_delivery import create_claim, is_enabled
        if not is_enabled():
            raise HTTPException(status_code=503, detail="YANDEX_DELIVERY_API_KEY не настроен")

        result = create_claim(
            req_id=req_id,
            customer_name=row["customer_name"],
            customer_phone=row["customer_phone"],
            to_fullname=to_fullname,
            items_description=row.get("items_to_send", "Комплектующие"),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Ошибка Яндекс API: {e}")

    claim_id     = result["claim_id"]
    claim_status = result["status"]
    token        = secrets.token_hex(32)

    db.execute(text("""
        UPDATE reshipment_requests SET
            status = 'shipped',
            delivery_method = 'yandex',
            track_number = :track,
            shipped_at = NOW(),
            confirmation_token = :token,
            processed_by = COALESCE(:by, processed_by)
        WHERE id = :id
    """), {
        "track": claim_id,
        "token": token,
        "by":    processed_by,
        "id":    req_id,
    })
    db.commit()

    confirm_url = f"https://box.vidovito.com/reshipment/confirm/{token}"

    # Уведомить клиента через Salebot
    salebot_notified = False
    salebot_id = row.get("salebot_client_id")
    if salebot_id and claim_status not in ("estimating_failed",):
        try:
            from ..salebot import notify_shipped, is_enabled as sb_enabled
            if sb_enabled():
                salebot_notified = notify_shipped(
                    salebot_id, req_id,
                    track_number=claim_id,
                    delivery_method="yandex",
                    confirm_url=confirm_url,
                )
        except Exception:
            pass

    return {
        "status":           "success",
        "claim_id":         claim_id,
        "claim_status":     claim_status,
        "confirmation_url": confirm_url,
        "salebot_notified": salebot_notified,
        "warning":          "Яндекс не смог рассчитать маршрут — обработайте заявку вручную в портале" if claim_status == "estimating_failed" else None,
    }


@router.post("/requests/{req_id}/request-review")
def request_review(req_id: int, db: Session = Depends(get_db)):
    _check_exists(req_id, db)
    db.execute(text("UPDATE reshipment_requests SET review_requested = TRUE WHERE id = :id"), {"id": req_id})
    db.commit()
    return {"status": "success"}


@router.post("/register-cdek-webhook")
def register_cdek_webhook(db: Session = Depends(get_db)):
    """Регистрирует webhook в СДЭК (вызвать один раз после деплоя)."""
    try:
        from ..cdek import register_webhook
        result = register_webhook("https://cxvo.ru/api/v1/reshipment/cdek-webhook")
        return {"status": "success", "data": result}
    except Exception as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.get("/stats")
def get_stats(db: Session = Depends(get_db)):
    rows = db.execute(text("""
        SELECT status, COUNT(*) as cnt,
               COALESCE(SUM(cdek_cost), SUM(shipping_cost), 0) as total_cost
        FROM reshipment_requests GROUP BY status
    """)).mappings().all()
    return {"status": "success", "data": [dict(r) for r in rows]}


# ── Хелпер ───────────────────────────────────────────────────────────────────


import io
import datetime
from fastapi.responses import StreamingResponse


# ── Склад: очередь заявок к отправке ────────────────────────────────────────

@router.get("/warehouse/queue")
def warehouse_queue(db: Session = Depends(get_db)):
    """Список одобренных заявок для менеджера склада."""
    rows = db.execute(text("""
        SELECT id, customer_name, customer_phone,
               address_city, address_street, address_house,
               client_pvz_code, client_pvz_address,
               items_to_send, cdek_uuid, cdek_number, cdek_cost,
               moderator_comment, processed_at, processed_by,
               status, matched_srid, shipped_at
        FROM reshipment_requests
        WHERE status = 'approved'
        ORDER BY processed_at ASC
    """)).mappings().all()
    return {"status": "success", "count": len(rows), "data": [dict(r) for r in rows]}


# ── Склад: список к упаковке (Excel) ────────────────────────────────────────

@router.get("/warehouse/packlist")
def warehouse_packlist(db: Session = Depends(get_db)):
    """Скачать список к отправке в формате Excel."""
    rows = db.execute(text("""
        SELECT id, customer_name, customer_phone,
               address_city, client_pvz_code, client_pvz_address,
               items_to_send, cdek_uuid, cdek_cost,
               matched_srid, moderator_comment, processed_at
        FROM reshipment_requests
        WHERE status = 'approved'
        ORDER BY processed_at ASC
    """)).mappings().all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "К отправке"

    header_fill = PatternFill("solid", fgColor="4F46E5")
    hdr_font    = Font(bold=True, color="FFFFFF", size=10)
    thin        = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = [
        ("№", 5), ("Заявка", 8), ("Клиент", 22), ("Телефон", 14),
        ("Город", 16), ("ПВЗ СДЭК", 12), ("Адрес ПВЗ", 35),
        ("Что отправить", 40), ("SRID", 14), ("Стоимость", 11),
        ("UUID СДЭК", 38), ("Одобрено", 16),
    ]
    for col, (title, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font      = hdr_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    ws.row_dimensions[1].height = 22

    alt_fill = PatternFill("solid", fgColor="F0F0FF")
    for i, row in enumerate(rows, 1):
        fill = alt_fill if i % 2 == 0 else None
        dt = row["processed_at"]
        dt_str = dt.strftime("%d.%m %H:%M") if dt else ""
        vals = [
            i,
            row["id"],
            row["customer_name"] or "",
            row["customer_phone"] or "",
            row["address_city"] or "",
            row["client_pvz_code"] or "",
            row["client_pvz_address"] or "",
            row["items_to_send"] or "",
            row["matched_srid"] or "",
            float(row["cdek_cost"]) if row["cdek_cost"] else "",
            row["cdek_uuid"] or "",
            dt_str,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=i+1, column=col, value=val)
            cell.border    = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(col in (8,)))
            if fill:
                cell.fill = fill
        ws.row_dimensions[i+1].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.date.today().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="packlist_{today}.xlsx"'},
    )


# ── Склад: PDF этикетки 58×40мм ─────────────────────────────────────────────

@router.get("/warehouse/labels")
def warehouse_labels(ids: Optional[str] = Query(None), db: Session = Depends(get_db)):
    """PDF с этикетками 58×40мм. ids=1,2,3 или все approved."""
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        where   = "WHERE id = ANY(:ids) AND status IN ('approved','shipped')"
        params  = {"ids": id_list}
    else:
        where  = "WHERE status = 'approved'"
        params = {}

    rows = db.execute(text(f"""
        SELECT id, customer_name, address_city, client_pvz_city,
               items_to_send, cdek_uuid, cdek_number, client_pvz_code
        FROM reshipment_requests {where}
        ORDER BY id ASC
    """), params).mappings().all()

    if not rows:
        raise HTTPException(status_code=404, detail="Нет заявок для печати")

    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import mm
    from reportlab.graphics.barcode import code128
    from reportlab.lib import colors

    W = 58 * mm
    H = 40 * mm

    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=(W, H))

    for order in rows:
        req_id  = order["id"]
        name    = (order["customer_name"] or "")[:24]
        city    = (order["address_city"] or order["client_pvz_city"] or "—")[:20]
        items   = (order["items_to_send"] or "")[:36]
        pvz     = order["client_pvz_code"] or ""
        cdek_no = order["cdek_number"] or ""

        barcode_val = f"REQ{req_id:06d}"
        try:
            bc = code128.Code128(
                barcode_val,
                barWidth=0.55 * mm,
                barHeight=14 * mm,
                quiet=False,
                checksum=False,
            )
            bc_w = bc.width
            x_bc = (W - bc_w) / 2
            bc.drawOn(c, x_bc, 22 * mm)
        except Exception:
            c.setFont("Helvetica-Bold", 8)
            c.drawCentredString(W / 2, 28 * mm, barcode_val)

        c.setFont("Helvetica", 5.5)
        c.setFillColor(colors.HexColor("#555555"))
        c.drawCentredString(W / 2, 20 * mm, barcode_val)

        c.setStrokeColor(colors.HexColor("#CCCCCC"))
        c.setLineWidth(0.3)
        c.line(2 * mm, 19 * mm, W - 2 * mm, 19 * mm)

        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 7.5)
        c.drawString(2 * mm, 15.5 * mm, f"#{req_id}  {city}")
        if pvz:
            c.setFont("Helvetica", 6.5)
            c.drawString(2 * mm, 11.5 * mm, f"ПВЗ: {pvz}")
        c.setFont("Helvetica", 6.5)
        c.drawString(2 * mm, 8 * mm, name)
        c.setFont("Helvetica", 6)
        c.setFillColor(colors.HexColor("#444444"))
        c.drawString(2 * mm, 4.5 * mm, items[:50])
        if cdek_no:
            c.setFont("Helvetica-Bold", 5.5)
            c.setFillColor(colors.HexColor("#2563EB"))
            c.drawString(2 * mm, 1.5 * mm, f"СДЭК: {cdek_no}")

        c.showPage()

    c.save()
    buf.seek(0)

    today = datetime.date.today().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="labels_{today}.pdf"'},
    )


# ── Склад: отметить как физически отправленную ────────────────────────────────

@router.post("/requests/{req_id}/mark-shipped")
def mark_shipped(req_id: int, processed_by: Optional[str] = None, db: Session = Depends(get_db)):
    """Отмечает что посылка физически сдана на ПВЗ СДЭК."""
    row = _check_exists(req_id, db)
    if row["status"] != "approved":
        raise HTTPException(status_code=400, detail="Заявка должна быть в статусе 'одобрена'")

    token = secrets.token_hex(32)
    db.execute(text("""
        UPDATE reshipment_requests SET
            status = 'shipped',
            shipped_at = NOW(),
            confirmation_token = COALESCE(confirmation_token, :token),
            processed_by = COALESCE(:by, processed_by)
        WHERE id = :id
    """), {"token": token, "by": processed_by, "id": req_id})
    db.commit()

    return {"status": "success", "message": f"Заявка #{req_id} отмечена как отправленная"}


def _check_exists(req_id: int, db: Session) -> dict:
    row = db.execute(
        text("SELECT * FROM reshipment_requests WHERE id = :id"), {"id": req_id}
    ).mappings().first()
    if not row:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    return dict(row)
