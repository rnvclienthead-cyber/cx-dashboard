import io
import json
import time
import base64
import requests
import os
import sys
import uuid
import concurrent.futures
import pandas as pd
import numpy as np
from datetime import datetime
from urllib.parse import quote
from PIL import Image as PILImage

if sys.platform == 'darwin':
    homebrew_lib = '/opt/homebrew/lib'
    if os.path.exists(homebrew_lib):
        os.environ['DYLD_FALLBACK_LIBRARY_PATH'] = f"{homebrew_lib}:{os.environ.get('DYLD_FALLBACK_LIBRARY_PATH', '')}"

from fastapi import APIRouter, Depends, HTTPException, Body, Query
from fastapi.responses import HTMLResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import text
from pydantic import BaseModel
from typing import Dict, Any, List, Optional

from weasyprint import HTML
from weasyprint.text.fonts import FontConfiguration

from ..database import get_db
from .auth import get_current_user

router = APIRouter(prefix="/api/v1/analytics", tags=["PPM Analytics"], dependencies=[Depends(get_current_user)])

class ClaimExportRequest(BaseModel):
    supplier: str; period: str; sku: str; name: str; name_cn: str
    defects: int; ppm_pct: float; desc_ru: str; desc_cn: str; cause_ru: str; cause_cn: str
    photo_groups: List[Dict[str, Any]] = []; chart_data: List[Dict[str, Any]] = []
    invoices_list: List[Dict[str, Any]] = []
    is_redownload: bool = False; existing_act_number: Optional[str] = None; is_test: bool = False
    # Новые поля реестра
    report_month: Optional[str] = None
    abc_group: Optional[str] = None
    product_name_ru: Optional[str] = None
    invoice_ref: Optional[str] = None
    deviation: Optional[str] = None
    claim_status: Optional[str] = 'На проверке'
    initiator: Optional[str] = 'Wildberries'
    factory_type: Optional[str] = 'завод'
    who_sent: Optional[str] = None

class ReplyRequest(BaseModel):
    reply: str

class ManualClaimModel(BaseModel):
    send_date: str; send_text: str
    invoice_ref: Optional[str] = None
    factory_name: Optional[str] = None
    container_num: Optional[str] = None
    manual_status: Optional[str] = None
    who_sent: Optional[str] = None

class ManualClaimReply(BaseModel):
    reply_date: Optional[str] = None
    reply_text: Optional[str] = None
    invoice_ref: Optional[str] = None
    factory_name: Optional[str] = None
    container_num: Optional[str] = None
    manual_status: Optional[str] = None
    who_sent: Optional[str] = None

class PhotoFilterRequest(BaseModel):
    urls: List[str]

class ClaimLogUpdate(BaseModel):
    report_month: Optional[str] = None
    initiator: Optional[str] = None
    product_name: Optional[str] = None
    qty: Optional[int] = None
    invoice_ref: Optional[str] = None
    ra_date: Optional[str] = None
    batch_num: Optional[str] = None
    production_date: Optional[str] = None
    sale_date: Optional[str] = None
    stage: Optional[str] = None
    object_type: Optional[str] = None
    controlled_params: Optional[str] = None
    deviation: Optional[str] = None
    deviation_desc: Optional[str] = None
    repeatability: Optional[str] = None
    claim_status: Optional[str] = None
    deviation_cause: Optional[str] = None
    abc_group: Optional[str] = None
    factory_type: Optional[str] = None
    who_sent: Optional[str] = None
    send_date: Optional[str] = None
    chat_name: Optional[str] = None
    send_status: Optional[str] = None
    factory_reply_date: Optional[str] = None
    factory_reply: Optional[str] = None
    comments: Optional[str] = None
    act_number: Optional[str] = None
    estimated_improvement_date: Optional[str] = None
    correction_invoice: Optional[str] = None

_PPM_CACHE = {}

def get_from_cache(key: str, ttl: int = 5):
    now = time.time()
    if key in _PPM_CACHE:
        data, ts = _PPM_CACHE[key]
        if now - ts < ttl: return data
    return None

def save_to_cache(key: str, data: any):
    _PPM_CACHE[key] = (data, time.time())

def init_log_table(db: Session):
    query = text("""
        CREATE TABLE IF NOT EXISTS cx_claim_logs (
            id SERIAL PRIMARY KEY, act_number VARCHAR(50) UNIQUE NOT NULL, sku VARCHAR(100), supplier VARCHAR(255),
            defects_count INT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, status VARCHAR(50) DEFAULT 'Активен',
            factory_reply TEXT, pdf_payload JSONB, period_str VARCHAR(255)
        );
        ALTER TABLE cx_claim_logs ADD COLUMN IF NOT EXISTS factory_reply TEXT;
        ALTER TABLE cx_claim_logs ADD COLUMN IF NOT EXISTS pdf_payload JSONB;
        ALTER TABLE cx_claim_logs ADD COLUMN IF NOT EXISTS period_str VARCHAR(255);
        ALTER TABLE cx_claim_logs ADD COLUMN IF NOT EXISTS estimated_improvement_date DATE;
        ALTER TABLE cx_claim_logs ADD COLUMN IF NOT EXISTS correction_invoice VARCHAR(255);
    """)
    with db.bind.connect() as conn:
        conn.execute(query)
        conn.commit()

def init_manual_claims_table(db: Session):
    query = text("""
        CREATE TABLE IF NOT EXISTS cx_manual_claims (
            id SERIAL PRIMARY KEY, ticket_number VARCHAR(50) UNIQUE NOT NULL, send_date VARCHAR(50),
            send_text TEXT, reply_date VARCHAR(50), reply_text TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        ALTER TABLE cx_manual_claims ADD COLUMN IF NOT EXISTS invoice_ref VARCHAR(255);
        ALTER TABLE cx_manual_claims ADD COLUMN IF NOT EXISTS factory_name VARCHAR(255);
        ALTER TABLE cx_manual_claims ADD COLUMN IF NOT EXISTS container_num VARCHAR(255);
        ALTER TABLE cx_manual_claims ADD COLUMN IF NOT EXISTS manual_status VARCHAR(100);
        ALTER TABLE cx_manual_claims ADD COLUMN IF NOT EXISTS who_sent VARCHAR(255);
    """)
    with db.bind.connect() as conn:
        conn.execute(query)
        conn.commit()

_LOG_DATE_FIELDS = ['ra_date', 'production_date', 'sale_date', 'send_date', 'factory_reply_date']

def _format_log_row(r: dict) -> dict:
    d = dict(r)
    if d.get('created_at'): d['created_at'] = d['created_at'].strftime('%d.%m.%Y %H:%M')
    for f in _LOG_DATE_FIELDS:
        if d.get(f) and hasattr(d[f], 'strftime'): d[f] = d[f].strftime('%Y-%m-%d')
    d['factory_reply'] = d.get('factory_reply') or ""
    d['period_str'] = d.get('period_str') or ""
    return d

@router.get("/claim-logs")
def get_claim_logs(db: Session = Depends(get_db)):
    init_log_table(db)
    try:
        with db.bind.connect() as conn:
            rows = conn.execute(text("""
                SELECT * FROM cx_claim_logs
                ORDER BY
                  CASE
                    WHEN act_number = 'б/н' THEN 109.5
                    WHEN act_number ~ '^[0-9]'
                      THEN CAST(REGEXP_REPLACE(SPLIT_PART(act_number, ',', 1), '[^0-9]', '', 'g') AS FLOAT)
                    ELSE 0
                  END DESC
                LIMIT 500
            """)).mappings().all()
            return {"status": "success", "data": [_format_log_row(dict(r)) for r in rows]}
    except Exception as e: return {"status": "error", "message": str(e), "data": []}

@router.get("/claim-logs/by-sku/{sku}")
def get_claim_logs_by_sku(sku: str, db: Session = Depends(get_db)):
    """Акты по конкретному SKU — для маркеров на графике."""
    try:
        with db.bind.connect() as conn:
            rows = conn.execute(text(
                "SELECT id, act_number, send_date, ra_date, created_at, factory_reply FROM cx_claim_logs "
                "WHERE LOWER(TRIM(sku)) = LOWER(TRIM(:sku)) AND status != 'Аннулирован' "
                "ORDER BY COALESCE(send_date, ra_date::date, created_at::date) ASC"
            ), {"sku": sku}).mappings().all()
            data = []
            for r in rows:
                d = dict(r)
                for f in ['send_date', 'ra_date']:
                    if d.get(f) and hasattr(d[f], 'strftime'): d[f] = d[f].strftime('%Y-%m-%d')
                if d.get('created_at') and hasattr(d['created_at'], 'strftime'): d['created_at'] = d['created_at'].strftime('%Y-%m-%d')
                data.append(d)
            return {"status": "success", "data": data}
    except Exception as e: return {"status": "success", "data": []}

@router.put("/claim-logs/{log_id}/cancel")
def cancel_claim_log(log_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    try:
        with db.bind.connect() as conn:
            row = conn.execute(text("SELECT send_date, send_status FROM cx_claim_logs WHERE id = :id"), {"id": log_id}).mappings().fetchone()
            if row:
                if row['send_date'] or (row['send_status'] and row['send_status'] not in ('', 'Не отправлен', None)):
                    raise HTTPException(status_code=400, detail="Акт уже отправлен на завод — аннулирование заблокировано.")
            conn.execute(text("UPDATE cx_claim_logs SET status = 'Аннулирован' WHERE id = :id"), {"id": log_id})
            conn.commit()
        return {"status": "success"}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@router.delete("/claim-logs/{log_id}")
def delete_claim_log(log_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Полное удаление акта — только для администратора."""
    if current_user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Только администратор может удалять акты.")
    try:
        with db.bind.connect() as conn:
            conn.execute(text("DELETE FROM cx_claim_logs WHERE id = :id"), {"id": log_id})
            conn.commit()
        return {"status": "success"}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@router.put("/claim-logs/{log_id}/unlock")
def unlock_claim_log(log_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Разблокировать акт — сбросить статус отправки. Только для администратора."""
    if current_user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Только администратор может разблокировать акты.")
    try:
        with db.bind.connect() as conn:
            conn.execute(text("UPDATE cx_claim_logs SET send_status = 'Не отправлен', send_date = NULL WHERE id = :id"), {"id": log_id})
            conn.commit()
        return {"status": "success"}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@router.put("/claim-logs/{log_id}")
def update_claim_log(log_id: int, req: ClaimLogUpdate, db: Session = Depends(get_db)):
    """Обновление всех редактируемых полей акта."""
    fields = req.dict(exclude_none=True)
    if not fields: return {"status": "success"}
    # Автоматически ставим дату отправки если статус меняется на "Отправлен"
    if fields.get('send_status') == 'Отправлен' and 'send_date' not in fields:
        fields['send_date'] = datetime.now().date()
    try:
        set_parts = [f"{k} = :{k}" for k in fields]
        fields['id'] = log_id
        with db.bind.connect() as conn:
            conn.execute(text(f"UPDATE cx_claim_logs SET {', '.join(set_parts)} WHERE id = :id"), fields)
            conn.commit()
        return {"status": "success"}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@router.put("/claim-logs/{log_id}/reply")
def update_claim_reply(log_id: int, req: ReplyRequest, db: Session = Depends(get_db)):
    try:
        with db.bind.connect() as conn:
            conn.execute(text("UPDATE cx_claim_logs SET factory_reply = :reply WHERE id = :id"), {"reply": req.reply, "id": log_id})
            conn.commit()
        return {"status": "success"}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@router.get("/manual-claims")
def get_manual_claims(db: Session = Depends(get_db)):
    init_manual_claims_table(db)
    try:
        with db.bind.connect() as conn:
            rows = conn.execute(text("SELECT * FROM cx_manual_claims ORDER BY created_at DESC")).mappings().all()
            return {"status": "success", "data": [dict(r) for r in rows]}
    except Exception as e: return {"status": "error", "message": str(e), "data": []}

@router.post("/manual-claims")
def create_manual_claim(req: ManualClaimModel, db: Session = Depends(get_db)):
    init_manual_claims_table(db)
    try:
        with db.bind.connect() as conn:
            res = conn.execute(text("""
                INSERT INTO cx_manual_claims (ticket_number, send_date, send_text, invoice_ref, factory_name, container_num, manual_status, who_sent)
                VALUES ('M-TEMP', :sd, :st, :ir, :fn, :cn, :ms, :ws) RETURNING id
            """), {"sd": req.send_date, "st": req.send_text, "ir": req.invoice_ref,
                  "fn": req.factory_name, "cn": req.container_num, "ms": req.manual_status, "ws": req.who_sent})
            new_id = res.scalar()
            t_num = f"MC-{datetime.now().strftime('%y%m')}-{new_id:04d}"
            conn.execute(text("UPDATE cx_manual_claims SET ticket_number = :t WHERE id = :id"), {"t": t_num, "id": new_id})
            conn.commit()
        return {"status": "success"}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@router.put("/manual-claims/{claim_id}")
def update_manual_claim(claim_id: int, req: ManualClaimReply, db: Session = Depends(get_db)):
    try:
        with db.bind.connect() as conn:
            conn.execute(text("""
                UPDATE cx_manual_claims SET
                    reply_date = :rd, reply_text = :rt,
                    invoice_ref = COALESCE(:invoice_ref, invoice_ref),
                    factory_name = COALESCE(:factory_name, factory_name),
                    container_num = COALESCE(:container_num, container_num),
                    manual_status = COALESCE(:manual_status, manual_status),
                    who_sent = COALESCE(:who_sent, who_sent)
                WHERE id = :id
            """), {
                "rd": req.reply_date, "rt": req.reply_text, "id": claim_id,
                "invoice_ref": req.invoice_ref, "factory_name": req.factory_name,
                "container_num": req.container_num, "manual_status": req.manual_status,
                "who_sent": req.who_sent
            })
            conn.commit()
        return {"status": "success"}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

def compute_abc_xyz(df_monthly: pd.DataFrame,
                    df_weekly: pd.DataFrame = None,
                    window_months: int = 12) -> pd.DataFrame:
    """Динамический расчёт ABC (по заказам) и XYZ (по вариативности) из wb_orders.

    ABC — метод cumulative-share (те же пороги, что у Litestat):
      A = артикулы, формирующие первые 80% суммарного объёма
      B = следующие 15% (80–95%)
      C = оставшиеся 5%

    XYZ — коэффициент вариации (CV = std/mean):
      X = CV < 0.25  — стабильный спрос
      Y = 0.25–0.50  — умеренная вариативность
      Z ≥ 0.50       — нестабильный / сезонный

    ABC считается по месячным данным за window_months.
    XYZ считается по df_weekly (недельные отрезки, ~26 точек на артикул) —
    это ближе к методологии Litestat (10 отрезков по ~10 дней). Если
    df_weekly не передан — используются месячные данные как fallback.
    """
    if df_monthly.empty:
        return pd.DataFrame(columns=['article', 'abc_group', 'xyz_class'])

    ord_col = 'orders' if 'orders' in df_monthly.columns else 'pure_orders'

    # ── ABC: суммарные заказы за окно ────────────────────────────────────
    cutoff = pd.Timestamp.now().normalize() - pd.DateOffset(months=window_months)
    df_abc_src = df_monthly[df_monthly['month_dt'] >= cutoff].copy()
    if df_abc_src.empty:
        df_abc_src = df_monthly.copy()

    totals = (df_abc_src.groupby('article')[ord_col].sum()
              .reset_index(name='total')
              .sort_values('total', ascending=False)
              .reset_index(drop=True))
    grand = totals['total'].sum()
    if grand > 0:
        totals['cum_pct'] = totals['total'].cumsum() / grand
        totals['prev_cum'] = totals['cum_pct'].shift(1, fill_value=0.0)
    else:
        totals['cum_pct'] = 0.0
        totals['prev_cum'] = 0.0
    totals['abc_group'] = 'C'
    totals.loc[totals['prev_cum'] < 0.80, 'abc_group'] = 'A'
    totals.loc[(totals['prev_cum'] >= 0.80) & (totals['prev_cum'] < 0.95), 'abc_group'] = 'B'

    # ── XYZ: коэффициент вариации ─────────────────────────────────────────
    # Приоритет — недельные данные (ближе к Litestat-методологии).
    # Fallback — месячные (если df_weekly не передан или пуст).
    if df_weekly is not None and not df_weekly.empty:
        # Отсекаем "нерабочие" недели: периоды когда суммарный объём < 10% от
        # пиковой недели. Это убирает паразитные исторические/тестовые записи
        # (1–3 заказа в неделю) которые иначе вздувают CV до Z для всех артикулов.
        week_totals = df_weekly.groupby('week_dt')['orders'].sum()
        threshold = week_totals.max() * 0.10
        active_weeks = week_totals[week_totals >= threshold].index
        xyz_src = df_weekly[df_weekly['week_dt'].isin(active_weeks)]
        time_col = 'week_dt'
        xyz_ord = 'orders'
    else:
        xyz_src = df_abc_src
        time_col = 'month_dt'
        xyz_ord = ord_col

    pivot = xyz_src.pivot_table(index='article', columns=time_col,
                                values=xyz_ord, aggfunc='sum', fill_value=0)
    means = pivot.mean(axis=1).replace(0, float('nan'))
    cv = (pivot.std(axis=1) / means).fillna(0.0)
    xyz = pd.DataFrame({'article': cv.index, 'cv': cv.values})
    xyz['xyz_class'] = 'Z'
    xyz.loc[xyz['cv'] < 0.25, 'xyz_class'] = 'X'
    xyz.loc[(xyz['cv'] >= 0.25) & (xyz['cv'] < 0.50), 'xyz_class'] = 'Y'

    # ── Объединяем ────────────────────────────────────────────────────────
    result = (totals[['article', 'abc_group']]
              .merge(xyz[['article', 'xyz_class']], on='article', how='outer'))
    result['abc_group'] = result['abc_group'].fillna('C')
    result['xyz_class'] = result['xyz_class'].fillna('Z')
    return result

def load_hybrid_data(db: Session, platform: str = 'wb') -> pd.DataFrame:
    if platform == 'ym':
        return _load_ym_claims(db)
    query = text('SELECT * FROM view_wb_claims_enriched')
    try:
        with db.bind.connect() as conn:
            df = pd.read_sql(query, conn)
        if df.empty: return df
        if 'supplier_article' in df.columns:
            df['supplier_article'] = df['supplier_article'].fillna('Без артикула').astype(str).str.strip()
        if 'created_dt' in df.columns:
            df['created_dt'] = pd.to_datetime(df['created_dt'], errors='coerce').dt.strftime('%Y-%m-%d')
        if 'claim_date' in df.columns:
            df['claim_date'] = pd.to_datetime(df['claim_date'], errors='coerce').dt.strftime('%Y-%m-%d')
        return df
    except Exception as e:
        print(f"Ошибка load_hybrid_data: {e}")
        return pd.DataFrame()

def _load_ym_claims(db: Session) -> pd.DataFrame:
    """Возвраты ЯМ в формате view_wb_claims_enriched для совместимости с PPM."""
    try:
        with db.bind.connect() as conn:
            df = pd.read_sql(text("""
                SELECT
                    r.return_id::text          AS srid,
                    r.created_at               AS created_dt,
                    r.supplier_article,
                    r.return_comment           AS user_comment,
                    r.status_ru                AS status,
                    NULL::text                 AS status_ex,
                    NULL::text                 AS income_id,
                    r.photos,
                    NULL::text                 AS video_paths,
                    r.cat_1,  r.cat_2,  r.cat_3,  r.cat_4,  r.cat_5,
                    r.cat_6,  r.cat_7,  r.cat_8,  r.cat_9,  r.cat_10,
                    r.cat_11, r.cat_12, r.cat_13,
                    NULL::date                 AS order_date,
                    r.created_at               AS claim_date,
                    f.factory_name,
                    NULL::text                 AS container_num,
                    NULL::date                 AS shipment_date,
                    inv.invoice_num
                FROM ym_returns r
                LEFT JOIN (
                    SELECT DISTINCT ON (supplier_article)
                        supplier_article, invoice_num
                    FROM wb_invoices
                    WHERE marketplace = 'ym'
                      AND invoice_num NOT IN ('', '0')
                    ORDER BY supplier_article, supply_id DESC
                ) inv ON inv.supplier_article = r.supplier_article
                LEFT JOIN wb_factories f
                    ON trim(lower(f.invoice_num)) = trim(lower(inv.invoice_num))
                   AND trim(lower(f.supplier_article)) = trim(lower(r.supplier_article))
                WHERE r.return_type = 'RETURN' OR r.return_type IS NULL
                ORDER BY r.created_at DESC
            """), conn)
        if df.empty: return df
        df['supplier_article'] = df['supplier_article'].fillna('Без артикула').astype(str).str.strip()
        df['created_dt'] = pd.to_datetime(df['created_dt'], errors='coerce').dt.strftime('%Y-%m-%d')
        df['claim_date']  = pd.to_datetime(df['claim_date'],  errors='coerce').dt.strftime('%Y-%m-%d')
        return df
    except Exception as e:
        print(f"Ошибка _load_ym_claims: {e}")
        return pd.DataFrame()

def load_orders_data(db: Session, platform: str = 'wb') -> pd.DataFrame:
    if platform == 'ym':
        query = text("""
            SELECT TRIM(supplier_article) AS supplier_article,
                   DATE_TRUNC('month', created_at) AS month_dt,
                   COUNT(*) AS pure_orders
            FROM ym_orders
            WHERE status NOT IN ('CANCELLED','PENDING_CANCELLED','PROCESSING_EXPIRED')
            GROUP BY 1, 2
        """)
    else:
        query = text("""
            SELECT TRIM(supplier_article) AS supplier_article,
                   DATE_TRUNC('month', CAST(dt AS TIMESTAMP)) AS month_dt, COUNT(*) AS pure_orders
            FROM wb_orders WHERE cancel_dt IS NULL GROUP BY 1, 2
        """)
    try:
        with db.bind.connect() as conn:
            df = pd.read_sql(query, conn)
        if not df.empty: df['month_dt'] = pd.to_datetime(df['month_dt'])
        return df
    except Exception as e:
        return pd.DataFrame(columns=['supplier_article', 'month_dt', 'pure_orders'])

@router.get("/orders-by-factory")
def get_orders_by_factory(
    platform: str = Query("wb", pattern="^(wb|ym|ozon|all)$"),
    db: Session = Depends(get_db)
):
    """Чистые заказы с разбивкой по заводам."""
    if platform == 'ym':
        query = text("""
            SELECT
                TRIM(o.supplier_article) AS article,
                DATE_TRUNC('month', o.created_at) AS month_dt,
                COALESCE(f.factory_name, 'Уточняется') AS factory_name,
                COUNT(*) AS orders
            FROM ym_orders o
            LEFT JOIN (
                SELECT DISTINCT ON (supplier_article)
                    supplier_article, invoice_num
                FROM wb_invoices
                WHERE marketplace = 'ym'
                  AND invoice_num NOT IN ('', '0')
                ORDER BY supplier_article, supply_id DESC
            ) inv ON inv.supplier_article = o.supplier_article
            LEFT JOIN wb_factories f
                ON trim(lower(f.invoice_num)) = trim(lower(inv.invoice_num))
               AND trim(lower(f.supplier_article)) = trim(lower(o.supplier_article))
            WHERE o.status NOT IN ('CANCELLED','PENDING_CANCELLED','PROCESSING_EXPIRED')
            GROUP BY 1, 2, 3
        """)
        try:
            with db.bind.connect() as conn:
                df = pd.read_sql(query, conn)
            if df.empty: return {"status": "success", "data": []}
            df['month_dt'] = pd.to_datetime(df['month_dt']).dt.strftime('%Y-%m-%d')
            df['factory_name'] = df['factory_name'].fillna('Уточняется')
            df = df.replace({np.nan: None})
            return {"status": "success", "data": df.to_dict(orient='records')}
        except Exception as e:
            print("Ошибка orders-by-factory YM:", e)
            return {"status": "success", "data": []}
    TR = "translate(lower(trim({})),'авсенкмортху','abcehkmoptxy')"
    query = text(f"""
        SELECT art AS article, month_dt, factory_name, COUNT(*) AS orders
        FROM (
          SELECT TRIM(o.supplier_article) AS art,
                 DATE_TRUNC('month', CAST(o.dt AS TIMESTAMP)) AS month_dt,
                 o.srid,
                 -- Приоритет: 1) завод через цепочку инвойс→wb_factories,
                 -- 2) завод из ассортиментной матрицы — ТОЛЬКО если у артикула один завод.
                 -- Если в wb_factories несколько заводов — не используем assortment fallback,
                 -- чтобы заказы без цепочки оставались NULL (Уточняется) и во фронтенде
                 -- прикреплялись к заводу с наибольшим числом претензий.
                 COALESCE(
                   MAX(f.factory_name),
                   CASE WHEN mf.supplier_article IS NULL THEN MAX(a.manufacturer) ELSE NULL END
                 ) AS factory_name
          FROM wb_orders o
          LEFT JOIN wb_logistics l ON l.srid = o.srid
          LEFT JOIN wb_invoices inv
              ON split_part(l.income_id::text,'.',1) = split_part(inv.supply_id::text,'.',1)
             AND {TR.format('inv.supplier_article')} = {TR.format('o.supplier_article')}
          LEFT JOIN wb_factories f
              ON trim(lower(f.invoice_num)) = trim(lower(inv.invoice_num))
             AND {TR.format('f.supplier_article')} = {TR.format('o.supplier_article')}
          LEFT JOIN wb_assortment a
              ON {TR.format('a.supplier_article')} = {TR.format('o.supplier_article')}
             AND a.manufacturer IS NOT NULL AND a.manufacturer <> ''
          LEFT JOIN (
              SELECT {TR.format('supplier_article')} AS supplier_article
              FROM wb_factories
              GROUP BY {TR.format('supplier_article')}
              HAVING COUNT(DISTINCT factory_name) > 1
          ) mf ON mf.supplier_article = {TR.format('o.supplier_article')}
          WHERE o.cancel_dt IS NULL
          GROUP BY 1, 2, o.srid, mf.supplier_article
        ) s
        GROUP BY art, month_dt, factory_name
    """)
    try:
        with db.bind.connect() as conn:
            df = pd.read_sql(query, conn)
        if df.empty:
            return {"status": "success", "data": []}
        df['month_dt'] = pd.to_datetime(df['month_dt']).dt.strftime('%Y-%m-%d')
        df['factory_name'] = df['factory_name'].fillna('Уточняется')
        df = df.replace({np.nan: None})
        return {"status": "success", "data": df.to_dict(orient='records')}
    except Exception as e:
        print("Ошибка orders-by-factory:", e)
        return {"status": "success", "data": []}

@router.get("/assortment-names")
def get_assortment_names(db: Session = Depends(get_db)):
    """Возвращает карту supplier_article → name_ru для подстановки в акт."""
    try:
        rows = db.execute(text("SELECT supplier_article, name_ru FROM wb_assortment WHERE name_ru IS NOT NULL AND name_ru != ''")).mappings().all()
        return {"status": "success", "data": {r['supplier_article']: r['name_ru'] for r in rows}}
    except Exception as e:
        return {"status": "success", "data": {}}

@router.get("/ppm-claims")
def get_production_claims(
    platform: str = Query("wb", pattern="^(wb|ym|ozon|all)$"),
    db: Session = Depends(get_db)
):
    df = load_hybrid_data(db, platform)
    if df.empty: return {"status": "success", "data": []}
    df = df.replace({np.nan: None})
    return {"status": "success", "data": df.to_dict(orient="records")}


def _analyze_photo(url: str) -> dict:
    """Анализирует одно фото: детектирует штрих-коды и оценивает «коробочность»."""
    result = {"url": url, "keep": True, "reason": "ok"}
    try:
        resp = requests.get(url, timeout=8, headers={"User-Agent": "Mozilla/5.0"})
        if not resp.ok:
            return result  # не смогли скачать — не фильтруем

        img = PILImage.open(io.BytesIO(resp.content)).convert("RGB")

        # 1. Детектирование штрих-кода / QR-кода через pyzbar
        try:
            from pyzbar.pyzbar import decode as pyzbar_decode
            if len(pyzbar_decode(img)) > 0:
                result["keep"] = False; result["reason"] = "barcode"; return result
        except Exception:
            pass

        img_small = img.resize((80, 80))
        pixels = list(img_small.getdata())
        total = len(pixels)
        r_vals = [p[0] for p in pixels]
        g_vals = [p[1] for p in pixels]
        b_vals = [p[2] for p in pixels]
        brightness_vals = [(p[0]+p[1]+p[2])/3 for p in pixels]
        avg_brightness = sum(brightness_vals) / total

        # 2. Пустое/чёрное фото
        if avg_brightness > 248 or avg_brightness < 8:
            result["keep"] = False; result["reason"] = "blank"; return result

        # 3. Определяем «упаковочные» пиксели (белые, светло-серые, картон, крафт, бежевые)
        def is_packaging(r, g, b):
            brightness = (r + g + b) / 3
            saturation = max(r, g, b) - min(r, g, b)
            # Белый/светло-серый
            if brightness > 195 and saturation < 35: return True
            # Картонный/бежевый (R > G > B, умеренная яркость)
            if r > 150 and g > 120 and b > 80 and r > g > b and saturation < 60 and brightness > 140: return True
            # Крафт-бумага/коричневый
            if r > 140 and g > 100 and b < 90 and r - b > 40 and brightness > 120 and brightness < 200: return True
            return False

        box_pixels = sum(1 for r, g, b in pixels if is_packaging(r, g, b))
        box_ratio = box_pixels / total

        if box_ratio > 0.72:
            result["keep"] = False; result["reason"] = "box_packaging"; return result

        # 4. Низкое разнообразие цветов — монотонное изображение (фон без содержимого)
        import statistics
        r_std = statistics.stdev(r_vals)
        g_std = statistics.stdev(g_vals)
        b_std = statistics.stdev(b_vals)
        color_variance = (r_std + g_std + b_std) / 3
        # Если стандартное отклонение по всем каналам < 18 — очень монотонно
        if color_variance < 18 and avg_brightness > 160:
            result["keep"] = False; result["reason"] = "monotone_background"; return result

        # 5. Почти чисто-белая/серая фотография упаковки с чуть большим разнообразием
        # Проверяем среднее насыщение — упаковочные фото крайне ненасыщенные
        avg_sat = sum(max(p)-min(p) for p in pixels) / total
        if avg_sat < 12 and avg_brightness > 185:
            result["keep"] = False; result["reason"] = "low_saturation_packaging"; return result

    except Exception as e:
        result["reason"] = f"error: {e}"
    return result


@router.post("/filter-photos")
def filter_photos(req: PhotoFilterRequest):
    """Фильтрует список URL фото: убирает коробки, штрих-коды, пустые изображения."""
    if not req.urls:
        return {"status": "success", "results": []}
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as pool:
        results = list(pool.map(_analyze_photo, req.urls[:50]))  # максимум 50 фото за раз
    return {"status": "success", "results": results}

@router.get("/ppm-dataset")
def get_ppm_dataset(
    platform: str = Query("wb", pattern="^(wb|ym|ozon|all)$"),
    db: Session = Depends(get_db)
):
    cache_key = f"ppm-dataset-{platform}"
    cached = get_from_cache(cache_key)
    if cached: return cached
    try:
        df_sys = load_hybrid_data(db, platform)
        df_orders_sys = load_orders_data(db, platform)

        # Исторические данные — только для WB
        df_hist = pd.DataFrame()
        if platform == 'wb':
            try:
                with db.bind.connect() as conn:
                    df_hist = pd.read_sql(text("SELECT article as article, month_date as month_dt, defects as defects, orders as orders, source as source FROM historical_ppm"), conn)
            except: pass
        if not df_hist.empty: df_hist['month_dt'] = pd.to_datetime(df_hist['month_dt'])

        def is_defect(val):
            if val is True: return True
            return str(val).strip().lower() in ['1', '1.0', '+', 'true', 'да', 't']

        tag_cols = [f"cat_{i}" for i in range(1, 14) if f"cat_{i}" in df_sys.columns]
        if tag_cols and not df_sys.empty:
            mask = False
            for col in tag_cols: mask = mask | df_sys[col].apply(is_defect)
            df_sys['is_defect_marked'] = mask
        else:
            if not df_sys.empty: df_sys['is_defect_marked'] = False

        sys_metrics = pd.DataFrame(columns=['article', 'month_dt', 'defects'])

        date_col = 'created_dt'
        if not df_sys.empty and date_col in df_sys.columns:
            df_app_sys = df_sys[df_sys['is_defect_marked'] == True].copy()
            if not df_app_sys.empty:
                df_app_sys['month_dt'] = pd.to_datetime(df_app_sys[date_col]).dt.to_period('M').dt.to_timestamp()
                sys_metrics = df_app_sys.groupby(['supplier_article', 'month_dt']).size().reset_index(name='defects')
                sys_metrics.rename(columns={'supplier_article': 'article'}, inplace=True)

        if df_orders_sys.empty: df_orders_sys = pd.DataFrame(columns=['article', 'month_dt', 'orders'])
        else: df_orders_sys.rename(columns={'supplier_article': 'article', 'pure_orders': 'orders'}, inplace=True)

        # ABC/XYZ: для WB — полный расчёт с еженедельными данными; для ЯМ — только ABC по месяцам
        df_abc = pd.DataFrame()
        df_weekly = pd.DataFrame()
        if platform == 'wb':
            try:
                with db.bind.connect() as conn:
                    df_weekly = pd.read_sql(text("""
                        SELECT TRIM(supplier_article) AS article,
                               DATE_TRUNC('week', CAST(dt AS TIMESTAMP)) AS week_dt,
                               COUNT(*) AS orders
                        FROM wb_orders
                        WHERE cancel_dt IS NULL
                          AND dt >= CURRENT_DATE - INTERVAL '26 weeks'
                        GROUP BY 1, 2
                    """), conn)
            except Exception as e:
                print("XYZ weekly query error:", e)
        elif platform == 'ym':
            try:
                with db.bind.connect() as conn:
                    df_weekly = pd.read_sql(text("""
                        SELECT TRIM(supplier_article) AS article,
                               DATE_TRUNC('week', created_at) AS week_dt,
                               COUNT(*) AS orders
                        FROM ym_orders
                        WHERE status NOT IN ('CANCELLED','PENDING_CANCELLED','PROCESSING_EXPIRED')
                          AND created_at >= CURRENT_DATE - INTERVAL '26 weeks'
                        GROUP BY 1, 2
                    """), conn)
            except Exception as e:
                print("YM XYZ weekly query error:", e)

        df_abc = compute_abc_xyz(df_orders_sys, df_weekly=df_weekly)

        sys_metrics = pd.merge(df_orders_sys, sys_metrics, on=['article', 'month_dt'], how='outer').fillna(0)
        sys_metrics['source'] = 'System'

        df_total = pd.concat([df_hist, sys_metrics], ignore_index=True)
        if not df_abc.empty: df_total = pd.merge(df_total, df_abc, on='article', how='left')
        else: df_total['abc_group'] = 'C'; df_total['xyz_class'] = 'Z'

        df_total['abc_group'] = df_total['abc_group'].fillna('C')
        df_total['xyz_class'] = df_total['xyz_class'].fillna('Z')

        months_ru = {1:'Янв', 2:'Фев', 3:'Мар', 4:'Апр', 5:'Май', 6:'Июн', 7:'Июл', 8:'Авг', 9:'Сен', 10:'Окт', 11:'Ноя', 12:'Дек'}
        df_total['month_str'] = df_total['month_dt'].dt.month.map(months_ru) + " " + df_total['month_dt'].dt.year.astype(str)
        df_total['month_dt'] = df_total['month_dt'].dt.strftime('%Y-%m-%d')

        df_total = df_total.replace({np.nan: None})
        res = {"status": "success", "data": df_total.to_dict(orient="records")}
        save_to_cache(cache_key, res)
        return res
    except Exception as e:
        print("Ошибка де датасета:", e)
        return {"status": "success", "data": []}

def _fetch_compress(url, max_px=1200, quality=82, timeout=6):
    """Скачивает фото, сжимает до max_px по большей стороне, возвращает data-URI.
    При любой ошибке возвращает исходный URL — WeasyPrint скачает сам (graceful fallback)."""
    try:
        from PIL import Image
        r = requests.get(url, timeout=timeout, headers={"User-Agent": "Mozilla/5.0"})
        r.raise_for_status()
        img = Image.open(io.BytesIO(r.content)).convert("RGB")
        if max(img.width, img.height) > max_px:
            img.thumbnail((max_px, max_px), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, "JPEG", quality=quality, optimize=True)
        return f"data:image/jpeg;base64,{base64.b64encode(buf.getvalue()).decode()}"
    except Exception:
        return url  # WeasyPrint сам скачает

def build_photo_grid(urls):
    """Динамическая раскладка фото на странице А4: чем меньше фото — тем крупнее.
    Число колонок зависит от количества, высота ячейки считается так, чтобы ряды
    заполняли страницу по вертикали. object-fit: contain — без обрезки (сохраняет пропорции)."""
    n = len(urls)
    if n == 0:
        return ""
    if n == 1:    cols = 1
    elif n == 2:  cols = 2
    elif n <= 4:  cols = 2
    elif n <= 9:  cols = 3
    else:         cols = 4
    rows = (n + cols - 1) // cols
    PAGE_PHOTO_H = 980  # ~доступная высота под фото на листе A4 (px @96dpi) за вычетом заголовка
    gap = 6
    item_h = int((PAGE_PHOTO_H - rows * gap) / rows)
    item_h = max(120, min(item_h, 920))
    width_pct = 100.0 / cols
    cells = ""
    for u in urls:
        cells += (f'<div class="photo-cell" style="width:{width_pct:.4f}%; height:{item_h + gap}px;">'
                  f'<img style="height:{item_h}px;" src="{u}"></div>')
    return cells

@router.post("/export-claim")
def export_claim_pdf(payload: ClaimExportRequest = Body(...), db: Session = Depends(get_db)):
    init_log_table(db)
    act_number = payload.existing_act_number or "—"
    if payload.is_test:
        # Тестовый акт: генерируем PDF, но НЕ пишем в журнал и не присваиваем номер
        act_number = "TEST"
    elif not payload.is_redownload:
        try:
            pay_json = payload.model_dump_json() if hasattr(payload, 'model_dump_json') else payload.json()
            with db.bind.connect() as conn:
                res = conn.execute(text("""
                    INSERT INTO cx_claim_logs
                      (act_number, sku, supplier, defects_count, period_str, pdf_payload,
                       report_month, abc_group, product_name, invoice_ref, deviation,
                       deviation_desc, deviation_cause, claim_status, initiator, factory_type,
                       who_sent, qty, ra_date, send_status)
                    VALUES
                      ('TEMP-' || gen_random_uuid(), :sku, :sup, :def, :per, CAST(:pay AS jsonb),
                       :report_month, :abc_group, :product_name, :invoice_ref, :deviation,
                       :deviation_desc, :deviation_cause, :claim_status, :initiator, :factory_type,
                       :who_sent, :qty, CURRENT_DATE, 'Не отправлен')
                    RETURNING id
                """), {
                    "sku": payload.sku, "sup": payload.supplier, "def": payload.defects,
                    "per": payload.period, "pay": pay_json,
                    "report_month": payload.report_month,
                    "abc_group": payload.abc_group,
                    "product_name": payload.product_name_ru or payload.name,
                    "invoice_ref": payload.invoice_ref,
                    "deviation": payload.deviation,
                    "deviation_desc": payload.desc_ru,
                    "deviation_cause": payload.cause_ru,
                    "claim_status": payload.claim_status or 'На проверке',
                    "initiator": payload.initiator or 'Wildberries',
                    "factory_type": payload.factory_type or 'завод',
                    "who_sent": payload.who_sent,
                    "qty": payload.defects,
                })
                nid = res.scalar()
                next_num = conn.execute(text("""
                    SELECT COALESCE(MAX(CAST(TRIM(SPLIT_PART(act_number, ',', 1)) AS INTEGER)), 115) + 1
                    FROM cx_claim_logs
                    WHERE act_number ~ '^[0-9]' AND id != :nid
                """), {"nid": nid}).scalar()
                act_number = str(next_num)
                conn.execute(text("UPDATE cx_claim_logs SET act_number = :act WHERE id = :id"), {"act": act_number, "id": nid}); conn.commit()
        except Exception as e: print(f"Ошибка лога: {e}")

    # Высота графика приходит с фронта (зависит от числа инвойсов): мало инвойсов -> выше.
    chart_h = 170
    chart_html = ""
    if payload.chart_data and len(payload.chart_data) > 0 and payload.chart_data[0].get("image"):
        try: chart_h = int(payload.chart_data[0].get("height") or 170)
        except (TypeError, ValueError): chart_h = 170
        # Лимит 500px: при ширине контейнера ~718px и нат. ширине 850px отображаемая высота
        # = declared × (718/850) ≈ declared × 0.845. При 500 → ~422px, вместе с остальным
        # содержимым (шапка + KPI + таблица + заголовок + подписи ≈ 525px) итого ≈ 947px < 1046px A4.
        chart_h = max(120, min(chart_h, 500))
        chart_html = f'<div class="chart-wrapper"><img src="data:image/png;base64,{payload.chart_data[0]["image"]}"></div>'

    # Параллельный prefetch + сжатие всех фотографий (ThreadPoolExecutor, до 8 потоков).
    # WeasyPrint тянул их последовательно во время рендера — так скорость генерации падала
    # пропорционально числу фото. Теперь все фото готовы как data-URI до начала рендера.
    all_urls = [u for c in payload.photo_groups for u in c.get('urls', []) if u]
    if all_urls:
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(8, len(all_urls))) as ex:
            fetched = dict(zip(all_urls, ex.map(_fetch_compress, all_urls)))
    else:
        fetched = {}

    # Динамическая раскладка фото: размер и число колонок зависят от количества фото в группе.
    photos_html = ""
    for c in payload.photo_groups:
        urls = c.get('urls', [])
        if not urls: continue
        resolved = [fetched.get(u, u) for u in urls]
        grid_cells = build_photo_grid(resolved)
        photos_html += f"""
        <div class="photo-page-block">
            <div class="section-title">
                <span><span class="pin">&#9650;</span> {c.get("ru","")}</span>
                <span><span class="pin">&#9650;</span> {c.get("cn","")} <span class="count-badge">{len(urls)} шт / 件</span></span>
            </div>
            <div class="photo-grid">
                {grid_cells}
            </div>
        </div>
        """

    def fmt_cn_date(ru_date_str: str) -> str:
        """Конвертирует ДД.ММ.ГГГГ → YYYY年MM月DD日. Возвращает пустую строку если не удалось."""
        if not ru_date_str or ru_date_str == '—': return ''
        try:
            parts = ru_date_str.strip().split('.')
            if len(parts) == 3:
                return f"{parts[2]}年{parts[1]}月{parts[0]}日"
        except: pass
        return ''

    def fmt_today_cn() -> str:
        return datetime.now().strftime("%Y年%m月%d日")

    # Сортируем инвойсы по дате отгрузки
    def sort_date(inv):
        d = inv.get('date', '') or ''
        if d == '—' or not d: return '9999'
        try:
            parts = d.split('.')
            if len(parts) == 3: return f"{parts[2]}-{parts[1]}-{parts[0]}"
            return d
        except: return '9999'

    sorted_invoices = sorted(payload.invoices_list or [], key=sort_date)
    inv_table_rows = ""
    if sorted_invoices:
        for inv in sorted_invoices:
            inv_name = inv.get('invoice', '—')
            cont = inv.get('container', '—')
            dt = inv.get('date', '—')
            cnt = inv.get('count', '')
            cnt_str = f" <span style='color:#e11d48;font-weight:bold;'>({cnt} шт./件)</span>" if cnt else ""
            dt_cn = fmt_cn_date(dt)
            dt_display = f"{dt}<br><span style='color:#64748b;font-size:9px;'>{dt_cn}</span>" if dt_cn else dt
            inv_table_rows += f"<tr><td style='border:1px solid #cbd5e1;padding:4px 6px;color:#0f172a;'>{inv_name}{cnt_str}</td><td style='border:1px solid #cbd5e1;padding:4px 6px;color:#0f172a;'>{cont}</td><td style='border:1px solid #cbd5e1;padding:4px 6px;color:#0f172a;'>{dt_display}</td></tr>"
    else:
        inv_table_rows = "<tr><td colspan='3' style='border:1px solid #cbd5e1;padding:4px 6px;text-align:center;color:#94a3b8;'>Нет данных</td></tr>"

    inv_html_table = f"""
    <tr>
        <td colspan="2" style="padding: 0; border: none;">
            <table style="width:100%;border-collapse:collapse;text-align:left;font-size:11px;table-layout:fixed;">
                <colgroup>
                    <col style="width:35%;">
                    <col style="width:33%;">
                    <col style="width:32%;">
                </colgroup>
                <tr style="background:#f1f5f9;">
                    <td style="border:1px solid #cbd5e1;padding:4px 6px;font-weight:bold;color:#1e3a8a;font-size:10px;">Инвойс / 发票号码</td>
                    <td style="border:1px solid #cbd5e1;padding:4px 6px;font-weight:bold;color:#1e3a8a;font-size:10px;">Контейнер / 集装箱号</td>
                    <td style="border:1px solid #cbd5e1;padding:4px 6px;font-weight:bold;color:#1e3a8a;font-size:10px;">Дата отгрузки / 装运日期</td>
                </tr>
                {inv_table_rows}
            </table>
        </td>
    </tr>
    """

    html_content = f"""
    <!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8"><style>
        @font-face {{ font-family: 'CJKFallback'; src: url('file:///usr/share/fonts/truetype/wqy/wqy-zenhei.ttc'); }}
        .pin {{ display: inline-block; background: #facc15; color: #1e3a8a; font-weight: 900; font-size: 9px; padding: 1px 4px; border-radius: 3px; margin-right: 4px; vertical-align: middle; }}
        @page {{ size: A4 portrait; margin: 10mm; @bottom-right {{ content: "Стр. " counter(page) " из " counter(pages); font-family: 'Helvetica', sans-serif; font-size: 8pt; color: #94a3b8; }} }}
        body {{ font-family: 'Helvetica Neue', Helvetica, Arial, 'CJKFallback', sans-serif; color: #0f172a; margin: 0; padding: 0; line-height: 1.2; }}
        .first-page-layout {{ page-break-after: always; }}
        .header {{ border-bottom: 2px solid #1e3a8a; padding-bottom: 5px; margin-bottom: 10px; display: flex; justify-content: space-between; }}
        .header table {{ width: 100%; border: none; }} .header td {{ border: none; padding: 0; vertical-align: middle; }}
        .main-title {{ font-size: 16px; font-weight: 900; color: #1e3a8a; text-transform: uppercase; }}
        .sub-title {{ font-size: 11px; font-weight: bold; margin-top: 2px; }}
        .doc-badge {{ text-align: right; }} .doc-badge span {{ background: #1e3a8a; color: white; padding: 4px 10px; font-size: 11px; font-weight: bold; border-radius: 4px; }}
        .kpi-container {{ display: table; width: 100%; margin-bottom: 10px; table-layout: fixed; border-spacing: 6px 0; }}
        .kpi-card {{ display: table-cell; background: #f1f5f9; border-radius: 4px; padding: 6px; text-align: center; border: 1px solid #e2e8f0; }}
        .kpi-card.danger {{ background: #fff1f2; border-color: #fecdd3; }}
        .kpi-title {{ font-size: 10px; font-weight: bold; }} .kpi-value {{ font-size: 16px; font-weight: 900; margin-top: 2px; }} .kpi-value.danger {{ color: #e11d48; }}
        .info-table {{ width: 100%; border-collapse: collapse; margin-bottom: 5px; }}
        .info-table td {{ border: 1px solid #cbd5e1; padding: 4px 6px; vertical-align: top; }}
        .cell-label {{ background: #f8fafc; width: 35%; }} .cell-value {{ width: 65%; }}
        .text-pair {{ font-size: 11px; font-weight: bold; margin-bottom: 2px; }} .text-val {{ font-size: 12px; font-weight: bold; color: #1e3a8a; }}
        .desc-box {{ background: #fff1f2; padding: 6px; border-radius: 4px; border-left: 3px solid #e11d48; }} .desc-text {{ font-size: 11px; font-weight: bold; color: #9f1239; }}
        .section-title {{ background: #1e3a8a; color: white; padding: 4px 8px; font-size: 11px; font-weight: bold; text-transform: uppercase; margin-bottom: 5px; border-radius: 4px; display: flex; justify-content: space-between; page-break-inside: avoid; }}
        .chart-wrapper {{ text-align: center; background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 6px; padding: 2px; margin-bottom: 5px; page-break-inside: avoid; }} .chart-wrapper img {{ width: 100%; max-width: 100%; height: auto; max-height: {chart_h}px; object-fit: contain; }}
        .signatures {{ width: 100%; margin-top: 5px; border-top: 2px solid #e2e8f0; padding-top: 5px; }} .signatures td {{ width: 33%; text-align: center; border: none; padding: 0 10px; }}
        .sig-title {{ font-size: 10px; font-weight: bold; }} .sig-line {{ border-bottom: 1px solid #0f172a; margin: 3px 20px 0 20px; height: 12px; }}
        
        .photo-page-block {{ page-break-before: always; margin-top: 10px; }}
        .photo-grid {{ display: flex; flex-wrap: wrap; align-content: flex-start; justify-content: center; margin-top: 8px; }}
        .photo-cell {{ box-sizing: border-box; padding: 3px; display: flex; align-items: center; justify-content: center; }}
        .photo-cell img {{ width: 100%; max-width: 100%; object-fit: contain; background: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 4px; display: block; }}
    </style></head><body>
        <div class="first-page-layout">
            <div class="header"><table><tr><td><div class="main-title">Рекламационный акт качества</div><div class="sub-title">ПРЕТЕНЗИЯ ПО ФАБРИЧНОМУ БРАКУ / 质量投诉报告</div></td><td class="doc-badge"><span>Акт № {act_number}</span></td></tr></table></div>
            <div class="kpi-container">
                <div class="kpi-card"><div class="kpi-title">Артикул / 产品编号</div><div class="kpi-value sku-highlight" style="color:#1e3a8a;">{payload.sku}</div></div>
                <div class="kpi-card danger"><div class="kpi-title">Кол-во брака / 不合格数量</div><div class="kpi-value danger">{payload.defects} шт.</div></div>
                <div class="kpi-card danger"><div class="kpi-title">Процент брака / 不合格率</div><div class="kpi-value danger">{payload.ppm_pct}%</div></div>
            </div>
            <table class="info-table">
                <tr><td class="cell-label"><div class="text-pair">Дата / 编制日期</div></td><td class="cell-value"><div class="text-val">{datetime.now().strftime("%d.%m.%Y")}</div><div class="text-pair" style="color:#475569;">{fmt_today_cn()}</div></td></tr>
                <tr><td class="cell-label"><div class="text-pair">Период выборки / 抽样期</div></td><td class="cell-value"><div class="text-val">{payload.period or "Не указан"}</div></td></tr>
                <tr><td class="cell-label"><div class="text-pair">Завод / 工厂名称</div></td><td class="cell-value"><div class="text-val">{payload.supplier}</div></td></tr>
                <tr><td class="cell-label"><div class="text-pair">Наименование товара / 产品名称</div></td><td class="cell-value"><div class="text-val">{payload.name or 'Наименование товара'}</div><div class="text-val">{payload.name_cn}</div></td></tr>
                
                {inv_html_table}

                <tr><td class="cell-label"><div class="text-pair" style="color: #b91c1c;">Описание брака / 不符合项描述</div></td><td class="cell-value" style="padding: 0;"><div class="desc-box"><div class="desc-text">{str(payload.desc_ru).replace(chr(10), '<br>')}</div><div class="desc-text">{str(payload.desc_cn).replace(chr(10), '<br>')}</div></div></td></tr>
                <tr><td class="cell-label"><div class="text-pair">Ключевая причина / 初步原因分析</div></td><td class="cell-value"><div class="text-pair" style="color: #475569;">{str(payload.cause_ru).replace(chr(10), '<br>')}</div><div class="text-pair" style="color: #475569;">{str(payload.cause_cn).replace(chr(10), '<br>')}</div></td></tr>
            </table>
            <div class="section-title"><span>Аналитика отклонений</span><span>异常动态</span></div> {chart_html}
            <table class="signatures"><tr><td><div class="sig-title">Дата отправки уведомления / 通知发送日期</div><div class="text-val" style="margin-top:5px;">{datetime.now().strftime("%d.%m.%Y")}</div><div style="font-size:10px;color:#475569;margin-top:2px;">{fmt_today_cn()}</div></td><td><div class="sig-title">Дата ответа производителя / 厂方回复日期</div><div class="sig-line"></div></td><td><div class="sig-title">Согласие с предварительной причиной / 对初步原因的确认</div><div class="sig-line"></div></td></tr></table>
        </div>
        {photos_html}
    </body></html>
    """
    try:
        pdf_bytes = HTML(string=html_content).write_pdf(font_config=FontConfiguration())
        filename = f"{act_number}.pdf"
        ascii_fallback = filename.encode("latin-1", "replace").decode("latin-1")
        disposition = f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(filename)}"
        
        # Возвращаем файл и прокидываем чистый номер акта в кастомном заголовке
        return Response(content=pdf_bytes, media_type="application/pdf", headers={
            "Content-Disposition": disposition,
            "X-Act-Number": str(act_number)  # <- Добавлено для синхронизации с фронтендом
        })
    except Exception as e: return HTMLResponse(content=f"<h1>Ошибка генерации PDF: {e}</h1>", status_code=500)


@router.get("/claims-metadata")
def get_claims_metadata(db: Session = Depends(get_db)):
    """Уникальные заводы, инвойсы и контейнеры для автодополнения в ручных обращениях."""
    result = {"factories": [], "invoices": [], "containers": []}
    try:
        with db.bind.connect() as conn:
            try:
                rows = conn.execute(text(
                    "SELECT DISTINCT factory_name FROM wb_factories "
                    "WHERE factory_name IS NOT NULL AND factory_name != '' ORDER BY factory_name"
                )).fetchall()
                result["factories"] = [r[0] for r in rows]
            except Exception: pass
            try:
                rows = conn.execute(text(
                    "SELECT DISTINCT invoice_num FROM wb_invoices "
                    "WHERE invoice_num IS NOT NULL AND invoice_num NOT IN ('', '0') "
                    "ORDER BY invoice_num DESC LIMIT 500"
                )).fetchall()
                result["invoices"] = [r[0] for r in rows]
            except Exception: pass
            try:
                rows = conn.execute(text(
                    "SELECT DISTINCT container_num FROM view_wb_claims_enriched "
                    "WHERE container_num IS NOT NULL AND container_num NOT IN ('', 'None') "
                    "ORDER BY container_num DESC LIMIT 500"
                )).fetchall()
                result["containers"] = [r[0] for r in rows]
            except Exception: pass
    except Exception: pass
    return {"status": "success", "data": result}


@router.get("/correction-notifications")
def get_correction_notifications(db: Session = Depends(get_db)):
    """Возвращает акты, у которых инвойс с исправлением уже появился в системе."""
    try:
        with db.bind.connect() as conn:
            rows = conn.execute(text("""
                SELECT cl.id, cl.act_number, cl.sku, cl.supplier, cl.correction_invoice, cl.estimated_improvement_date
                FROM cx_claim_logs cl
                WHERE cl.correction_invoice IS NOT NULL AND cl.correction_invoice != ''
                  AND cl.status = 'Активен'
                  AND EXISTS (
                      SELECT 1 FROM wb_invoices wi
                      WHERE LOWER(TRIM(wi.invoice_num)) = LOWER(TRIM(cl.correction_invoice))
                  )
            """)).mappings().all()
            return {"status": "success", "data": [dict(r) for r in rows]}
    except Exception as e:
        return {"status": "success", "data": []}


@router.get("/ppm-report-excel")
def generate_ppm_report_excel(
    start_date: str = Query(...),
    end_date: str = Query(...),
    factories: Optional[str] = Query(None),
    period_type: str = Query("custom"),
    include_containers: bool = Query(False),
    include_costs: bool = Query(False),
    platform: str = Query("wb", pattern="^(wb|ym|ozon|all)$"),
    db: Session = Depends(get_db)
):
    """Генерация отчёта PPM по заводу в формате Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import timedelta

        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        duration = (end_dt - start_dt).days + 1

        factory_list = [f.strip() for f in (factories or '').split(',') if f.strip()]

        def week_label(s, e):
            wn_s = s.isocalendar()[1]; wn_e = e.isocalendar()[1]
            if wn_s == wn_e:
                return f"н{wn_s}  /  第{wn_s}周"
            else:
                return f"н{wn_s}–н{wn_e}  /  第{wn_s}–{wn_e}周"

        # 3 предыдущих периода той же длины: [n-3, n-2, n-1]
        prev_periods = [
            (start_dt - timedelta(days=duration*(n+1)), end_dt - timedelta(days=duration*(n+1)))
            for n in reversed(range(3))
        ]

        df_all = load_hybrid_data(db, platform)
        df_ord_raw = load_orders_data(db, platform)
        if not df_ord_raw.empty:
            df_ord_raw['month_dt'] = pd.to_datetime(df_ord_raw['month_dt'])
            oc = 'pure_orders' if 'pure_orders' in df_ord_raw.columns else 'orders'
            if oc != 'orders':
                df_ord_raw = df_ord_raw.rename(columns={oc: 'orders'})

        # Себестоимость из wb_cogs
        exact_cogs, latest_cogs = {}, {}
        if include_costs:
            try:
                with db.bind.connect() as conn:
                    for r in conn.execute(text(
                        "SELECT supplier_article, invoice_num, cost_value FROM wb_cogs "
                        "WHERE cost_value IS NOT NULL AND cost_value > 0"
                    )).fetchall():
                        sk, inv, val = str(r[0] or '').strip(), str(r[1] or '').strip(), float(r[2] or 0)
                        if sk and inv: exact_cogs[(sk, inv)] = val
                        if sk: latest_cogs[sk] = val
            except Exception as ex:
                print(f"wb_cogs error: {ex}")

        validVals = {'1', '1.0', '+', 'true', 'да', 't'}

        PROBLEM_GROUPS = [
            ("Не хватает комплект.\n配件缺失", [1, 2]),
            ("Повреждения деталей\n零件损坏", [4, 5]),
            ("Качество / хлипкость\n质量/强度问题", [7, 9]),
            ("Прочие дефекты\n其他缺陷", None),
        ]

        def claims_for(s, e):
            if df_all.empty or 'created_dt' not in df_all.columns: return pd.DataFrame()
            return df_all[(df_all['created_dt'] >= s.strftime('%Y-%m-%d')) & (df_all['created_dt'] <= e.strftime('%Y-%m-%d'))]

        def orders_for(s, e):
            if df_ord_raw.empty: return {}
            m1 = pd.Timestamp(s.year, s.month, 1)
            df_p = df_ord_raw[(df_ord_raw['month_dt'] >= m1) & (df_ord_raw['month_dt'] <= pd.Timestamp(e))]
            return df_p.groupby('supplier_article')['orders'].sum().to_dict() if not df_p.empty else {}

        df_cur = claims_for(start_dt, end_dt)
        ord_cur = orders_for(start_dt, end_dt)
        prev_claims = [claims_for(ps, pe) for ps, pe in prev_periods]
        prev_orders = [orders_for(ps, pe) for ps, pe in prev_periods]

        # Строим строки данных
        rows_data = []
        if not df_cur.empty:
            df_cur = df_cur.copy()
            df_cur['factory_name'] = df_cur['factory_name'].fillna('Уточняется')
            df_cur['supplier_article'] = df_cur['supplier_article'].astype(str).str.strip()
            for (fac, sku), grp in df_cur.groupby(['factory_name', 'supplier_article']):
                fac = str(fac or 'Уточняется')
                sku_s = str(sku).strip()
                if factory_list and fac not in factory_list:
                    continue
                total = len(grp)
                orders = int(ord_cur.get(sku_s, 0))
                pct = round(total / orders * 100, 2) if orders > 0 else 0.0

                # Группы проблем (без двойного счёта)
                g_counts = []
                assigned = set()
                for _, ids in PROBLEM_GROUPS[:-1]:
                    cnt = 0
                    for idx, r in grp.iterrows():
                        if idx not in assigned and any(str(r.get(f'cat_{i}', '') or '').strip().lower() in validVals for i in ids):
                            cnt += 1; assigned.add(idx)
                    g_counts.append(cnt)
                g_counts.append(max(0, total - sum(g_counts)))

                # % за прошлые периоды
                pp_pcts = []
                for df_p, ord_p in zip(prev_claims, prev_orders):
                    if df_p.empty: pp_pcts.append(0.0); continue
                    mask = (df_p['factory_name'].fillna('Уточняется') == fac) & (df_p['supplier_article'].astype(str).str.strip() == sku_s)
                    p_ord = int(ord_p.get(sku_s, 0))
                    pp_pcts.append(round(int(mask.sum()) / p_ord * 100, 2) if p_ord > 0 else 0.0)

                delta = round(pct - pp_pcts[-1], 2)

                containers = ''
                if include_containers:
                    containers = '; '.join(sorted({str(r.get('container_num','') or '') for _, r in grp.iterrows() if str(r.get('container_num','') or '') not in ('', 'nan', 'None', 'Не указан')}))

                cost_loss = 0.0
                if include_costs:
                    for _, r in grp.iterrows():
                        inv = str(r.get('invoice_num', '') or '').strip()
                        cost_loss += exact_cogs.get((sku_s, inv), latest_cogs.get(sku_s, 0.0))

                rows_data.append({'factory': fac, 'sku': sku_s, 'orders': orders, 'defects': total, 'pct': pct,
                                  'delta': delta, 'pp_pcts': pp_pcts, 'g_counts': g_counts,
                                  'containers': containers, 'cost_loss': round(cost_loss, 2)})

        rows_data.sort(key=lambda r: (r['factory'], -r['defects']))

        # === Excel ===
        wb_xl = openpyxl.Workbook()
        ws = wb_xl.active
        ws.title = "PPM Отчёт"

        thin = Side(style='thin', color='CBD5E1')
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
        HDR_FILL = PatternFill("solid", start_color="1E3A8A")
        SUB_FILL = PatternFill("solid", start_color="DBEAFE")
        TOTAL_FILL = PatternFill("solid", start_color="E2E8F0")
        ALT_FILL = PatternFill("solid", start_color="F8FAFC")
        RED_FILL = PatternFill("solid", start_color="FEE2E2")

        def fnt(bold=False, red=False, white=False, blue=False, size=10):
            color = "B91C1C" if red else ("FFFFFF" if white else ("1E3A8A" if blue else "000000"))
            return Font(name="Arial", size=size, bold=bold, color=color)

        HDR_ROW, SUB_ROW, DATA_START = 3, 4, 5
        period_str = (f"{start_dt.strftime('%d.%m.%Y')} – {end_dt.strftime('%d.%m.%Y')}  "
                      f"({start_dt.strftime('%Y年%m月%d日')} – {end_dt.strftime('%Y年%m月%d日')})")

        CONT_COL = 14 if include_containers else None
        COST_COL = (15 if include_containers else 14) if include_costs else None
        total_cols = 13 + (1 if include_containers else 0) + (1 if include_costs else 0)

        # Информационные строки
        ws['A1'] = "Период  /  时间段"; ws['A1'].font = fnt(bold=True)
        ws['B1'] = period_str; ws['B1'].font = fnt(bold=True, blue=True)
        ws.merge_cells('B1:F1')
        ws['A2'] = "Номер недели  /  周数"; ws['A2'].font = fnt(bold=True)
        ws['B2'] = week_label(start_dt, end_dt); ws['B2'].font = fnt(bold=True, blue=True)
        ws.merge_cells('B2:F2')

        # Хелпер: заголовочная ячейка с опциональным merge
        def make_hdr(row, col, val, r2=None, c2=None):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = HDR_FILL; cell.font = fnt(bold=True, white=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if r2 or c2:
                ws.merge_cells(start_row=row, start_column=col, end_row=r2 or row, end_column=c2 or col)

        def make_sub(row, col, val):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = SUB_FILL; cell.font = fnt(bold=True, blue=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Основные заголовки: колонки 1–6 занимают обе строки HDR_ROW и SUB_ROW
        for cidx, lbl in [
            (1, "Завод\n工厂"),
            (2, "Артикул\n产品编号"),
            (3, "Продано, шт\n销售量（件）"),
            (4, "Брак, шт\n缺陷数（件）"),
            (5, "% брака\n缺陷率"),
            (6, "Δ к прошл. периоду\n与上期对比△"),
        ]:
            make_hdr(HDR_ROW, cidx, lbl, r2=SUB_ROW)

        # "% брака за прошлые периоды" — занимает cols 7-9 в HDR_ROW; sub-headers в SUB_ROW
        make_hdr(HDR_ROW, 7, "% брака за прошлые периоды\n过去各期缺陷率", c2=9)
        for i, (ps, pe) in enumerate(prev_periods):
            make_sub(SUB_ROW, 7+i, week_label(ps, pe))

        # "Распределение по типам проблем" — cols 10-13
        make_hdr(HDR_ROW, 10, "Распределение по типам проблем (шт / %)\n问题类型分布（件/%）", c2=13)
        for i, (nm, cn) in enumerate(PROBLEM_GROUPS):
            make_sub(SUB_ROW, 10+i, nm)

        if CONT_COL: make_hdr(HDR_ROW, CONT_COL, "Контейнеры\n集装箱", r2=SUB_ROW)
        if COST_COL: make_hdr(HDR_ROW, COST_COL, "Издержки, руб.\n损失（руб）", r2=SUB_ROW)

        for rv in [HDR_ROW, SUB_ROW]:
            for c in range(1, total_cols + 1):
                ws.cell(row=rv, column=c).border = BORDER
        ws.row_dimensions[HDR_ROW].height = 42
        ws.row_dimensions[SUB_ROW].height = 28

        # Строки данных
        for i, row in enumerate(rows_data):
            r = DATA_START + i
            is_red = row['pct'] > 1.0
            alt = ALT_FILL if i % 2 == 0 else None

            def dc(col, val, fmt=None, bold=False, center=True):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = BORDER
                if col == 5 and is_red:
                    cell.fill = RED_FILL; cell.font = fnt(bold=True, red=True)
                else:
                    if alt: cell.fill = alt
                    cell.font = fnt(bold=bold)
                if fmt: cell.number_format = fmt
                cell.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=(col in (1, 2)))
                return cell

            dc(1, row['factory'], center=False)
            dc(2, row['sku'], center=False)
            dc(3, row['orders'])
            dc(4, row['defects'])
            dc(5, row['pct'], fmt='0.00"%"')

            delta_cell = dc(6, row['delta'], fmt='0.00"%"')
            if row['delta'] > 0: delta_cell.font = fnt(red=True)
            elif row['delta'] < 0: delta_cell.font = Font(name="Arial", size=10, color="15803D")

            for j, pp in enumerate(row['pp_pcts']):
                dc(7+j, pp, fmt='0.00"%"')

            tot_def = row['defects']
            for j, cnt in enumerate(row['g_counts']):
                share = round(cnt / tot_def * 100, 1) if tot_def > 0 else 0.0
                val_str = f"{cnt} шт/件 / {share}%" if cnt > 0 else "—"
                cell = ws.cell(row=r, column=10+j, value=val_str)
                cell.border = BORDER
                if alt: cell.fill = alt
                cell.font = fnt()
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            if CONT_COL:
                cell = ws.cell(row=r, column=CONT_COL, value=row['containers'])
                cell.border = BORDER; cell.font = Font(name="Arial", size=9)
                if alt: cell.fill = alt
                cell.alignment = Alignment(wrap_text=True, vertical='center')

            if COST_COL:
                cell = ws.cell(row=r, column=COST_COL, value=row['cost_loss'])
                cell.border = BORDER; cell.font = fnt()
                if alt: cell.fill = alt
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')

        # Объединяем ячейки одинакового завода
        i = 0
        while i < len(rows_data):
            j = i + 1
            while j < len(rows_data) and rows_data[j]['factory'] == rows_data[i]['factory']:
                j += 1
            if j - i > 1:
                r1, r2_ = DATA_START + i, DATA_START + j - 1
                ws.merge_cells(start_row=r1, start_column=1, end_row=r2_, end_column=1)
                cell = ws.cell(row=r1, column=1)
                cell.value = rows_data[i]['factory']
                cell.font = fnt(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = BORDER
            i = j

        # Итоговая строка
        if rows_data:
            tr = DATA_START + len(rows_data)
            for c in range(1, total_cols + 1):
                ws.cell(row=tr, column=c).fill = TOTAL_FILL
                ws.cell(row=tr, column=c).border = BORDER
            ws.cell(row=tr, column=1, value="ИТОГО  /  合计")
            ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
            ws.cell(row=tr, column=1).font = fnt(bold=True)
            ws.cell(row=tr, column=1).alignment = Alignment(horizontal='center', vertical='center')

            tot_ord = sum(r['orders'] for r in rows_data)
            tot_def = sum(r['defects'] for r in rows_data)
            for col, val, fmt in [(3,tot_ord,None),(4,tot_def,None),(5,round(tot_def/tot_ord*100,2) if tot_ord else 0,'0.00"%"')]:
                cell = ws.cell(row=tr, column=col, value=val)
                cell.font = fnt(bold=True); cell.alignment = Alignment(horizontal='center', vertical='center')
                if fmt: cell.number_format = fmt
            for j in range(4):
                ws.cell(row=tr, column=10+j, value=sum(r['g_counts'][j] for r in rows_data)).font = fnt(bold=True)
            if COST_COL:
                cell = ws.cell(row=tr, column=COST_COL, value=round(sum(r['cost_loss'] for r in rows_data), 2))
                cell.font = fnt(bold=True); cell.number_format = '#,##0.00'

        # Ширины колонок
        for cidx, w in {1:22,2:18,3:12,4:10,5:10,6:13,7:10,8:10,9:10,10:18,11:18,12:18,13:16}.items():
            ws.column_dimensions[get_column_letter(cidx)].width = w
        if CONT_COL: ws.column_dimensions[get_column_letter(CONT_COL)].width = 35
        if COST_COL: ws.column_dimensions[get_column_letter(COST_COL)].width = 15

        ws.freeze_panes = 'C5'

        buf = io.BytesIO()
        wb_xl.save(buf)
        buf.seek(0)

        fn = f"PPM_Отчет_{start_date}_{end_date}.xlsx"
        ascii_fn = fn.encode("latin-1", "replace").decode("latin-1")
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=\"{ascii_fn}\"; filename*=UTF-8''{quote(fn)}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка генерации отчёта: {str(e)}")